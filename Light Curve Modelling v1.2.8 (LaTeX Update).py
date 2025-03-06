import tkinter as tk
import tkinter.filedialog as tkfile
import tkinter.ttk as ttk
from PIL import Image, ImageTk

from matplotlib import pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)

import math
import numpy as np

from scipy.optimize import curve_fit

import openpyxl

import string

import webbrowser

import time
import sys

import pylatex #All the needed latex stuff

import pylatex.utils #Bold, italic etc.

import os

main_bg="#0B0B0B"
secondary_bg="#252525"
tetriary_bg="#3A3A3A"
title_bg="#FF7A37"

button_bg="#f0e130"

text_color_primary="white"
text_color_secondary="black"

color_frame_bg="#0C0B0B"

# splash_window=tk.Tk() #To make it centered
# splash_window.resizable(width=False, height=False)
# splash_window.overrideredirect(True)
# splash_window.configure(bg=main_bg)

# screen_width=splash_window.winfo_screenwidth()
# screen_height=splash_window.winfo_screenheight()

# x_cordinate = int((screen_width/2) - (700/2))
# y_cordinate = int((screen_height/2) - (450/2))

# splash_window.geometry("{}x{}+{}+{}".format(700, 450, x_cordinate, y_cordinate))

# splash_frame=tk.Frame(master=splash_window, width=700, height=450, padx=0, pady=0)
# splash_frame.pack()
# splash_frame.place(anchor='center', relx=0.5, rely=0.5)

# image=Image.open(r"splashscreen.png")
# splash_image=ImageTk.PhotoImage(image.resize((700, 450), Image.ANTIALIAS))
# splash_label=tk.Label(master=splash_frame, image=splash_image)
# splash_label.image=splash_image
# splash_label.pack()
    
# def main_window_init():
#     splash_window.destroy()
    
main_window=tk.Tk()
main_window.title("Light Curve Modelling v1.2.8")
main_window.resizable(width=False, height=False)
main_window.iconbitmap(r'icon_lcf.ico')
main_window.configure(bg=main_bg)

frame_top_left=tk.Frame(master=main_window, bg=secondary_bg, padx=10, pady=5)
frame_middle_left=tk.Frame(master=main_window, bg=secondary_bg, padx=10, pady=5)
frame_bottom_left=tk.Frame(master=main_window, bg=secondary_bg, padx=10, pady=5)

frame_top_right=tk.Frame(master=main_window, bg=secondary_bg, padx=10, pady=5)
frame_middle_right=tk.Frame(master=main_window, bg=secondary_bg, padx=10, pady=5)
frame_bottom_right=tk.Frame(master=main_window, bg=secondary_bg, padx=10, pady=5)

frame_top_left.grid(row=0, column=0, sticky="n", padx=10, pady=10)
frame_top_right.grid(row=0, column=1, rowspan=2, padx=10, pady=10)
frame_middle_left.grid(row=1, column=0, padx=10, pady=10)
frame_middle_right.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=20, pady=10)
frame_bottom_left.grid(row=3, column=0, pady=5)
frame_bottom_right.grid(row=3, column=1, pady=5)

#Light Curve Data
lc_time, lc_flux = [], []
Epoch_List, Delta_phase_List=[], []

Mean_Anomaly_values=[]
Eccentric_Anomaly_values=[]
True_Anomaly_values=[]
Rho_values=[]
Delta_values=[]

T_T0_values=[]
Epoch_deg=[]

planet_parameter_names_list=["T0", "P", "e", "\u03C9", "i", "rp"]
star_parameter_names_list=["r⋆", "ua", "ub"]
alt_star_parameter_names_list=["T eff", "log g", "Metallicity", "Micro Turbulance"]
model_parameter_names_list=["T0", "P", "e", "\u03C9", "i", "rp", "r⋆", "ua", "ub"]
model_parameter_round_list=[4,6,3,3,3,4,4,3,3]

def model_parameters_update():
    model_parameter_fit_value_list=[]
    model_parameter_error_value_list=[]

def lcexport():
    lcexportplanet_parameter_values_list=[]
    lcexportstar_parameter_values_list=[]

planet_parameter_values_list=[]
star_parameter_values_list=[]
alt_star_parameter_values_list=[]
model_parameter_values_list=[]

#Radial Velocity Data
obs_time_list, v_radial_list, v_error_list, weights = [], [], [], []

rvmodel_parameter_names_list=["v\u03B3", "K", "T0", "P", "e", "\u03C9"]
rvstar_parameter_names_list=["v\u03B3", "K", "r⋆", "M⋆"]
rvplanet_parameter_names_list=["T0", "P", "e", "\u03C9", "i", "rp"]
rvmodel_parameter_round_list=[4,4,4,6,3,3]

def rvmodel_parameters_update():
    rvmodel_parameter_fit_value_list=[]
    rvmodel_parameter_error_value_list=[]
    
def rvexport():
    exportplanet_parameter_values_list=[]
    exportstar_parameter_values_list=[]

rvmodel_parameter_values_list=[]
rvstar_parameter_values_list=[]
rvplanet_parameter_values_list=[]

rv_result_parameter_names_list=["M⋆", "M (Planet)", "a (Planet)", "R (Planet)", "g (Planet)", "log g (Planet)", "R⋆", "g⋆", "log g⋆"]
rv_result_parameter_unit_list=["M\u2609", "Mj", "AU", "Rj", "cm/sn^2", "---", "R\u2609", "cm/sn^2", "---"]
rv_mass_star=[]
rv_constant_values=[1988500e24, 1.89813e27, 695508, 71492, 6.67259e-8] #Sun's Mass (kg), Jupiter's Mass (kg), Sun's Radius (km), Jupiter's Radius (km), Gravitational Constant (cgs)
rv_result_parameter_values_list=[]

all_planet_parameter_names_list=["Inferior Conjunction (Light Curve)", "Orbital Period", "Eccentricity", "Argument of Periapsis", "Inclination", "Planet Fractional Radius", "Orbital Radius", "Planet Radius", "Planet Mass", "Planet Surface Gravity", "Logarithm of Planet Surface Gravity"]
all_planet_parameter_notation_list=["T0", "P", "e", "\u03C9", "i", "rp", "a", "Rp", "Mp", "gp", "log gp"]
all_planet_parameter_unit_list=["Days", "Days", "---", "Radians", "Radians", "---", "AU", "Rj (Jupiter Radius)", "Mj (Jupiter Mass)", "cm/sn^2", "---"]
all_planet_parameter_values_list=[]

all_star_parameter_names_list=["Star Fractional Radius", "Limb Darkening Coefficient 'a'", "Limb Darkening Coefficient 'b'", "Star Radius", "Star Mass", "Star Surface Gravity", "Logarithm of Star Surface Gravity"]
all_star_parameter_notation_list=["r⋆", "ua", "ub", "R⋆", "M⋆", "g⋆", "log g⋆"]
all_star_parameter_unit_list=["---", "---", "---", "R\u2609 (Sun Radius)", "M\u2609 (Sun Mass)", "cm/sn^2", "---"]
all_star_parameter_values_list=[]

latex_planet_parameter_names=[pylatex.utils.NoEscape("$T_{0}$"), pylatex.utils.NoEscape("$P$"), pylatex.utils.NoEscape("$e$"), pylatex.utils.NoEscape("$\omega$"), pylatex.utils.NoEscape("$i$"), pylatex.utils.NoEscape("$r_{p}$"), pylatex.utils.NoEscape("$a$"), pylatex.utils.NoEscape("$R_{p}$"), pylatex.utils.NoEscape("$M_{p}$"), pylatex.utils.NoEscape("$g_{p}$"), pylatex.utils.NoEscape("$\log g_{p}$")]
latex_planet_parameter_values=["---", "---", "---", "---", "---", "---", "---", "---", "---", "---", "---"]
latex_planet_parameter_units=["Days", "Days", "---", "Radians", "Radians", "---", "AU", pylatex.utils.NoEscape("$R_{J}$"), pylatex.utils.NoEscape("$M_{J}$"), pylatex.utils.NoEscape("$cm/sn^2$"), "---"]

latex_star_parameter_names=[pylatex.utils.NoEscape("$r_{\star}$"), pylatex.utils.NoEscape("$u_{a}$"), pylatex.utils.NoEscape("$u_{b}$"), pylatex.utils.NoEscape("$R_{\star}$"), pylatex.utils.NoEscape("$M_{\star}$"), pylatex.utils.NoEscape("$g_{\star}$"), pylatex.utils.NoEscape("$\log g_{\star}$")]
latex_star_parameter_values=["---", "---", "---", "---", "---", "---", "---"]
latex_star_parameter_units=["---", "---", "---", pylatex.utils.NoEscape("R{\(\odot\)}"), pylatex.utils.NoEscape("M{\(\odot\)}"), pylatex.utils.NoEscape("$cm/sn^2$"), "---"]

#Button Configurations + Functions
#--> Window for Light Curve Data Showcase + Data Collection
# --> Open Data Button Configuration
def lcdataentry_window():
    if dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        
        button_bg="#f0e130"
        
        text_color_primary="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
    if dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        
        button_bg="#C9C6C6"
        
        text_color_primary="black"
        text_color_secondary="black"
            
        color_frame_bg="#0C0B0B"

    lcdata_window=tk.Toplevel(master=main_window)
    lcdata_window.title("Light Curve Data Entry")
    lcdata_window.iconbitmap(r'icon_lcf.ico')
    lcdata_window.configure(bg=main_bg)
    lcdata_window.wm_transient(main_window)
    lcdata_window.resizable(width=False, height=False)
    
    datawindow=tk.Text(master=lcdata_window, bg=tetriary_bg, fg=text_color_primary)
    datawindow.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

    if len(lc_time)!=0 or len(lc_flux)!=0:
        for i in range(len(lc_time)):
            datawindow.insert(tk.END, lc_time[i])
            datawindow.insert(tk.END, f"	{lc_flux[i]}\n")
    
    def Open_Data():
        data_path=tkfile.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        
        if not data_path:
            return
        
        datawindow.configure(state="normal")
        datawindow.delete("1.0", tk.END)
        lc_time.clear()
        lc_flux.clear()
        Delta_phase_List.clear()
    
        with open(data_path, mode="r", encoding="utf-8") as user_input_file:
            lc_input=user_input_file.read()
            data_input=datawindow.insert(tk.END, lc_input)
        
            for line in open(data_path, mode="r"):
                values = [float(s) for s in line.split()]
                lc_time.append(float(values[0]))
                lc_flux.append(float(values[1]))
    
        lcdata_window.title(f"Light Curve Data Entry - {data_path}")
        datawindow.configure(state="disabled")
    
    def Clear_Data():
        lc_time.clear()
        lc_flux.clear()
        Delta_phase_List.clear()
    
        datawindow.configure(state="normal")
        datawindow.delete("1.0", tk.END)
        lcdata_window.title("Light Curve Data Entry")
    
    def Apply_Button():
        if datawindow.get("1.0", tk.END)=="\n":
            putlcerr=tk.messagebox.showerror("Error", "There is no data applied!")
    
        else:
            lcsuccess=tk.messagebox.showinfo("Successful", "Data has been applied!")
            lcdata_window.withdraw()

    dataopen=tk.Button(master=lcdata_window, text="Open Data", command=Open_Data, bg="green").grid(row=1, column=1, pady=5, sticky="ew")
    dataclear=tk.Button(master=lcdata_window, text="Clear Data", command=Clear_Data, bg="red").grid(row=2, column=1, pady=5, sticky="ew")
    applybutton=tk.Button(master=lcdata_window, text="Apply", command=Apply_Button, bg=button_bg).grid(row=3, column=1, pady=15, sticky="ew")

    lcdata_window.mainloop()

#--> Bottom Left Frame --> Preferences Button Configuration
#Preferences Window
preferences_list=[] #1 Title, #2 Data, #3 Fit, #4 Frame Bg, #5 Graph Bg, #6 Label/Title Color, #7 Axis Color
def preferences_window():
    if dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        
        button_bg="#f0e130"
        
        text_color_primary="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
    if dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        
        button_bg="#C9C6C6"
        
        text_color_primary="black"
        text_color_secondary="black"
            
        color_frame_bg="#0C0B0B"

    preferences_window=tk.Toplevel(master=main_window)
    preferences_window.title("Graph Preferences")
    preferences_window.iconbitmap(r'icon_lcf.ico')
    preferences_window.configure(bg=main_bg)
    preferences_window.wm_transient(main_window)
    preferences_window.resizable(width=False, height=False)
    
    first_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    first_frame_2=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    second_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    third_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    fourth_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    fifth_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    sixth_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    seventh_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    eighth_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    ninth_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    tenth_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    eleventh_frame=tk.Frame(master=preferences_window, bg=secondary_bg, padx=10, pady=5)
    
    first_frame.grid(row=0, column=0, padx=10, pady=10)
    first_frame_2.grid(row=0, column=1, padx=10, pady=10)
    second_frame.grid(row=1, column=0, padx=10, pady=10)
    third_frame.grid(row=1, column=1, padx=10, pady=10)
    fourth_frame.grid(row=2, column=0, padx=10, pady=10)
    fifth_frame.grid(row=2, column=1, padx=10, pady=10)
    sixth_frame.grid(row=3, column=0, padx=10, pady=10)
    seventh_frame.grid(row=3, column=1, padx=10, pady=10)
    eighth_frame.grid(row=4, column=0, padx=10, pady=10)
    ninth_frame.grid(row=4, column=1, padx=10, pady=10)
    tenth_frame.grid(row=5, column=0, padx=10, pady=10)
    eleventh_frame.grid(row=5, column=1, padx=10, pady=10)
    
    lcgraph_title_label=tk.Label(master=first_frame, text="Light Curve (LC)\nGraph Title", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    lcgraph_title_entry=tk.Entry(master=first_frame, bg=tetriary_bg, fg=text_color_primary)
    lcgraph_title_entry.insert(tk.END, "Light Curve Model")
    lcgraph_title_entry.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    rvgraph_title_label=tk.Label(master=first_frame_2, text="Radial Velocity (RV)\nGraph Title", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=5, sticky="nsew")
    rvgraph_title_entry=tk.Entry(master=first_frame_2, bg=tetriary_bg, fg=text_color_primary)
    rvgraph_title_entry.insert(tk.END, "Radial Velocity Model")
    rvgraph_title_entry.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #OPTION MENU!!!
    color_list=["Black", "White", "Blue", "Red", "Yellow", "Orange", "Green", "Purple", "Gray", "Pink"]
    data_styles=[".", "o", "'", "*", "x", "+", "D", "d", "H", "h", "s", "p"]
    line_styles=["-", "--", "-.", ":"]
    
    #Data Color
    graph_data_preference_label=tk.Label(master=second_frame, text="Graph Data Color", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    data_color_variable=tk.StringVar(second_frame)
    data_color_variable.set("Select Color")
    
    data_color_menu=tk.OptionMenu(second_frame, data_color_variable, *color_list)
    data_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Fit Color
    graph_fit_preference_label=tk.Label(master=third_frame, text="Graph Fit Color", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    fit_color_variable=tk.StringVar(third_frame)
    fit_color_variable.set("Select Color")
    
    fit_color_menu=tk.OptionMenu(third_frame, fit_color_variable, *color_list)
    fit_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Frame Background Color
    bg_preference_label=tk.Label(master=fourth_frame, text="Frame Background Color", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    bg_color_variable=tk.StringVar(fourth_frame)
    bg_color_variable.set("Select Color")
    
    bg_color_menu=tk.OptionMenu(fourth_frame, bg_color_variable, *color_list)
    bg_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Graph Background Color
    graph_bg_preference_label=tk.Label(master=fifth_frame, text="Graph Background Color", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    graph_bg_color_variable=tk.StringVar(fifth_frame)
    graph_bg_color_variable.set("Select Color")
    
    graph_bg_color_menu=tk.OptionMenu(fifth_frame, graph_bg_color_variable, *color_list)
    graph_bg_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Label and Title Color
    title_preference_label=tk.Label(master=sixth_frame, text="Label and Title Color", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    title_color_variable=tk.StringVar(sixth_frame)
    title_color_variable.set("Select Color")
    
    title_color_menu=tk.OptionMenu(sixth_frame, title_color_variable, *color_list)
    title_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Axis Color
    axis_preference_label=tk.Label(master=seventh_frame, text="Axis Color", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    axis_color_variable=tk.StringVar(seventh_frame)
    axis_color_variable.set("Select Color")
    
    axis_color_menu=tk.OptionMenu(seventh_frame, axis_color_variable, *color_list)
    axis_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #System Velocity Color (RV)
    rvelocity_preference_label=tk.Label(master=eighth_frame, text="System Velocity Color (RV)", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    rvelocity_color_variable=tk.StringVar(eighth_frame)
    rvelocity_color_variable.set("Select Color")
    
    rvelocity_color_menu=tk.OptionMenu(eighth_frame, rvelocity_color_variable, *color_list)
    rvelocity_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Error Bar Color (RV)
    errrvelocity_preference_label=tk.Label(master=ninth_frame, text="Error Bar Color (RV)", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    errrvelocity_color_variable=tk.StringVar(ninth_frame)
    errrvelocity_color_variable.set("Select Color")
    
    errrvelocity_color_menu=tk.OptionMenu(ninth_frame, errrvelocity_color_variable, *color_list)
    errrvelocity_color_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Plot Point Style
    plotpoint_preference_label=tk.Label(master=tenth_frame, text="Data Point Style", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    plotpoint_variable=tk.StringVar(tenth_frame)
    plotpoint_variable.set(".")
    
    plotpoint_menu=tk.OptionMenu(tenth_frame, plotpoint_variable, *data_styles)
    plotpoint_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    #Line Style
    line_preference_label=tk.Label(master=eleventh_frame, text="Line (Fit) Style", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=10, sticky="nsew")
    
    line_variable=tk.StringVar(eleventh_frame)
    line_variable.set("-")
    
    line_menu=tk.OptionMenu(eleventh_frame, line_variable, *line_styles)
    line_menu.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
    
    if len(preferences_list)==12:
        lcgraph_title_entry.delete(0, "end")
        lcgraph_title_entry.insert(tk.END, preferences_list[0])
        rvgraph_title_entry.delete(0, "end")
        rvgraph_title_entry.insert(tk.END, preferences_list[9])
        data_color_variable.set(string.capwords(preferences_list[1]))
        fit_color_variable.set(string.capwords(preferences_list[2]))
        bg_color_variable.set(string.capwords(preferences_list[3]))
        graph_bg_color_variable.set(string.capwords(preferences_list[4]))
        title_color_variable.set(string.capwords(preferences_list[5]))
        axis_color_variable.set(string.capwords(preferences_list[6]))
        rvelocity_color_variable.set(string.capwords(preferences_list[7]))
        errrvelocity_color_variable.set(string.capwords(preferences_list[8]))
        plotpoint_variable.set((preferences_list[10]))
        line_variable.set((preferences_list[11]))
        
    def Apply_preferences():
        preferences_list.clear()
        preferences_list.append(str(lcgraph_title_entry.get()))
        preferences_list.append(str(data_color_variable.get()).lower())
        preferences_list.append(str(fit_color_variable.get()).lower())
        preferences_list.append(str(bg_color_variable.get()).lower())
        preferences_list.append(str(graph_bg_color_variable.get()).lower())
        preferences_list.append(str(title_color_variable.get()).lower())
        preferences_list.append(str(axis_color_variable.get()).lower())
        preferences_list.append(str(rvelocity_color_variable.get()).lower())
        preferences_list.append(str(errrvelocity_color_variable.get()).lower())
        preferences_list.append(str(rvgraph_title_entry.get()))
        preferences_list.append(str(plotpoint_variable.get()))
        preferences_list.append(str(line_variable.get()))
        
        if (preferences_list[1]=="select color" or preferences_list[2]=="select color" or preferences_list[3]=="select color" 
        or preferences_list[4]=="select color" or preferences_list[5]=="select color" or preferences_list[6]=="select color" 
        or preferences_list[7]=="select color" or preferences_list[8]=="select color"):
            applyerror=tk.messagebox.showerror("Error", "Please select the color for each given option.")
        
        else:
            applysuccess=tk.messagebox.showinfo("Successful", "Preferences have been applied!")
            preferences_window.withdraw()
        
    #Apply Button
    apply_colors=tk.Button(master=preferences_window, text="Apply Preferences", command=Apply_preferences, bg=button_bg)
    apply_colors.grid(row=6, column=0, columnspan=2, padx=10, pady=(10,5))
    
    #An important note
    important_note=tk.Label(master=preferences_window, text="(Note: For changes to take effect, please redraw the graphs.)", bg=main_bg, fg=text_color_primary, font=("Times New Roman", 10, "bold"))
    important_note.grid(row=7, column=0, columnspan=2, padx=10, pady=(0,5))
    
    preferences_window.mainloop()
    
preferences_button=tk.Button(master=frame_bottom_left, text="Preferences", bg=button_bg, command=preferences_window, font=("Times New Roman", 10, "bold"))
preferences_button.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

# --> Show Light Curve Button Configuration
def lightcurvewindow():
    graphtitle="Light Curve"
    
    if len(preferences_list)==12:
        graphtitle=preferences_list[0]
        data_bg=preferences_list[1]
        main_bg=preferences_list[3]
        tetriary_bg=preferences_list[4]
        text_color_primary_label=preferences_list[5]
        text_color_primary_axis=preferences_list[6]
        sys_vel_color=preferences_list[7]
        err_bar_color=preferences_list[8]
        data_style=preferences_list[10]
        line_style=preferences_list[11]
        
    elif dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        data_bg="red"
        data_style="."
        line_style="-"
        
        button_bg="#f0e130"
        
        text_color_primary_label="white"
        text_color_primary_axis="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
    elif dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        data_bg="red"
        data_style="."
        line_style="-"
    
        button_bg="#C9C6C6"
        
        text_color_primary_label="black"
        text_color_primary_axis="black"
        text_color_secondary="black"
    
        color_frame_bg="#0C0B0B"

    lc_window=tk.Toplevel(master=main_window)
    lc_window.title("Light Curve")
    lc_window.iconbitmap(r'icon_lcf.ico')
    lc_window.configure(bg=main_bg)
    lc_window.resizable(width=False, height=False)
    
    figure=Figure(figsize=(7,4), dpi=100)
    figure.patch.set_facecolor(main_bg)

    lc=figure.add_subplot(111)
    lc.plot(lc_time, lc_flux, data_style, color=data_bg)
    lc.patch.set_facecolor(tetriary_bg)
    
    lc.xaxis.label.set_color(text_color_primary_label)
    lc.yaxis.label.set_color(text_color_primary_label)
    
    lc.tick_params(axis='x', colors=text_color_primary_axis)
    lc.tick_params(axis='y', colors=text_color_primary_axis)
    
    lc.spines['left'].set_color(text_color_primary_axis)
    lc.spines['bottom'].set_color(text_color_primary_axis)
    lc.spines['right'].set_color(text_color_primary_axis)
    lc.spines['top'].set_color(text_color_primary_axis)

    lc.set_title(graphtitle, fontsize=12, color=text_color_primary_label, fontweight="bold")
    lc.set_ylabel("Intensity", fontsize=12)
    lc.set_xlabel("Time [days]", fontsize=12)

    plot_box=FigureCanvasTkAgg(figure, master=lc_window)
    plot_box.draw()
    plot_box.get_tk_widget().pack()

    toolbar=NavigationToolbar2Tk(plot_box, lc_window)
    toolbar.update()
    plot_box.get_tk_widget().pack()

# --> Star Parameters Button Configurations
def starparameters():
    if dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"

        button_bg="#f0e130"

        text_color_primary="white"
        text_color_secondary="black"

        color_frame_bg="#0C0B0B"
    
    if dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        
        button_bg="#C9C6C6"
        
        text_color_primary="black"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
    
    sp_window=tk.Toplevel(master=main_window)
    sp_window.title("Star Parameters - Light Curve")
    sp_window.iconbitmap(r'icon_lcf.ico')
    sp_window.configure(bg=main_bg)
    sp_window.wm_transient(main_window)
    sp_window.resizable(width=False, height=False)
    
    spfirst_frame=tk.Frame(master=sp_window, bg=main_bg, padx=10, pady=5)
    spsecond_frame=tk.Frame(master=sp_window, bg=main_bg, padx=10, pady=5, highlightthickness=1, highlightbackground= "red")
    
    spfirst_frame.grid(row=0, column=0, padx=10, pady=5)
    spsecond_frame.grid(row=1, column=0, padx=10, pady=10)
    
    for i in range(len(star_parameter_names_list)):
        pp_name=tk.Label(master=spfirst_frame, text=star_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=main_bg, fg=text_color_primary)
        pp_name.grid(row=1+i, column=0, sticky="nsew")
    
    for i in range(len(alt_star_parameter_names_list)):
        pp_name=tk.Label(master=spsecond_frame, text=alt_star_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=main_bg, fg=text_color_primary)
        pp_name.grid(row=5+len(star_parameter_names_list)+i, column=0, sticky="nsew")
    
    parameter_name=tk.Label(master=spfirst_frame, text="Parameter", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=5)
    parameter_value=tk.Label(master=spfirst_frame, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=1, padx=5, pady=5)
    parameter_unit=tk.Label(master=spfirst_frame, text="Unit", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=2, padx=5, pady=5)
    
    ua_ub_calculation=tk.Label(master=spsecond_frame, text="ua and ub Calculation Area", font=("Times New Roman", 10, "bold"), bg=secondary_bg, fg=text_color_primary).grid(row=7, column=0, columnspan=3, padx=5, pady=(15,5))
    
    rstar_entry=tk.Entry(master=spfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    ua_entry=tk.Entry(master=spfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    ub_entry=tk.Entry(master=spfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    
    Teff_entry=tk.Entry(master=spsecond_frame, bg=tetriary_bg, fg=text_color_primary)
    logg_entry=tk.Entry(master=spsecond_frame, bg=tetriary_bg, fg=text_color_primary)
    metallicity_entry=tk.Entry(master=spsecond_frame, bg=tetriary_bg, fg=text_color_primary)
    microturb_entry=tk.Entry(master=spsecond_frame, bg=tetriary_bg, fg=text_color_primary)
    
    if len(star_parameter_values_list)==3:        
        rstar_entry.insert(tk.END, star_parameter_values_list[0])
        ua_entry.insert(tk.END, star_parameter_values_list[1])
        ub_entry.insert(tk.END, star_parameter_values_list[2])
    
    elif len(star_parameter_values_list)==1:
        rstar_entry.insert(tk.END, star_parameter_values_list[0])
    
    if len(alt_star_parameter_values_list)==4:
        Teff_entry.insert(tk.END, alt_star_parameter_values_list[0])
        logg_entry.insert(tk.END, alt_star_parameter_values_list[1])
        metallicity_entry.insert(tk.END, alt_star_parameter_values_list[2])
        microturb_entry.insert(tk.END, alt_star_parameter_values_list[3])
    
    rstar_entry.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
    ua_entry.grid(row=2, column=1, padx=5, pady=5, sticky="nsew")
    ub_entry.grid(row=3, column=1, padx=5, pady=5, sticky="nsew")
    
    Teff_entry.grid(row=8, column=1, padx=5, pady=5, sticky="nsew")
    logg_entry.grid(row=9, column=1, padx=5, pady=5, sticky="nsew")
    metallicity_entry.grid(row=10, column=1, padx=5, pady=5, sticky="nsew")
    microturb_entry.grid(row=11, column=1, padx=5, pady=5, sticky="nsew")
    
    rstar_unit=tk.Label(master=spfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=1, column=2, padx=5, pady=5, sticky="nsew")
    ua_unit=tk.Label(master=spfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=2, column=2, padx=5, pady=5, sticky="nsew")
    ub_unit=tk.Label(master=spfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=3, column=2, padx=5, pady=5, sticky="nsew")
    
    Teff_unit=tk.Label(master=spsecond_frame, text="Kelvin", bg=main_bg, fg=text_color_primary).grid(row=8, column=2, padx=5, pady=5, sticky="nsew")
    logg_unit=tk.Label(master=spsecond_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=9, column=2, padx=5, pady=5, sticky="nsew")
    metallicity_unit=tk.Label(master=spsecond_frame, text="[Fe/H]", bg=main_bg, fg=text_color_primary).grid(row=10, padx=5, pady=5,column=2, sticky="nsew")
    microturb_unit=tk.Label(master=spsecond_frame, text="km/sec", bg=main_bg, fg=text_color_primary).grid(row=11, column=2, padx=5, pady=5, sticky="nsew")
    
    def sp():
        alt_star_parameter_values_list.clear()
        star_parameter_values_list.clear()
        
        if (rstar_entry.get()=="" or ua_entry.get()==""  or ub_entry.get()==""):
            sp_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
        
        if (Teff_entry.get()!="" and logg_entry.get()!="" and metallicity_entry.get()!="" and microturb_entry.get()!=""):
            alt_star_parameter_values_list.append(float(Teff_entry.get()))
            alt_star_parameter_values_list.append(float(logg_entry.get()))
            alt_star_parameter_values_list.append(float(metallicity_entry.get()))
            alt_star_parameter_values_list.append(float(microturb_entry.get()))
        
        star_parameter_values_list.append(float(rstar_entry.get()))
        star_parameter_values_list.append(float(ua_entry.get()))
        star_parameter_values_list.append(float(ub_entry.get()))
        
        for i in range(len(star_parameter_names_list)):
            starvalue_first=tk.Label(master=frame_middle_left, text=star_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
            starvalue_first.grid(row=2+i+len(planet_parameter_names_list), column=1, sticky="nsew")
        
        spsuccess=tk.messagebox.showinfo("Successful", "Star Parameters have been applied!")
        sp_window.withdraw()
    
    def sp_open():
        data_path=tkfile.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        
        if not data_path:
            return
    
        rstar_entry.delete(0, "end")
        ua_entry.delete(0, "end")
        ub_entry.delete(0, "end")
        
        alt_star_parameter_values_list.clear()
        star_parameter_values_list.clear()
    
        with open(data_path, mode="r", encoding="utf-8") as user_input_file:
            for line in open(data_path, mode="r"):
                values = [s for s in line.split("=")]
                star_parameter_values_list.append(float(values[1]))
            
            rstar_entry.insert(tk.END, round(star_parameter_values_list[0],4))
            ua_entry.insert(tk.END, round(star_parameter_values_list[1],3))
            ub_entry.insert(tk.END, round(star_parameter_values_list[2],3))
    
    def sp_clear():
        alt_star_parameter_values_list.clear()
        star_parameter_values_list.clear()
        
        Teff_entry.delete(0, "end")
        logg_entry.delete(0, "end")
        metallicity_entry.delete(0, "end")
        microturb_entry.delete(0, "end")
        rstar_entry.delete(0, "end")
        ua_entry.delete(0, "end")
        ub_entry.delete(0, "end")
        
        for i in range(len(star_parameter_names_list)):
            starvalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
            starvalue_first.grid(row=2+i+len(planet_parameter_names_list), column=1, sticky="nsew")
    
    def find_ua_and_ub():
        if (Teff_entry.get()=="" or logg_entry.get()=="" or metallicity_entry.get()=="" or microturb_entry.get()==""):
            ua_ub_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to calculate the ua and ub values.")
        
        elif len(star_parameter_values_list)==3:
            star_parameter_values_list.pop(2)
            star_parameter_values_list.pop(1)
        
            ua_entry.delete(0, "end")
            ub_entry.delete(0, "end")
        
        Teff=float(Teff_entry.get())
        Log_g=float(logg_entry.get())
        Metallicity=float(metallicity_entry.get())
        Micro_Turb=float(microturb_entry.get())
        
        limbdark=openpyxl.load_workbook(r"Limb Darkening Coefficients (Claret 2000).xlsx")
        limbdark=limbdark.active
        
        x=(limbdark.max_row)-1 #Maksimum Satır Sayısının 1 eksiği bize data sayısını verecektir.
        y=1 #1. Satırlarda parametrelerin ne olduğu yazıyor. Bu nedenle yukarıda 1 eksiği alıyoruz.
        Teff_values=[]
        Log_g_values=[]
        Micro_Turb_values=[]
        Metallicity_values=[]

        for i in range(x):
            y+=1
            Teff_value=float((limbdark.cell(row=y,column=4).value))
            if Teff_value in Teff_values:
                continue
    
            else:
                Teff_values.append(Teff_value)

        y=1
        for i in range(x):
            y+=1
            Log_g_value=float((limbdark.cell(row=y,column=3).value))
            if Log_g_value in Log_g_values:
                continue
        
            else:
                Log_g_values.append(Log_g_value)

        y=1
        for i in range(x):
            y+=1
            Micro_Turb_value=float((limbdark.cell(row=y,column=2).value))
            if Micro_Turb_value in Micro_Turb_values:
                continue
    
            else:
                Micro_Turb_values.append(Micro_Turb_value)
    
        y=1
        for i in range(x):
            y+=1
            Metallicity_value=float((limbdark.cell(row=y,column=5).value))
            if Metallicity_value in Metallicity_values:
                continue
    
            else:
                Metallicity_values.append(Metallicity_value)
        
        difference_values_Teff=[]

        for i in range(len(Teff_values)):
            difference=abs(Teff-Teff_values[i])
            difference_values_Teff.append(difference)

        index_Teff=difference_values_Teff.index(min(difference_values_Teff))
        Teff_limb_dark=Teff_values[index_Teff]

        difference_values_Log_g=[]

        for i in range(len(Log_g_values)):
            difference=abs(Log_g-Log_g_values[i])
            difference_values_Log_g.append(difference)

        index_Log_g=difference_values_Log_g.index(min(difference_values_Log_g))
        Log_g_limb_dark=Log_g_values[index_Log_g]

        difference_values_Micro_Turb=[]

        for i in range(len(Micro_Turb_values)):
            difference=abs(Micro_Turb-Micro_Turb_values[i])
            difference_values_Micro_Turb.append(difference)

        index_Micro_Turb=difference_values_Micro_Turb.index(min(difference_values_Micro_Turb))
        Micro_Turb_limb_dark=Micro_Turb_values[index_Micro_Turb]

        difference_values_Metallicity=[]

        for i in range(len(Metallicity_values)):
            difference=abs(Metallicity-Metallicity_values[i])
            difference_values_Metallicity.append(difference)

        index_Metallicity=difference_values_Metallicity.index(min(difference_values_Metallicity))
        Metallicity_limb_dark=Metallicity_values[index_Metallicity]
        
        y=1
        for i in range(x):
            y+=1
            Teff_value=float((limbdark.cell(row=y,column=4).value))
            Log_g_value=float((limbdark.cell(row=y,column=3).value))
            Limb_Dark_Coeff=((limbdark.cell(row=y,column=1).value))
            Micro_Turb_value=float((limbdark.cell(row=y,column=2).value))
            Metallicity_value=float((limbdark.cell(row=y,column=5).value))
    
            if Teff_value==Teff_limb_dark and Log_g_value==Log_g_limb_dark and Micro_Turb_value==Micro_Turb_limb_dark and Metallicity_value==Metallicity_limb_dark and Limb_Dark_Coeff=="a":
                ua_entry.delete(0, "end")
                ua=float((limbdark.cell(row=y,column=6).value))
                ua_entry.insert(tk.END, ua)
                star_parameter_values_list.append(ua)
        
            if Teff_value==Teff_limb_dark and Log_g_value==Log_g_limb_dark and Micro_Turb_value==Micro_Turb_limb_dark and Metallicity_value==Metallicity_limb_dark and Limb_Dark_Coeff=="b":
                ub_entry.delete(0, "end")
                ub=float((limbdark.cell(row=y,column=6).value))
                ub_entry.insert(tk.END, ub)
                star_parameter_values_list.append(ub)
    
            else:
                continue
        
        if ua_entry.get()!="" and ub_entry.get()!="" and ua_entry.get()!="0.1" and ub_entry.get()!="0.1":
            ua_ub_success=tk.messagebox.showinfo("Successful", "ua and ub values have been calculated successfully.")
        
        if ua_entry.get()=="" or ub_entry.get()=="" or ua_entry.get()=="0.1" or ub_entry.get()=="0.1":
            ua_entry.delete(0, "end")
            ub_entry.delete(0, "end")
            
            ua_entry.insert(tk.END, float(0.1))
            ub_entry.insert(tk.END, float(0.1))
            
            star_parameter_values_list.append(0.1)
            star_parameter_values_list.append(0.1)
            
            ua_ub_success=tk.messagebox.showinfo("Successful", "ua and ub values have been set to standard 0.1 value. \n\nTo get a better ua and ub value please check the representative boxes in the ''Model Parameters'' area when fitting. \n\n(There isn't a suitable ua and ub value for the Star Parameters entered.)")
            
            if len(star_parameter_values_list)==3:
                star_parameter_values_list.pop(2)
                star_parameter_values_list.pop(1)
            
            elif len(star_parameter_values_list)==2:
                star_parameter_values_list.pop(1)
    
    sp_open_button=tk.Button(master=spfirst_frame, text="Open Data", command=sp_open, bg="green").grid(row=4, column=0, columnspan=3, sticky="nsew", padx=45, pady=(15,5))
    sp_clear_button=tk.Button(master=spfirst_frame, text="Clear Data", command=sp_clear, bg="red").grid(row=5, column=0, columnspan=3, sticky="nsew", padx=45, pady=5)  
    sp_apply_button=tk.Button(master=spfirst_frame, text="Apply", command=sp, bg=button_bg).grid(row=6, column=0, columnspan=3, sticky="nsew", padx=45, pady=15)
    sp_ua_ub_button=tk.Button(master=spsecond_frame, text="Calculate ua and ub", command=find_ua_and_ub, bg=button_bg).grid(row=14, column=0, columnspan=3, sticky="nsew", padx=45, pady=15)

# --> Planet Parameters Button Configuration
def planetparameters():
    if dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"

        button_bg="#f0e130"

        text_color_primary="white"
        text_color_secondary="black"

        color_frame_bg="#0C0B0B"
    
    if dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        
        button_bg="#C9C6C6"
        
        text_color_primary="black"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
    
    pp_window=tk.Toplevel(master=main_window)
    pp_window.title("Planet Parameters - Light Curve")
    pp_window.iconbitmap(r'icon_lcf.ico')
    pp_window.configure(bg=main_bg)
    pp_window.wm_transient(main_window)
    pp_window.resizable(width=False, height=False)
    
    ppfirst_frame=tk.Frame(master=pp_window, bg=main_bg, padx=10, pady=5)
    ppfirst_frame.grid(row=0, column=0, padx=10, pady=5)
    
    for i in range(len(planet_parameter_names_list)):
        pp_name=tk.Label(master=ppfirst_frame, text=planet_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=main_bg, fg=text_color_primary)
        pp_name.grid(row=1+i, column=0, sticky="nsew")
    
    parameter_name=tk.Label(master=ppfirst_frame, text="Parameter", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=5)
    parameter_value=tk.Label(master=ppfirst_frame, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=1, padx=5, pady=5)
    parameter_unit=tk.Label(master=ppfirst_frame, text="Unit", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=2, padx=5, pady=5)
    
    T0_entry=tk.Entry(master=ppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    period_entry=tk.Entry(master=ppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    eccentricity_entry=tk.Entry(master=ppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    argumentofperiapsis_entry=tk.Entry(master=ppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    inclination_entry=tk.Entry(master=ppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    rplanet_entry=tk.Entry(master=ppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
    
    if len(planet_parameter_values_list)==6:
        T0_entry.insert(tk.END, planet_parameter_values_list[0])
        period_entry.insert(tk.END, planet_parameter_values_list[1])
        eccentricity_entry.insert(tk.END, planet_parameter_values_list[2])
        argumentofperiapsis_entry.insert(tk.END, planet_parameter_values_list[3])
        inclination_entry.insert(tk.END, planet_parameter_values_list[4])
        rplanet_entry.insert(tk.END, planet_parameter_values_list[5])
    
    T0_entry.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
    period_entry.grid(row=2, column=1, padx=5, pady=5, sticky="nsew")
    eccentricity_entry.grid(row=3, column=1, padx=5, pady=5, sticky="nsew")
    argumentofperiapsis_entry.grid(row=4, column=1, padx=5, pady=5, sticky="nsew")
    inclination_entry.grid(row=5, column=1, padx=5, pady=5, sticky="nsew")
    rplanet_entry.grid(row=6, column=1, padx=5, pady=5, sticky="nsew")
    
    T0_unit=tk.Label(master=ppfirst_frame, text="Days", bg=main_bg, fg=text_color_primary).grid(row=1, column=2, padx=5, pady=5, sticky="nsew")
    period_unit=tk.Label(master=ppfirst_frame, text="Days", bg=main_bg, fg=text_color_primary).grid(row=2, column=2, padx=5, pady=5, sticky="nsew")
    eccentricity_unit=tk.Label(master=ppfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=3, column=2, padx=5, pady=5, sticky="nsew")
    argumentofperiapsis_unit=tk.Label(master=ppfirst_frame, text="Radians", bg=main_bg, fg=text_color_primary).grid(row=4, column=2, padx=5, pady=5, sticky="nsew")
    inclination_unit=tk.Label(master=ppfirst_frame, text="Radians", bg=main_bg, fg=text_color_primary).grid(row=5, column=2, padx=5, pady=5, sticky="nsew")
    rplanet_unit=tk.Label(master=ppfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=6, column=2, padx=5, pady=5, sticky="nsew")
    
    def pp():
        planet_parameter_values_list.clear()
        
        if (T0_entry.get()=="" or period_entry.get()==""
        or eccentricity_entry.get()=="" or argumentofperiapsis_entry.get()==""
        or inclination_entry.get()=="" or rplanet_entry.get()==""):
            pp_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
        
        planet_parameter_values_list.append(float(T0_entry.get()))
        planet_parameter_values_list.append(float(period_entry.get()))
        planet_parameter_values_list.append(float(eccentricity_entry.get()))
        planet_parameter_values_list.append(float(argumentofperiapsis_entry.get()))
        planet_parameter_values_list.append(float(inclination_entry.get()))
        planet_parameter_values_list.append(float(rplanet_entry.get()))

        for i in range(len(planet_parameter_names_list)):
            parametervalue=tk.Label(master=frame_middle_left, text=planet_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
            parametervalue.grid(row=2+i, column=1, sticky="nsew")
        
        ppsuccess=tk.messagebox.showinfo("Successful", "Planet Parameters have been applied!")
        pp_window.withdraw()
    
    def pp_open():
        data_path=tkfile.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
        
        if not data_path:
            return
    
        T0_entry.delete(0, "end")
        period_entry.delete(0, "end")
        eccentricity_entry.delete(0, "end")
        argumentofperiapsis_entry.delete(0, "end")
        inclination_entry.delete(0, "end")
        rplanet_entry.delete(0, "end")
        planet_parameter_values_list.clear()
    
        with open(data_path, mode="r", encoding="utf-8") as user_input_file:
            for line in open(data_path, mode="r"):
                values = [s for s in line.split("=")]
                planet_parameter_values_list.append(float(values[1]))
            
            T0_entry.insert(tk.END, round(planet_parameter_values_list[0],4))
            period_entry.insert(tk.END, round(planet_parameter_values_list[1],6))
            eccentricity_entry.insert(tk.END, round(planet_parameter_values_list[2],3))
            argumentofperiapsis_entry.insert(tk.END, round(planet_parameter_values_list[3],3))
            inclination_entry.insert(tk.END, round(planet_parameter_values_list[4],3))
            rplanet_entry.insert(tk.END, round(planet_parameter_values_list[5],4))
    
    def pp_clear():
        planet_parameter_values_list.clear()
       
        T0_entry.delete(0, "end")
        period_entry.delete(0, "end")
        eccentricity_entry.delete(0, "end")
        argumentofperiapsis_entry.delete(0, "end")
        inclination_entry.delete(0, "end")
        rplanet_entry.delete(0, "end")
        
        for i in range(len(planet_parameter_names_list)):
            parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=2+i, column=1, sticky="nsew")
    
    pp_open_button=tk.Button(master=ppfirst_frame, text="Open Data", command=pp_open, bg="green").grid(row=11, column=0, columnspan=3, sticky="nsew", padx=45, pady=(15,5))
    pp_clear_button=tk.Button(master=ppfirst_frame, text="Clear Data", command=pp_clear, bg="red").grid(row=12, column=0, columnspan=3, sticky="nsew", padx=45, pady=5)  
    pp_apply_button=tk.Button(master=ppfirst_frame, text="Apply", command=pp, bg=button_bg).grid(row=13, column=0, columnspan=3, sticky="nsew", padx=45, pady=15)

def rvanalysis_window():
    if dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        
        button_bg="#f0e130"
        
        text_color_primary="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
    if dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        
        button_bg="#C9C6C6"
        
        text_color_primary="black"
        text_color_secondary="black"
            
        color_frame_bg="#0C0B0B"

    rv_window=tk.Toplevel(master=main_window)
    rv_window.title("Radial Velocity Analysis")
    rv_window.iconbitmap(r'icon_lcf.ico')
    rv_window.configure(bg=main_bg)
    rv_window.wm_transient(main_window)
    rv_window.resizable(width=False, height=False)
    
    frame_top_leftrv=tk.Frame(master=rv_window, bg=secondary_bg, padx=10, pady=0)
    frame_middle_leftrv=tk.Frame(master=rv_window, bg=secondary_bg, padx=10, pady=0)
    frame_bottom_leftrv=tk.Frame(master=rv_window, bg=secondary_bg, padx=10, pady=5)

    frame_top_rightrv=tk.Frame(master=rv_window, bg=secondary_bg, padx=10, pady=5)
    frame_middle_rightrv=tk.Frame(master=rv_window, bg=secondary_bg, padx=10, pady=5)
    frame_bottom_rightrv=tk.Frame(master=rv_window, bg=secondary_bg, padx=10, pady=5)

    frame_top_leftrv.grid(row=0, column=0, sticky="new", padx=(15,10), pady=10)
    frame_top_rightrv.grid(row=0, column=1, rowspan=2, padx=(15,10), pady=10)
    frame_middle_leftrv.grid(row=1, column=0, sticky="new", padx=(15,10), pady=10)
    frame_middle_rightrv.grid(row=2, column=1, sticky="nsew", padx=(15,10), pady=10)
    frame_bottom_leftrv.grid(row=2, column=0, rowspan=2, sticky="nsew", padx=(15,10), pady=10)
    frame_bottom_rightrv.grid(row=3, column=1, padx=(15,10), pady=10)
    
    def rvdataentry_window():
        if dl_button['text']=="Light Mode":
            main_bg="#0B0B0B"
            secondary_bg="#252525"
            tetriary_bg="#3A3A3A"
            title_bg="#FF7A37"
            
            button_bg="#f0e130"
            
            text_color_primary="white"
            text_color_secondary="black"
            
            color_frame_bg="#0C0B0B"
            
        if dl_button['text']=="Dark Mode":
            main_bg="#ECDBDB"
            secondary_bg="#FFFFFF"
            tetriary_bg="#FFFFFF"
            title_bg="#7A7A7A"
            
            button_bg="#C9C6C6"
            
            text_color_primary="black"
            text_color_secondary="black"
                
            color_frame_bg="#0C0B0B"
    
        rvdata_window=tk.Toplevel(master=rv_window)
        rvdata_window.title("Radial Velocity Data Entry")
        rvdata_window.iconbitmap(r'icon_lcf.ico')
        rvdata_window.configure(bg=main_bg)
        rvdata_window.wm_transient(rv_window)
        rvdata_window.resizable(width=False, height=False)
        
        datawindow_rv=tk.Text(master=rvdata_window, bg=tetriary_bg, fg=text_color_primary)
        datawindow_rv.grid(row=0, column=0, columnspan=3, padx=10, pady=10)
    
        if len(obs_time_list)!=0 or len(v_radial_list)!=0:
            for i in range(len(obs_time_list)):
                datawindow_rv.insert(tk.END, obs_time_list[i])
                if v_error_list[i]!=0:
                    datawindow_rv.insert(tk.END, f"	{v_radial_list[i]}")
                    datawindow_rv.insert(tk.END, f"	{v_error_list[i]}\n")
                
                if v_error_list[i]==0:
                    datawindow_rv.insert(tk.END, f"	{v_radial_list[i]}\n")
    
        def Open_Data():
            data_path=tkfile.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
            
            if not data_path:
                return
        
            datawindow_rv.delete("1.0", tk.END)
            obs_time_list.clear()
            v_radial_list.clear()
            v_error_list.clear()
            weights.clear()
            rv_result_parameter_values_list.clear()
            
            unit_window=tk.Toplevel(master=rvdata_window)
            unit_window.title("Radial Velocity Unit")
            unit_window.iconbitmap(r'icon_lcf.ico')
            unit_window.configure(bg=main_bg)
            unit_window.wm_transient(datawindow_rv)
            unit_window.resizable(width=False, height=False)
            
            style_radiobutton=ttk.Style(unit_window)
            style_radiobutton.configure("TRadiobutton", background=main_bg, foreground=text_color_primary)
        
            unit_vary=tk.StringVar()
        
            unit_label=tk.Label(master=unit_window, text="Velocity Unit:", bg=title_bg, font=("Times New Roman", 10, "bold"))
            unit_check_msec=ttk.Radiobutton(master=unit_window, text="m/sec", variable=unit_vary, value="m/sec")
            unit_check_kmsec=ttk.Radiobutton(master=unit_window, text="km/sec", variable=unit_vary, value="km/sec")
        
            unit_label.grid(row=0, column=0, pady=10, padx=50)
            unit_check_msec.grid(row=1, column=0, pady=5, padx=50)
            unit_check_kmsec.grid(row=2, column=0, pady=5, padx=50)
            
            error_included=tk.StringVar()
            
            err_label=tk.Label(master=unit_window, text="Does the data\n include errors?", bg=title_bg, font=("Times New Roman", 10, "bold"))
            err_check_yes=ttk.Radiobutton(master=unit_window, text="Yes", variable=error_included, value="Yes")
            err_check_no=ttk.Radiobutton(master=unit_window, text="No", variable=error_included, value="No")
            
            err_label.grid(row=3, column=0, pady=(20,10), padx=50)
            err_check_yes.grid(row=4, column=0, pady=5, padx=50)
            err_check_no.grid(row=5, column=0, pady=5, padx=50)
            
            def apply_unit():
                if unit_vary.get()=="":
                    uniterr=tk.messagebox.showerror("Error", "Please select the Velocity Unit to continue.")
                
                elif error_included.get()=="":
                    errerrxd=tk.messagebox.showerror("Error", "Please select if your data includes error or not to continue.")
                
                else:
                    with open(data_path, mode="r", encoding="utf-8") as user_input_file:
                        rv_input=user_input_file.read()
            
                    for line in open(data_path, mode="r"):
                        values = [float(s) for s in line.split()]
                        if error_included.get()=="Yes" and unit_vary.get()=="m/sec":
                            if len(values)!=3:
                                err=tk.messagebox.showerror("Error", "The errors are not included in the data entered. Please check the data entry.")
                                break
                            
                            else:
                                v_error_list.append(float(values[2]/1000))
                                weight=1/(float(values[2]/1000))
                                weights.append(weight)
                        
                        elif error_included.get()=="Yes" and unit_vary.get()=="km/sec":
                            if len(values)!=3:
                                err=tk.messagebox.showerror("Error", "The errors are not included in the data entered. Please check the data entry.")
                                break
                            
                            else:
                                v_error_list.append(float(values[2]))
                                weight=1/(float(values[2]))
                                weights.append(weight)
                        
                        else:
                            v_error_list.append(float(0))
                            weight=1
                            weights.append(weight)
                        
                        obs_time_list.append(float(values[0]))
                        if unit_vary.get()=="m/sec":
                            v_radial_list.append(float(values[1]/1000))
                        
                        else:
                            v_radial_list.append(float(values[1]))
                    
                    if len(v_error_list)!=0:
                        unitsuccess=tk.messagebox.showinfo("Successful", f"Velocity Unit and Error Preference has been applied! (Unit: {unit_vary.get()})")
                    
                    datawindow_rv.configure(state="normal")
                    for i in range(len(obs_time_list)):
                        datawindow_rv.insert(tk.END, obs_time_list[i])
                        if error_included.get()=="Yes":
                            datawindow_rv.insert(tk.END, f"	{v_radial_list[i]}")
                            datawindow_rv.insert(tk.END, f"	{v_error_list[i]}\n")
                        
                        if error_included.get()=="No":
                            datawindow_rv.insert(tk.END, f"	{v_radial_list[i]}\n")
                    
                    unit_window.withdraw()
            
            apply_unit=tk.Button(master=unit_window, text="Apply", command=apply_unit, bg=button_bg).grid(row=6, column=0, pady=(20,10), padx=50, sticky="ew")
        
            rvdata_window.title(f"Radial Velocity Data Entry - {data_path}")
            datawindow_rv.configure(state="disabled")
            
            rvdata_window.mainloop()
        
        def Clear_Data():
            obs_time_list.clear()
            v_radial_list.clear()
            v_error_list.clear()
            weights.clear()
            rv_result_parameter_values_list.clear()
        
            datawindow_rv.configure(state="normal")
            datawindow_rv.delete("1.0", tk.END)
            rvdata_window.title("Radial Velocity Data Entry")
        
        def Apply_Button():
            if datawindow_rv.get("1.0", tk.END)=="\n":
                rverr=tk.messagebox.showerror("Error", "There is no data applied!")
            
            else:
                rvsuccess=tk.messagebox.showinfo("Successful", "Data has been applied!")
                rvdata_window.withdraw()
    
        dataopen=tk.Button(master=rvdata_window, text="Open Data", command=Open_Data, bg="green").grid(row=2, column=1, pady=5, sticky="ew")
        dataclear=tk.Button(master=rvdata_window, text="Clear Data", command=Clear_Data, bg="red").grid(row=3, column=1, pady=5, sticky="ew")
        applybutton=tk.Button(master=rvdata_window, text="Apply", command=Apply_Button, bg=button_bg).grid(row=4, column=1, pady=15, sticky="ew")
    
        rvdata_window.mainloop()

    def rvplanetparameters():
        if dl_button['text']=="Light Mode":
            main_bg="#0B0B0B"
            secondary_bg="#252525"
            tetriary_bg="#3A3A3A"
            title_bg="#FF7A37"
    
            button_bg="#f0e130"
    
            text_color_primary="white"
            text_color_secondary="black"
    
            color_frame_bg="#0C0B0B"
        
        if dl_button['text']=="Dark Mode":
            main_bg="#ECDBDB"
            secondary_bg="#FFFFFF"
            tetriary_bg="#FFFFFF"
            title_bg="#7A7A7A"
            
            button_bg="#C9C6C6"
            
            text_color_primary="black"
            text_color_secondary="black"
            
            color_frame_bg="#0C0B0B"
        
        rvpp_window=tk.Toplevel(master=rv_window)
        rvpp_window.title("Planet Parameters - Radial Velocity")
        rvpp_window.iconbitmap(r'icon_lcf.ico')
        rvpp_window.configure(bg=main_bg)
        rvpp_window.wm_transient(rv_window)
        rvpp_window.resizable(width=False, height=False)
        
        rvppfirst_frame=tk.Frame(master=rvpp_window, bg=main_bg, padx=10, pady=5)
        rvppfirst_frame.grid(row=0, column=0, padx=10, pady=5)
        
        for i in range(len(rvplanet_parameter_names_list)):
            pp_name=tk.Label(master=rvppfirst_frame, text=rvplanet_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=main_bg, fg=text_color_primary)
            pp_name.grid(row=1+i, column=0, sticky="nsew")
        
        parameter_name=tk.Label(master=rvppfirst_frame, text="Parameter", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=5)
        parameter_value=tk.Label(master=rvppfirst_frame, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=1, padx=5, pady=5)
        parameter_unit=tk.Label(master=rvppfirst_frame, text="Unit", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=2, padx=5, pady=5)
        
        rvT0_entry=tk.Entry(master=rvppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        rvperiod_entry=tk.Entry(master=rvppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        rveccentricity_entry=tk.Entry(master=rvppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        rvargumentofperiapsis_entry=tk.Entry(master=rvppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        rvinclination_entry=tk.Entry(master=rvppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        rvrplanet_entry=tk.Entry(master=rvppfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        
        if len(rvplanet_parameter_values_list)==6:
            rvT0_entry.insert(tk.END, rvplanet_parameter_values_list[0])
            rvperiod_entry.insert(tk.END, rvplanet_parameter_values_list[1])
            rveccentricity_entry.insert(tk.END, rvplanet_parameter_values_list[2])
            rvargumentofperiapsis_entry.insert(tk.END, rvplanet_parameter_values_list[3])
            rvinclination_entry.insert(tk.END, rvplanet_parameter_values_list[4])
            rvrplanet_entry.insert(tk.END, rvplanet_parameter_values_list[5])
        
        elif len(planet_parameter_values_list)==6:
            rvT0_entry.insert(tk.END, planet_parameter_values_list[0])
            rvperiod_entry.insert(tk.END, planet_parameter_values_list[1])
            rveccentricity_entry.insert(tk.END, planet_parameter_values_list[2])
            rvargumentofperiapsis_entry.insert(tk.END, planet_parameter_values_list[3])
            rvinclination_entry.insert(tk.END, planet_parameter_values_list[4])
            rvrplanet_entry.insert(tk.END, planet_parameter_values_list[5])
        
        rvT0_entry.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
        rvperiod_entry.grid(row=2, column=1, padx=5, pady=5, sticky="nsew")
        rveccentricity_entry.grid(row=3, column=1, padx=5, pady=5, sticky="nsew")
        rvargumentofperiapsis_entry.grid(row=4, column=1, padx=5, pady=5, sticky="nsew")
        rvinclination_entry.grid(row=5, column=1, padx=5, pady=5, sticky="nsew")
        rvrplanet_entry.grid(row=6, column=1, padx=5, pady=5, sticky="nsew")
        
        rvT0_unit=tk.Label(master=rvppfirst_frame, text="Days", bg=main_bg, fg=text_color_primary).grid(row=1, column=2, padx=5, pady=5, sticky="nsew")
        rvperiod_unit=tk.Label(master=rvppfirst_frame, text="Days", bg=main_bg, fg=text_color_primary).grid(row=2, column=2, padx=5, pady=5, sticky="nsew")
        rveccentricity_unit=tk.Label(master=rvppfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=3, column=2, padx=5, pady=5, sticky="nsew")
        rvargumentofperiapsis_unit=tk.Label(master=rvppfirst_frame, text="Radians", bg=main_bg, fg=text_color_primary).grid(row=4, column=2, padx=5, pady=5, sticky="nsew")
        rvinclination_unit=tk.Label(master=rvppfirst_frame, text="Radians", bg=main_bg, fg=text_color_primary).grid(row=5, column=2, padx=5, pady=5, sticky="nsew")
        rvrplanet_unit=tk.Label(master=rvppfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=6, column=2, padx=5, pady=5, sticky="nsew")
        
        def rvpp():
            rvplanet_parameter_values_list.clear()
            
            if (rvT0_entry.get()=="" or rvperiod_entry.get()==""
            or rveccentricity_entry.get()=="" or rvargumentofperiapsis_entry.get()==""
            or rvinclination_entry.get()=="" or rvrplanet_entry.get()==""):
                pp_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
            
            rvplanet_parameter_values_list.append(float(rvT0_entry.get()))
            rvplanet_parameter_values_list.append(float(rvperiod_entry.get()))
            rvplanet_parameter_values_list.append(float(rveccentricity_entry.get()))
            rvplanet_parameter_values_list.append(float(rvargumentofperiapsis_entry.get()))
            rvplanet_parameter_values_list.append(float(rvinclination_entry.get()))
            rvplanet_parameter_values_list.append(float(rvrplanet_entry.get()))
    
            parametervalue1=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[0], bg=secondary_bg, fg=text_color_primary)
            parametervalue1.grid(row=2+2, column=1, sticky="nsew")
            
            parametervalue2=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[1], bg=secondary_bg, fg=text_color_primary)
            parametervalue2.grid(row=2+2+1, column=1, sticky="nsew")
            
            parametervalue3=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[2], bg=secondary_bg, fg=text_color_primary)
            parametervalue3.grid(row=2+2+2, column=1, sticky="nsew")
            
            parametervalue4=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[3], bg=secondary_bg, fg=text_color_primary)
            parametervalue4.grid(row=2+2+3, column=1, sticky="nsew")

            rvppsuccess=tk.messagebox.showinfo("Successful", "Planet Parameters have been applied!")
            rvpp_window.withdraw()
        
        def rvpp_open():
            data_path=tkfile.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
            
            if not data_path:
                return
        
            rvT0_entry.delete(0, "end")
            rvperiod_entry.delete(0, "end")
            rveccentricity_entry.delete(0, "end")
            rvargumentofperiapsis_entry.delete(0, "end")
            rvinclination_entry.delete(0, "end")
            rvrplanet_entry.delete(0, "end")
            rvplanet_parameter_values_list.clear()
        
            with open(data_path, mode="r", encoding="utf-8") as user_input_file:
                for line in open(data_path, mode="r"):
                    values = [s for s in line.split("=")]
                    rvplanet_parameter_values_list.append(float(values[1]))
                
                rvT0_entry.insert(tk.END, round(rvplanet_parameter_values_list[0],4))
                rvperiod_entry.insert(tk.END, round(rvplanet_parameter_values_list[1],6))
                rveccentricity_entry.insert(tk.END, round(rvplanet_parameter_values_list[2],3))
                rvargumentofperiapsis_entry.insert(tk.END, round(rvplanet_parameter_values_list[3],3))
                rvinclination_entry.insert(tk.END, round(rvplanet_parameter_values_list[4],3))
                rvrplanet_entry.insert(tk.END, round(rvplanet_parameter_values_list[5],4))
    
        def rvpp_clear():
            rvplanet_parameter_values_list.clear()
           
            rvT0_entry.delete(0, "end")
            rvperiod_entry.delete(0, "end")
            rveccentricity_entry.delete(0, "end")
            rvargumentofperiapsis_entry.delete(0, "end")
            rvinclination_entry.delete(0, "end")
            rvrplanet_entry.delete(0, "end")
            
            for i in range(len(rvplanet_parameter_names_list)):
                parametervalue_first=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
                parametervalue_first.grid(row=2+i, column=1, sticky="nsew")
    
        rvpp_open_button=tk.Button(master=rvppfirst_frame, text="Open Data", command=rvpp_open, bg="green").grid(row=11, column=0, columnspan=3, sticky="nsew", padx=45, pady=(15,5))
        rvpp_clear_button=tk.Button(master=rvppfirst_frame, text="Clear Data", command=rvpp_clear, bg="red").grid(row=12, column=0, columnspan=3, sticky="nsew", padx=45, pady=5)  
        rvpp_apply_button=tk.Button(master=rvppfirst_frame, text="Apply", command=rvpp, bg=button_bg).grid(row=13, column=0, columnspan=3, sticky="nsew", padx=45, pady=15)

    def rvstarparameters():
        if dl_button['text']=="Light Mode":
            main_bg="#0B0B0B"
            secondary_bg="#252525"
            tetriary_bg="#3A3A3A"
            title_bg="#FF7A37"
    
            button_bg="#f0e130"
    
            text_color_primary="white"
            text_color_secondary="black"
    
            color_frame_bg="#0C0B0B"
        
        if dl_button['text']=="Dark Mode":
            main_bg="#ECDBDB"
            secondary_bg="#FFFFFF"
            tetriary_bg="#FFFFFF"
            title_bg="#7A7A7A"
            
            button_bg="#C9C6C6"
            
            text_color_primary="black"
            text_color_secondary="black"
            
            color_frame_bg="#0C0B0B"
        
        rvsp_window=tk.Toplevel(master=rv_window)
        rvsp_window.title("Star Parameters - Radial Velocity")
        rvsp_window.iconbitmap(r'icon_lcf.ico')
        rvsp_window.configure(bg=main_bg)
        rvsp_window.wm_transient(rv_window)
        rvsp_window.resizable(width=False, height=False)
        
        rvspfirst_frame=tk.Frame(master=rvsp_window, bg=main_bg, padx=10, pady=5)
        rvspfirst_frame.grid(row=0, column=0, padx=10, pady=5)
        
        if(len(obs_time_list)==0 or len(v_radial_list)==0):
            rvsp_error=tk.messagebox.showerror("Error", "Please enter the Radial Velocity data first to proceed.\n\n(Radial Velocity data is needed in order to calculate a few of the star parameters and the graph itself.)")
            rvsp_window.withdraw()
            
        elif(len(rvplanet_parameter_values_list)!=6):
            rvsp_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters first to proceed.\n\n(Planet Parameters are needed in order to calculate a few of the star parameters and the graph itself.)")
            rvsp_window.withdraw()
            
        for i in range(len(rvstar_parameter_names_list)):
            rvsp_name=tk.Label(master=rvspfirst_frame, text=rvstar_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=main_bg, fg=text_color_primary)
            rvsp_name.grid(row=1+i, column=0, sticky="nsew")
        
        parameter_name=tk.Label(master=rvspfirst_frame, text="Parameter", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=0, padx=5, pady=5)
        parameter_value=tk.Label(master=rvspfirst_frame, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=1, padx=5, pady=5)
        parameter_unit=tk.Label(master=rvspfirst_frame, text="Unit", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary).grid(row=0, column=2, padx=5, pady=5)
        
        vgama_entry=tk.Entry(master=rvspfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        K_entry=tk.Entry(master=rvspfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        rstar_entry=tk.Entry(master=rvspfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        mass_star_entry=tk.Entry(master=rvspfirst_frame, bg=tetriary_bg, fg=text_color_primary)
        
        T0=rvplanet_parameter_values_list[0]
        Period=rvplanet_parameter_values_list[1]
        
        def Epoch_func(obs_time_list, T0, Period):
            rvEpoch_List=[] #Phase List
            rvT_T0_values=[] #T-T0 values list
    
            for time in obs_time_list:
                if ((float(time)-T0)/Period)<0:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
            
                else:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360
            
                if Epoch_value<0.2:
                    Delta_phase=Epoch_value+1
            
                else:
                    Delta_phase=Epoch_value
                
                rvEpoch_List.append(Delta_phase)
            
                T_T0=time-T0
                rvT_T0_values.append(T_T0)
            
            return rvEpoch_List, rvT_T0_values
        
        rvEpoch_List, rvT_T0_values=Epoch_func(obs_time_list, T0, Period)

        phase_min=min(rvEpoch_List)
        periastron_1=v_radial_list[rvEpoch_List.index(phase_min)]
        
        phase_max=max(rvEpoch_List)
        periastron_2=v_radial_list[rvEpoch_List.index(phase_max)]
        
        periastron_1_dif=abs(round(periastron_1)-periastron_1)
        periastron_2_dif=abs(round(periastron_2)-periastron_2)
        
        if periastron_1_dif<=periastron_2_dif:
            v_gama=periastron_1
        
        if periastron_2_dif<periastron_1_dif:
            v_gama=periastron_2
         
        vr_max=max(v_radial_list)
        vr_min=min(v_radial_list)
        
        A=vr_max-v_gama
        B=v_gama-vr_min #Negatif değerlerle uğraşıldığından eksili yazıldı.
        
        K=(A+B)/2 #Unit: [km/sn]
        
        if len(rvstar_parameter_values_list)==0 or len(rvstar_parameter_values_list)==3:
            vgama_entry.insert(tk.END, round(v_gama,4))
            K_entry.insert(tk.END, round(K,4))
        
        if len(rvstar_parameter_values_list)==4:
            vgama_entry.insert(tk.END, rvstar_parameter_values_list[0])
            K_entry.insert(tk.END, rvstar_parameter_values_list[1])
            rstar_entry.insert(tk.END, rvstar_parameter_values_list[2])
            mass_star_entry.insert(tk.END, rvstar_parameter_values_list[3])
        
        elif len(star_parameter_values_list)==3:
            rstar_entry.insert(tk.END, star_parameter_values_list[0])
        
        vgama_entry.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")
        K_entry.grid(row=2, column=1, padx=5, pady=5, sticky="nsew")
        rstar_entry.grid(row=3, column=1, padx=5, pady=5, sticky="nsew")
        mass_star_entry.grid(row=4, column=1, padx=5, pady=5, sticky="nsew")
        
        vgama_unit=tk.Label(master=rvspfirst_frame, text="km/sec", bg=main_bg, fg=text_color_primary).grid(row=1, column=2, padx=5, pady=5, sticky="nsew")
        K_unit=tk.Label(master=rvspfirst_frame, text="km/sec", bg=main_bg, fg=text_color_primary).grid(row=2, column=2, padx=5, pady=5, sticky="nsew")
        rstar_unit=tk.Label(master=rvspfirst_frame, text="---", bg=main_bg, fg=text_color_primary).grid(row=3, column=2, padx=5, pady=5, sticky="nsew")
        mass_star_unit=tk.Label(master=rvspfirst_frame, text="M\u2609", bg=main_bg, fg=text_color_primary).grid(row=4, column=2, padx=5, pady=5, sticky="nsew")
        
        def rvsp():
            rvstar_parameter_values_list.clear()
            
            if (vgama_entry.get()=="" or K_entry.get()==""  or rstar_entry.get()=="" or mass_star_entry.get()==""):
                rvsp_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
                        
            rvstar_parameter_values_list.append(float(vgama_entry.get()))
            rvstar_parameter_values_list.append(float(K_entry.get()))
            rvstar_parameter_values_list.append(float(rstar_entry.get()))
            rvstar_parameter_values_list.append(round(float(mass_star_entry.get()),3))
            
            for i in range(len(rvstar_parameter_names_list)-2):
                starvalue_first=tk.Label(master=frame_middle_leftrv, text=rvstar_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
                starvalue_first.grid(row=2+i, column=1, sticky="nsew")
            
            rv_mass_star.clear()
            rv_mass_star.append(round(float(mass_star_entry.get()),3))
            
            spsuccess=tk.messagebox.showinfo("Successful", "Star Parameters have been applied!")
            rvsp_window.withdraw()
        
        def rvsp_open():
            data_path=tkfile.askopenfilename(filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")])
            
            if not data_path:
                return
        
            vgama_entry.delete(0, "end")
            K_entry.delete(0, "end")
            rstar_entry.delete(0, "end")
            mass_star_entry.delete(0, "end")

            rvstar_parameter_values_list.clear()
        
            with open(data_path, mode="r", encoding="utf-8") as user_input_file:
                for line in open(data_path, mode="r"):
                    values = [s for s in line.split("=")]
                    rvstar_parameter_values_list.append(float(values[1]))
                
                vgama_entry.insert(tk.END, round(rvstar_parameter_values_list[0],4))
                K_entry.insert(tk.END, round(rvstar_parameter_values_list[1],4))
                rstar_entry.insert(tk.END, round(rvstar_parameter_values_list[2],4))
                mass_star_entry.insert(tk.END, round(rvstar_parameter_values_list[3],3))
    
        def rvsp_clear():
            rvstar_parameter_values_list.clear()
            
            vgama_entry.delete(0, "end")
            K_entry.delete(0, "end")
            rstar_entry.delete(0, "end")
            mass_star_entry.delete(0, "end")
            
            for i in range(len(rvstar_parameter_names_list)-2):
                starvalue_first=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
                starvalue_first.grid(row=i+2, column=1, sticky="nsew")
        
        rvsp_open_button=tk.Button(master=rvspfirst_frame, text="Open Data", command=rvsp_open, bg="green").grid(row=6, column=0, columnspan=3, sticky="nsew", padx=45, pady=(15,5))
        rvsp_clear_button=tk.Button(master=rvspfirst_frame, text="Clear Data", command=rvsp_clear, bg="red").grid(row=7, column=0, columnspan=3, sticky="nsew", padx=45, pady=5)  
        rvsp_apply_button=tk.Button(master=rvspfirst_frame, text="Apply", command=rvsp, bg=button_bg).grid(row=8, column=0, columnspan=3, sticky="nsew", padx=45, pady=15)
    
    #Top Left Frame --> The Buttons
    entry_label=tk.Label(master=frame_top_leftrv, text="Data Entry Area", padx=10, pady=10, bg=title_bg, font=("Times New Roman", 10, "bold"), width=46)
    rv_entry_button=tk.Button(master=frame_top_leftrv, text="Radial Velocity Data Entry", padx=10, pady=5, command=rvdataentry_window, bg=button_bg)
    parametersofstar_button=tk.Button(master=frame_top_leftrv, text="Star Parameters", command=rvstarparameters, padx=10, pady=5, bg=button_bg)
    parametersofplanet_button=tk.Button(master=frame_top_leftrv, text="Planet Parameters", command=rvplanetparameters, padx=10, pady=5, bg=button_bg)
    
    entry_label.grid(row=0, column=0, sticky="nsew", pady=10)
    rv_entry_button.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
    parametersofplanet_button.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)
    parametersofstar_button.grid(row=3, column=0, sticky="nsew", padx=5, pady=(5,10))
    
    #Middle Left Frame --> Should it be varied or not? + The values of the parameters
    parameter_label=tk.Label(master=frame_middle_leftrv, text="Model Parameters", font=("Times New Roman", 10, "bold"), bg=title_bg, width=46)
    parameter_name_label=tk.Label(master=frame_middle_leftrv, text="Parameter", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    parameter_value_label=tk.Label(master=frame_middle_leftrv, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    variability_label=tk.Label(master=frame_middle_leftrv, text="Free/Fixed", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    
    parameter_label.grid(row=0, column=0, columnspan=3, sticky="nsew", padx=10, pady=(10,5))
    parameter_name_label.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
    parameter_value_label.grid(row=1, column=1, sticky="nsew", padx=10, pady=5)
    variability_label.grid(row=1, column=2, sticky="nsew", padx=10, pady=5)
    
    for i in range(len(rvmodel_parameter_names_list)):
        parametername=tk.Label(master=frame_middle_leftrv, text=rvmodel_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=2+i, column=0, sticky="nsew", padx=10)
        
        if len(rvstar_parameter_values_list)==4:
            for i in range(len(rvstar_parameter_names_list)-2):
                starvalue_first=tk.Label(master=frame_middle_leftrv, text=rvstar_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
                starvalue_first.grid(row=2+i, column=1, sticky="nsew")
        
        else:
            for i in range(len(rvstar_parameter_names_list)-2):
                parametervalue_first=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
                parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
        
        if len(rvplanet_parameter_values_list)==6:
            parametervalue1=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[0], bg=secondary_bg, fg=text_color_primary)
            parametervalue1.grid(row=2+2, column=1, sticky="nsew")
            
            parametervalue2=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[1], bg=secondary_bg, fg=text_color_primary)
            parametervalue2.grid(row=2+2+1, column=1, sticky="nsew")
            
            parametervalue3=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[2], bg=secondary_bg, fg=text_color_primary)
            parametervalue3.grid(row=2+2+2, column=1, sticky="nsew")
            
            parametervalue4=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[3], bg=secondary_bg, fg=text_color_primary)
            parametervalue4.grid(row=2+2+3, column=1, sticky="nsew")
        
        else:
            parametervalue1=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue1.grid(row=2+2, column=1, sticky="nsew")
            
            parametervalue2=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue2.grid(row=2+2+1, column=1, sticky="nsew")
            
            parametervalue3=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue3.grid(row=2+2+2, column=1, sticky="nsew")
            
            parametervalue4=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue4.grid(row=2+2+3, column=1, sticky="nsew")
    
    rvv_gama_vary=tk.IntVar()
    rvK_vary=tk.IntVar()
    rvT0_vary=tk.IntVar()
    rvPeriod_vary=tk.IntVar()
    rvEccentricity_vary=tk.IntVar()
    rvArgument_Of_Periapsis_vary=tk.IntVar()

    rvv_gama_vary_check=tk.Checkbutton(master=frame_middle_leftrv, variable=rvv_gama_vary, onvalue=1, offvalue=0, bg=secondary_bg)
    rvK_vary_check=tk.Checkbutton(master=frame_middle_leftrv, variable=rvK_vary, onvalue=1, offvalue=0, bg=secondary_bg)
    rvT0_vary_check=tk.Checkbutton(master=frame_middle_leftrv, variable=rvT0_vary, onvalue=1, offvalue=0, bg=secondary_bg)
    rvPeriod_vary_check=tk.Checkbutton(master=frame_middle_leftrv, variable=rvPeriod_vary, onvalue=1, offvalue=0, bg=secondary_bg)
    rvEccentricity_vary_check=tk.Checkbutton(master=frame_middle_leftrv, variable=rvEccentricity_vary, onvalue=1, offvalue=0, bg=secondary_bg)
    rvArgument_Of_Periapsis_vary_check=tk.Checkbutton(master=frame_middle_leftrv, variable=rvArgument_Of_Periapsis_vary, onvalue=1, offvalue=0, bg=secondary_bg)
    
    rvv_gama_vary_check.grid(row=2, column=2, padx=10)
    rvK_vary_check.grid(row=3, column=2, padx=10)
    rvT0_vary_check.grid(row=4, column=2, padx=10)
    rvPeriod_vary_check.grid(row=5, column=2, padx=10)
    rvEccentricity_vary_check.grid(row=6, column=2, padx=10)
    rvArgument_Of_Periapsis_vary_check.grid(row=7, column=2, padx=10)
    
    #Bottom Left Frame --> Calculated Parameters (Radius, Mass etc.)
    calculated_label=tk.Label(master=frame_bottom_leftrv, text="Calculated Parameters", padx=5, pady=10, bg=title_bg, font=("Times New Roman", 10, "bold"), width=45)
    
    rvcalcparameter_result_label=tk.Label(master=frame_bottom_leftrv, text="Parameter Name", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    rvcalcparameter_name_fit_label=tk.Label(master=frame_bottom_leftrv, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    rvcalcparameter_unit_fit_label=tk.Label(master=frame_bottom_leftrv, text="Unit", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    
    calculated_label.grid(row=0, column=0, columnspan=3, sticky="nsew", padx=10, pady=(10,5))
    rvcalcparameter_result_label.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
    rvcalcparameter_name_fit_label.grid(row=1, column=1, sticky="nsew", padx=10, pady=5)
    rvcalcparameter_unit_fit_label.grid(row=1, column=2, sticky="nsew", padx=10, pady=5)
    
    for i in range(len(rv_result_parameter_names_list)):
        parametername=tk.Label(master=frame_bottom_leftrv, text=rv_result_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=2+i, column=0, sticky="nsew", padx=10, pady=3)
    
        if len(rv_result_parameter_values_list)!=0:
            parametervalue_first=tk.Label(master=frame_bottom_leftrv, text=rv_result_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10, pady=3)
            
        elif len(rv_result_parameter_values_list)==0:
            parametervalue_first=tk.Label(master=frame_bottom_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10, pady=3)
        
        parameterunit=tk.Label(master=frame_bottom_leftrv, text=rv_result_parameter_unit_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parameterunit.grid(row=2+i, column=2, sticky="nsew", padx=10, pady=3)
    
    #Top Right Frame --> Fitting Area
    rvfit_text=tk.Label(master=frame_top_rightrv, text="Radial Velocity Model", font=("Times New Roman", 12, "bold"), bg=title_bg, width=60)
    rvfit_text.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
    rv_fitting=tk.Frame(master=frame_top_rightrv, width=(550), height=(320), highlightbackground="red", highlightthickness=0, bg=main_bg)
    rv_fitting.grid(row=1, column=1)
    
    def rv_phase_show():  
        if(len(obs_time_list)==0 or len(v_radial_list)==0):
            rv_show_error=tk.messagebox.showerror("Error", "Please enter the Radial Velocity data to proceed.")
        
        elif len(rvplanet_parameter_values_list)!=6:
            rv_show_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
        
        elif len(rvstar_parameter_values_list)!=4:
            rv_show_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
        
        graphtitle="Radial Velocity"
        
        if len(preferences_list)==12:
            graphtitle=preferences_list[9]
            data_bg=preferences_list[1]
            main_bg=preferences_list[3]
            tetriary_bg=preferences_list[4]
            text_color_primary_label=preferences_list[5]
            text_color_primary_axis=preferences_list[6]
            sys_vel_color=preferences_list[7]
            err_bar_color=preferences_list[8]
            data_style=preferences_list[10]
            line_style=preferences_list[11]
            
        elif dl_button['text']=="Light Mode":
            main_bg="#0B0B0B"
            secondary_bg="#252525"
            tetriary_bg="#3A3A3A"
            title_bg="#FF7A37"
            data_bg="red"
            fit_color="blue"
            data_style="."
            line_style="-"
            
            button_bg="#f0e130"
            
            text_color_primary_label="white"
            text_color_primary_axis="white"
            text_color_secondary="black"
            
            color_frame_bg="#0C0B0B"
            
            sys_vel_color="white"
            err_bar_color="orange"
            
        elif dl_button['text']=="Dark Mode":
            main_bg="#ECDBDB"
            secondary_bg="#FFFFFF"
            tetriary_bg="#FFFFFF"
            title_bg="#7A7A7A"
            data_bg="red"
            fit_color="blue"
            data_style="."
            line_style="-"
        
            button_bg="#C9C6C6"
            
            text_color_primary_label="black"
            text_color_primary_axis="black"
            text_color_secondary="black"
        
            color_frame_bg="#0C0B0B"
            
            sys_vel_color="black"
            err_bar_color="orange"
        
        T0=rvplanet_parameter_values_list[0]
        Period=rvplanet_parameter_values_list[1]
        v_gama=rvstar_parameter_values_list[0]
        
        def Epoch_func(obs_time_list, T0, Period):
            rvEpoch_List=[] #Phase List
            rvT_T0_values=[] #T-T0 values list
    
            for time in obs_time_list:
                if ((float(time)-T0)/Period)<0:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
            
                else:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360

                rvEpoch_List.append(Epoch_value)
            
                T_T0=time-T0
                rvT_T0_values.append(T_T0)
            
            return rvEpoch_List, rvT_T0_values
        
        rvEpoch_List, rvT_T0_values=Epoch_func(obs_time_list, T0, Period)
        
        for widgets in rv_fitting.winfo_children():
            widgets.destroy()
        
        rv_fig=Figure(figsize=(6.0,3.2), dpi=90)
        rv_fig.patch.set_facecolor(main_bg)
        
        rv_fit=rv_fig.add_subplot(111)
        rv_fit.plot(rvEpoch_List, v_radial_list, data_style, color=data_bg)
        rv_fit.plot([0,1], [v_gama, v_gama], '--', color=sys_vel_color)
        rv_fit.patch.set_facecolor(tetriary_bg)
        rv_fig.subplots_adjust(left=0.19, bottom=0.177, top=0.895, right=0.958)
        
        rv_fit.xaxis.label.set_color(text_color_primary_label)
        rv_fit.yaxis.label.set_color(text_color_primary_label)
        
        rv_fit.tick_params(axis='x', colors=text_color_primary_axis)
        rv_fit.tick_params(axis='y', colors=text_color_primary_axis)
        
        rv_fit.spines['left'].set_color(text_color_primary_axis)
        rv_fit.spines['bottom'].set_color(text_color_primary_axis)
        rv_fit.spines['right'].set_color(text_color_primary_axis)
        rv_fit.spines['top'].set_color(text_color_primary_axis)
        
        rv_fit.set_title(graphtitle, fontsize=10, color=text_color_primary_label, fontweight="bold")
        rv_fit.set_ylabel("Radial Velocity (km/sec)", fontsize=10)
        rv_fit.set_xlabel("Phase", fontsize=10)
        
        plot_box=FigureCanvasTkAgg(rv_fig, rv_fitting)
        plot_box.draw()
        plot_box.get_tk_widget().pack()
        
        toolbar=NavigationToolbar2Tk(plot_box, rv_fitting)
        toolbar.update()
        plot_box.get_tk_widget().pack()
    
    def rv_fit_first_show():
        graphtitle="Radial Velocity"
        
        if len(preferences_list)==12:
            graphtitle=preferences_list[0]
            data_bg=preferences_list[1]
            fit_color=preferences_list[2]
            main_bg=preferences_list[3]
            tetriary_bg=preferences_list[4]
            text_color_primary_label=preferences_list[5]
            text_color_primary_axis=preferences_list[6]
            sys_vel_color=preferences_list[7]
            err_bar_color=preferences_list[8]
            data_style=preferences_list[10]
            line_style=preferences_list[11]
        
            if dl_button['text']=="Light Mode":
                lcpmain_bg="#0B0B0B"
                lcpsecondary_bg="#252525"
                text_color_primary="white"
                lcpbutton_bg="#f0e130"
                
                secondary_bg="#252525"
            
            elif dl_button['text']=="Dark Mode":
                lcpmain_bg="#ECDBDB"
                lcpsecondary_bg="#FFFFFF"
                text_color_primary="black"
                lcpbutton_bg="#C9C6C6"
                
                secondary_bg="#FFFFFF"
        
        elif dl_button['text']=="Light Mode":
            main_bg="#0B0B0B"
            secondary_bg="#252525"
            tetriary_bg="#3A3A3A"
            title_bg="#FF7A37"
            data_bg="red"
            fit_color="blue"
            data_style="."
            line_style="-"
            
            button_bg="#f0e130"
            lcpbutton_bg="#f0e130"
            
            text_color_primary_label="white"
            text_color_primary_axis="white"
            text_color_secondary="black"
            
            color_frame_bg="#0C0B0B"
            
            lcpmain_bg="#0B0B0B"
            lcpsecondary_bg="#252525"
            text_color_primary="white"
            
            sys_vel_color="white"
            err_bar_color="orange"
        
        elif dl_button['text']=="Dark Mode":
            main_bg="#ECDBDB"
            secondary_bg="#FFFFFF"
            tetriary_bg="#FFFFFF"
            title_bg="#7A7A7A"
            data_bg="red"
            fit_color="blue"
            data_style="."
            line_style="-"
        
            button_bg="#C9C6C6"
            lcpbutton_bg="#C9C6C6"
            
            text_color_primary_label="black"
            text_color_primary_axis="black"
            text_color_secondary="black"
        
            color_frame_bg="#0C0B0B"
            
            lcpmain_bg="#ECDBDB"
            lcpsecondary_bg="#FFFFFF"
            text_color_primary="black"
            
            sys_vel_color="black"
            err_bar_color="orange"
        
        #Define Progress
        def progress_stage():
            return f"Current Progress: %{rvprogressbar['value']}"
        
        def progress():
            if rvprogressbar['value']<100:
                rvprogressbar['value']+=10
                rvprogresslabel['text']=progress_stage()
            
            if rvprogressbar['value']==100:
                progressfinish=tk.messagebox.showinfo("Successful", "The Light Curve Modelling is finished!")
                rvp_window.destroy()
        
        #Define Pop-up Progress Bar
        rvp_window=tk.Toplevel(master=rv_window)
        rvp_window.title("Radial Velocity Modelling Progress")
        rvp_window.iconbitmap(r'icon_lcf.ico')
        rvp_window.configure(bg=lcpmain_bg)
        rvp_window.wm_transient(rv_window)
        rvp_window.resizable(width=False, height=False)
        
        window_width=rvp_window.winfo_reqwidth()
        window_height=rvp_window.winfo_reqheight()

        screen_width=rvp_window.winfo_screenwidth()
        screen_height=rvp_window.winfo_screenheight()

        x_cordinate = int((screen_width/2) - (window_width/2))
        y_cordinate = int((screen_height/2) - (window_height/2))
        
        rvp_window.geometry("+{}+{}".format(x_cordinate, y_cordinate))
        
        rvinprogress=tk.Label(master=rvp_window, text="Radial Velocity Modelling in Progress...\n(Progress may take up to a few minutes)", bg=lcpsecondary_bg, fg=text_color_primary)
        rvinprogress.grid(row=0, column=0, pady=10, padx=10)
        
        rvprogressbar=ttk.Progressbar(master=rvp_window, orient="horizontal", mode="determinate", length=300)
        rvprogressbar.grid(row=1, column=0, pady=10, padx=10)
        
        rvprogresslabel=tk.Label(master=rvp_window, text=progress_stage(), bg=lcpbutton_bg)
        rvprogresslabel.grid(row=2, column=0, pady=10, padx=10)
        
        #Define progressbar value
        rvprogressbar['value']=0
        
        if(len(obs_time_list)==0 or len(v_radial_list)==0):
            rv_show_error=tk.messagebox.showerror("Error", "Please enter the Radial Velocity data to proceed.")
            rvp_window.destroy()
        
        elif len(rvplanet_parameter_values_list)!=6:
            rv_show_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
            rvp_window.destroy()
        
        elif len(rvstar_parameter_values_list)!=4:
            rv_show_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
            rvp_window.destroy()
        
        rvmodel_parameter_values_list.clear()
        
        rvmodel_parameter_values_list.append(rvstar_parameter_values_list[0])
        rvmodel_parameter_values_list.append(rvstar_parameter_values_list[1])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[0])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[1])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[2])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[3])
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        T0=rvplanet_parameter_values_list[0]
        Period=rvplanet_parameter_values_list[1]
        Eccentricity=rvplanet_parameter_values_list[2]
        Argument_Of_Periapsis=rvplanet_parameter_values_list[3]
        inclination=rvplanet_parameter_values_list[4]
        
        v_gama=rvstar_parameter_values_list[0]
        K=rvstar_parameter_values_list[1]
        
        def Epoch_func(obs_time_list, T0, Period):
            rvEpoch_List=[] #Phase List
            rvT_T0_values=[] #T-T0 values list
            
            sin_wave_time_data=[0]
            sinEpoch_List=[]
            sinT_T0_values=[]
            
            sintime=0
            while sintime<Period:
                sintime+=float(Period/100)
                sin_wave_time_data.append(sintime)
            
            for time in sin_wave_time_data:
                if ((float(time)-T0)/Period)<0:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
                
                else:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360
        
                sinEpoch_List.append(Epoch_value)
                
                T_T0=time-T0
                sinT_T0_values.append(T_T0)
        
            for time in obs_time_list:
                if ((float(time)-T0)/Period)<0:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
            
                else:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360

                rvEpoch_List.append(Epoch_value)
            
                T_T0=time-T0
                rvT_T0_values.append(T_T0)
            
            return rvEpoch_List, rvT_T0_values, sinEpoch_List, sinT_T0_values
        
        rvEpoch_List, rvT_T0_values, sinEpoch_List, sinT_T0_values=Epoch_func(obs_time_list, T0, Period)
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        def Anomalies(obs_time_list, T0, Period, Eccentricity, Argument_Of_Periapsis):
            rvMean_Anomaly_values, rvEccentric_Anomaly_values, rvTrue_Anomaly_values=[], [], []
            Mean_Anomaly_values_sin, Eccentric_Anomaly_values_sin, True_Anomaly_values_sin=[], [], []
            
            rvEpoch_List, rvT_T0_values, sinEpoch_List, sinT_T0_values=Epoch_func(obs_time_list, T0, Period)
            
            for time in sinEpoch_List:    
                Mean_Anomaly=((2*math.pi))*(time)
                Mean_Anomaly_values_sin.append((Mean_Anomaly)) #M değerleri Radyan Biriminde
                    
                Ei_old=Mean_Anomaly #Radyan Biriminde
                Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old)) #Radyan Biriminde
                error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                while error>0.001:
                    Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old))
                    error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                    Ei_old=Ei_new
            
                Eccentric_Anomaly_values_sin.append(Ei_new) #E değerleri Radyan Biriminde
                
                True_Anomaly_0=((math.pi/2)-Argument_Of_Periapsis) #ν değerleri Radyan Biriminde
            
                True_Anomaly=2*math.atan(math.sqrt((1+Eccentricity)/(1-Eccentricity))*math.tan((Ei_new)/2))
                if True_Anomaly<0:
                    True_Anomaly+=2*math.pi
                
                True_Anomaly_values_sin.append(True_Anomaly) #ν değerleri Radyan Biriminde
        
            for T_T0 in rvT_T0_values:    
                Mean_Anomaly=((2*math.pi)/Period)*(T_T0)
                rvMean_Anomaly_values.append((Mean_Anomaly)) #M değerleri Radyan Biriminde
                    
                Ei_old=Mean_Anomaly #Radyan Biriminde
                Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old)) #Radyan Biriminde
                error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                while error>0.001:
                    Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old))
                    error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                    Ei_old=Ei_new
            
                rvEccentric_Anomaly_values.append(Ei_new) #E değerleri Radyan Biriminde
                
                True_Anomaly_0=((math.pi/2)-Argument_Of_Periapsis) #ν değerleri Radyan Biriminde
            
                True_Anomaly=2*math.atan(math.sqrt((1+Eccentricity)/(1-Eccentricity))*math.tan((Ei_new)/2))
                if True_Anomaly<0:
                    True_Anomaly+=2*math.pi
                
                rvTrue_Anomaly_values.append(True_Anomaly) #ν değerleri Radyan Biriminde
            
            return rvMean_Anomaly_values, rvEccentric_Anomaly_values, rvTrue_Anomaly_values, True_Anomaly_values_sin
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        def vr_fit_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis):
            Mean_Anomaly_values, Eccentric_Anomaly_values, True_Anomaly_values, True_Anomaly_values_sin=Anomalies(obs_time_list, T0, Period, Eccentricity, Argument_Of_Periapsis)

            vr_fit=[]
            
            for i in range(len(True_Anomaly_values)):
                vr_fit_val=v_gama+(K*((Eccentricity*math.cos(Argument_Of_Periapsis))+(math.cos(True_Anomaly_values[i]+Argument_Of_Periapsis))))
                vr_fit.append(vr_fit_val)
            
            return vr_fit
        
        def vr_fit_sin_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis):
            Mean_Anomaly_values, Eccentric_Anomaly_values, True_Anomaly_values, True_Anomaly_values_sin=Anomalies(obs_time_list, T0, Period, Eccentricity, Argument_Of_Periapsis)
            
            vr_fit_sin_list=[]
            
            for i in range(len(True_Anomaly_values_sin)):
                vr_fit_sin=v_gama+(K*((Eccentricity*math.cos(Argument_Of_Periapsis))+(math.cos(True_Anomaly_values_sin[i]+Argument_Of_Periapsis))))
                vr_fit_sin_list.append(vr_fit_sin)
            
            return vr_fit_sin_list
        
        vr_fit=vr_fit_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis)
        vr_fit_sin_list=vr_fit_sin_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis)
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        rv_result_parameter_values_list.clear()
        
        M1=rv_mass_star[0]
        
        mass_func=((1.0385e-7)*(pow((1-pow(Eccentricity,2)),3/2))*pow(K,3)*Period)/(pow(math.sin(inclination),3)) #km

        M2_old=0
        M2_new=(1/1000)*13
        
        while abs(M2_old-M2_new)>0.0001:
            M2_old=M2_new
            
            f=(pow(M2_old,3))/(pow((M1+M2_old),2))-mass_func
            f_der=((pow(M2_old,2))*(M2_old+(3*M1)))/(pow((M1+M2_old),3))
        
            M2_new=M2_old-(f/f_der)
        
        M2=M2_new
        
        M_sun=rv_constant_values[0]
        M_jup=rv_constant_values[1]
        
        R_sun=rv_constant_values[2]
        R_jup=rv_constant_values[3]
        
        G=rv_constant_values[4] #cgs unit
        
        a1=(13751*K*Period*math.sqrt(1-pow(Eccentricity,2)))/math.sin(inclination) #km
        a2=(a1*M1)/M2 #km
        a=a1+a2 #km
        
        r_star=rvstar_parameter_values_list[2]
        r_planet=rvplanet_parameter_values_list[5]
        
        R_star=r_star*a
        R_planet=r_planet*a
        
        M1_g=(M1*M_sun)*1000 #g
        M2_g=(M2*M_sun)*1000 #g

        R_star_cm=R_star*(1000*100) #cm
        R_planet_cm=R_planet*(1000*100) #cm

        g_star=G*(M1_g/pow(R_star_cm,2))
        g_planet=G*(M2_g/pow(R_planet_cm,2))
        
        rv_result_parameter_values_list.append(round(M1,3)) #M Sun
        rv_result_parameter_values_list.append(round(((M2*M_sun)/M_jup),3)) #M Jup
        rv_result_parameter_values_list.append(round((a2/149597870.7),3)) #AU
        rv_result_parameter_values_list.append(round((R_planet/R_jup),3)) #R Jup
        rv_result_parameter_values_list.append(round(g_planet,3)) #cm/sn^2
        rv_result_parameter_values_list.append(round(math.log10(g_planet),3)) #unitless
        rv_result_parameter_values_list.append(round((R_star/R_sun),3)) #R Sun
        rv_result_parameter_values_list.append(round(g_star,3)) #cm/sn^2
        rv_result_parameter_values_list.append(round(math.log10(g_star),3)) #unitless
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        for i in range(len(rvmodel_parameter_names_list)):
            parameter_value=tk.Entry(master=frame_middle_rightrv)
            parameter_value.insert(tk.END, rvmodel_parameter_values_list[i])
            parameter_value.grid(row=2+i, column=1, sticky="nsew", padx=10)
            parameter_value.configure(state="readonly")
    
        for i in range(len(rvmodel_parameter_names_list)):
            fit_value=tk.Entry(master=frame_middle_rightrv)
            fit_value.insert(tk.END, rvmodel_parameter_values_list[i])
            fit_value.grid(row=2+i, column=2, sticky="nsew", padx=10)
            fit_value.configure(state="readonly")
    
        for i in range(len(rvmodel_parameter_names_list)):
            fit_error=tk.Entry(master=frame_middle_rightrv)
            fit_error.insert(tk.END, f"\u00B1 None")
            fit_error.grid(row=2+i, column=3, sticky="nsew", padx=10)
            fit_error.configure(state="readonly")
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        rvmodel_parameters_update.rvmodel_parameter_fit_value_list=rvmodel_parameter_values_list.copy()
        rvmodel_parameters_update.rvmodel_parameter_error_value_list=["\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None"]
    
        v_gama=rvstar_parameter_values_list[0]
        
        Epoch_List_backup=rvEpoch_List.copy()
        v_radial_list_backup=v_radial_list.copy()
        v_error_list_backup=v_error_list.copy()
        Epoch_List_sin_backup=sinEpoch_List.copy()
        vr_fit_sin_list_backup=vr_fit_sin_list.copy()

        Epoch_List_sorted, v_radial_list_sorted, v_error_list_sorted, Epoch_List_sin_sorted, vr_fit_sin_list_sorted = [], [], [], [], []

        for i in range(len(v_radial_list)):
            min_val=min(Epoch_List_backup) #Minimum zaman
            min_index=Epoch_List_backup.index(min_val) #Minimum zamanın index değeri
            
            Epoch_List_sorted.append(min_val)
            Epoch_List_backup.pop(min_index)
            
            v_radial_list_sorted.append(v_radial_list_backup[min_index])
            v_radial_list_backup.pop(min_index)
            
            v_error_list_sorted.append(v_error_list_backup[min_index])
            v_error_list_backup.pop(min_index)
        
        for i in range(len(sinEpoch_List)):
            min_val=min(Epoch_List_sin_backup) #Minimum zaman
            min_index=Epoch_List_sin_backup.index(min_val)
            
            Epoch_List_sin_sorted.append(Epoch_List_sin_backup[min_index])
            Epoch_List_sin_backup.pop(min_index)
            
            vr_fit_sin_list_sorted.append(vr_fit_sin_list_backup[min_index])
            vr_fit_sin_list_backup.pop(min_index)
        
        for widgets in rv_fitting.winfo_children():
            widgets.destroy()
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        rv_fig=Figure(figsize=(6.0,3.2), dpi=90)
        rv_fig.patch.set_facecolor(main_bg)
        
        rv_fit=rv_fig.add_subplot(111)
        rv_fit.plot(Epoch_List_sorted, v_radial_list_sorted, data_style, color=data_bg)
        rv_fit.plot(Epoch_List_sin_sorted, vr_fit_sin_list_sorted, line_style, color=fit_color)
        rv_fit.plot([0,1], [v_gama, v_gama], '--', color=sys_vel_color)
        rv_fit.patch.set_facecolor(tetriary_bg)
        rv_fig.subplots_adjust(left=0.19, bottom=0.177, top=0.895, right=0.958)
        
        rv_fit.xaxis.label.set_color(text_color_primary_label)
        rv_fit.yaxis.label.set_color(text_color_primary_label)
        
        rv_fit.tick_params(axis='x', colors=text_color_primary_axis)
        rv_fit.tick_params(axis='y', colors=text_color_primary_axis)
        
        rv_fit.spines['left'].set_color(text_color_primary_axis)
        rv_fit.spines['bottom'].set_color(text_color_primary_axis)
        rv_fit.spines['right'].set_color(text_color_primary_axis)
        rv_fit.spines['top'].set_color(text_color_primary_axis)
        
        rv_fit.set_title(graphtitle, fontsize=10, color=text_color_primary_label, fontweight="bold")
        rv_fit.set_ylabel("Radial Velocity (km/sec)", fontsize=10)
        rv_fit.set_xlabel("Phase", fontsize=10)
        
        plot_box=FigureCanvasTkAgg(rv_fig, rv_fitting)
        plot_box.draw()
        plot_box.get_tk_widget().pack()
        
        toolbar=NavigationToolbar2Tk(plot_box, rv_fitting)
        toolbar.update()
        plot_box.get_tk_widget().pack()
        
        for i in range(len(rv_result_parameter_values_list)):
            parametervalue_first=tk.Label(master=frame_bottom_leftrv, text=rv_result_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
        
        progress()
        rv_window.update()
        rvp_window.update()
    
    def rv_fit_improved_show():
        graphtitle="Radial Velocity"
        
        if len(preferences_list)==12:
            graphtitle=preferences_list[0]
            data_bg=preferences_list[1]
            fit_color=preferences_list[2]
            main_bg=preferences_list[3]
            tetriary_bg=preferences_list[4]
            text_color_primary_label=preferences_list[5]
            text_color_primary_axis=preferences_list[6]
            sys_vel_color=preferences_list[7]
            err_bar_color=preferences_list[8]
            data_style=preferences_list[10]
            line_style=preferences_list[11]
        
            if dl_button['text']=="Light Mode":
                lcpmain_bg="#0B0B0B"
                lcpsecondary_bg="#252525"
                text_color_primary="white"
                lcpbutton_bg="#f0e130"
                
                secondary_bg="#252525"
            
            elif dl_button['text']=="Dark Mode":
                lcpmain_bg="#ECDBDB"
                lcpsecondary_bg="#FFFFFF"
                text_color_primary="black"
                lcpbutton_bg="#C9C6C6"
                
                secondary_bg="#FFFFFF"
        
        elif dl_button['text']=="Light Mode":
            main_bg="#0B0B0B"
            secondary_bg="#252525"
            tetriary_bg="#3A3A3A"
            title_bg="#FF7A37"
            data_bg="red"
            fit_color="blue"
            data_style="."
            line_style="-"
            
            button_bg="#f0e130"
            lcpbutton_bg="#f0e130"
            
            text_color_primary_label="white"
            text_color_primary_axis="white"
            text_color_secondary="black"
            
            color_frame_bg="#0C0B0B"
            
            lcpmain_bg="#0B0B0B"
            lcpsecondary_bg="#252525"
            text_color_primary="white"
            
            sys_vel_color="white"
            err_bar_color="orange"
        
        elif dl_button['text']=="Dark Mode":
            main_bg="#ECDBDB"
            secondary_bg="#FFFFFF"
            tetriary_bg="#FFFFFF"
            title_bg="#7A7A7A"
            data_bg="red"
            fit_color="blue"
            data_style="."
            line_style="-"
        
            button_bg="#C9C6C6"
            lcpbutton_bg="#C9C6C6"
            
            text_color_primary_label="black"
            text_color_primary_axis="black"
            text_color_secondary="black"
        
            color_frame_bg="#0C0B0B"
            
            lcpmain_bg="#ECDBDB"
            lcpsecondary_bg="#FFFFFF"
            text_color_primary="black"
            
            sys_vel_color="black"
            err_bar_color="orange"
        
        #Define Progress
        def progress_stage():
            return f"Current Progress: %{rvprogressbar['value']}"
        
        def progress():
            if rvprogressbar['value']<100:
                rvprogressbar['value']+=10
                rvprogresslabel['text']=progress_stage()
            
            if rvprogressbar['value']==100:
                progressfinish=tk.messagebox.showinfo("Successful", "The Light Curve Modelling is finished!")
                rvp_window.destroy()
        
        #Define Pop-up Progress Bar
        rvp_window=tk.Toplevel(master=rv_window)
        rvp_window.title("Radial Velocity Modelling Progress")
        rvp_window.iconbitmap(r'icon_lcf.ico')
        rvp_window.configure(bg=lcpmain_bg)
        rvp_window.wm_transient(rv_window)
        rvp_window.resizable(width=False, height=False)
        
        window_width=rvp_window.winfo_reqwidth()
        window_height=rvp_window.winfo_reqheight()

        screen_width=rvp_window.winfo_screenwidth()
        screen_height=rvp_window.winfo_screenheight()

        x_cordinate = int((screen_width/2) - (window_width/2))
        y_cordinate = int((screen_height/2) - (window_height/2))
        
        rvp_window.geometry("+{}+{}".format(x_cordinate, y_cordinate))
        
        rvinprogress=tk.Label(master=rvp_window, text="Radial Velocity Modelling in Progress...\n(Progress may take up to a few minutes)", bg=lcpsecondary_bg, fg=text_color_primary)
        rvinprogress.grid(row=0, column=0, pady=10, padx=10)
        
        rvprogressbar=ttk.Progressbar(master=rvp_window, orient="horizontal", mode="determinate", length=300)
        rvprogressbar.grid(row=1, column=0, pady=10, padx=10)
        
        rvprogresslabel=tk.Label(master=rvp_window, text=progress_stage(), bg=lcpbutton_bg)
        rvprogresslabel.grid(row=2, column=0, pady=10, padx=10)
        
        #Define progressbar value
        rvprogressbar['value']=0
        
        if(len(obs_time_list)==0 or len(v_radial_list)==0):
            rv_show_error=tk.messagebox.showerror("Error", "Please enter the Radial Velocity data to proceed.")
            rvp_window.destroy()
        
        elif len(rvplanet_parameter_values_list)!=6:
            rv_show_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
            rvp_window.destroy()
            
        elif len(rvstar_parameter_values_list)!=4:
            rv_show_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
            rvp_window.destroy()
        
        rvmodel_parameter_values_list.clear()

        rvmodel_parameter_values_list.append(rvstar_parameter_values_list[0])
        rvmodel_parameter_values_list.append(rvstar_parameter_values_list[1])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[0])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[1])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[2])
        rvmodel_parameter_values_list.append(rvplanet_parameter_values_list[3])
        
        T0=rvplanet_parameter_values_list[0]
        Period=rvplanet_parameter_values_list[1]
        Eccentricity=rvplanet_parameter_values_list[2]
        Argument_Of_Periapsis=rvplanet_parameter_values_list[3]
        inclination=rvplanet_parameter_values_list[4]
        
        v_gama=rvstar_parameter_values_list[0]
        K=rvstar_parameter_values_list[1]
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        def Epoch_func(obs_time_list, T0, Period):
            rvEpoch_List=[] #Phase List
            rvT_T0_values=[] #T-T0 values list
            
            sin_wave_time_data=[0]
            sinEpoch_List=[]
            sinT_T0_values=[]
            
            sintime=0
            while sintime<Period:
                sintime+=float(Period/100)
                sin_wave_time_data.append(sintime)
            
            for time in sin_wave_time_data:
                if ((float(time)-T0)/Period)<0:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
                
                else:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360
        
                sinEpoch_List.append(Epoch_value)
                
                T_T0=time-T0
                sinT_T0_values.append(T_T0)
        
            for time in obs_time_list:
                if ((float(time)-T0)/Period)<0:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
            
                else:
                    Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                    Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360

                rvEpoch_List.append(Epoch_value)
            
                T_T0=time-T0
                rvT_T0_values.append(T_T0)
            
            return rvEpoch_List, rvT_T0_values, sinEpoch_List, sinT_T0_values
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        def Anomalies(obs_time_list, T0, Period, Eccentricity, Argument_Of_Periapsis):
            rvMean_Anomaly_values, rvEccentric_Anomaly_values, rvTrue_Anomaly_values=[], [], []
            Mean_Anomaly_values_sin, Eccentric_Anomaly_values_sin, True_Anomaly_values_sin=[], [], []
            
            rvEpoch_List, rvT_T0_values, sinEpoch_List, sinT_T0_values=Epoch_func(obs_time_list, T0, Period)
            
            for time in sinEpoch_List:    
                Mean_Anomaly=((2*math.pi))*(time)
                Mean_Anomaly_values_sin.append((Mean_Anomaly)) #M değerleri Radyan Biriminde
                    
                Ei_old=Mean_Anomaly #Radyan Biriminde
                Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old)) #Radyan Biriminde
                error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                while error>0.001:
                    Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old))
                    error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                    Ei_old=Ei_new
            
                Eccentric_Anomaly_values_sin.append(Ei_new) #E değerleri Radyan Biriminde
                
                True_Anomaly_0=((math.pi/2)-Argument_Of_Periapsis) #ν değerleri Radyan Biriminde
            
                True_Anomaly=2*math.atan(math.sqrt((1+Eccentricity)/(1-Eccentricity))*math.tan((Ei_new)/2))
                if True_Anomaly<0:
                    True_Anomaly+=2*math.pi
                
                True_Anomaly_values_sin.append(True_Anomaly) #ν değerleri Radyan Biriminde
        
            for T_T0 in rvT_T0_values:    
                Mean_Anomaly=((2*math.pi)/Period)*(T_T0)
                rvMean_Anomaly_values.append((Mean_Anomaly)) #M değerleri Radyan Biriminde
                    
                Ei_old=Mean_Anomaly #Radyan Biriminde
                Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old)) #Radyan Biriminde
                error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                while error>0.001:
                    Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old))
                    error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                    Ei_old=Ei_new
            
                rvEccentric_Anomaly_values.append(Ei_new) #E değerleri Radyan Biriminde
                
                True_Anomaly_0=((math.pi/2)-Argument_Of_Periapsis) #ν değerleri Radyan Biriminde
            
                True_Anomaly=2*math.atan(math.sqrt((1+Eccentricity)/(1-Eccentricity))*math.tan((Ei_new)/2))
                if True_Anomaly<0:
                    True_Anomaly+=2*math.pi
                
                rvTrue_Anomaly_values.append(True_Anomaly) #ν değerleri Radyan Biriminde
            
            return rvMean_Anomaly_values, rvEccentric_Anomaly_values, rvTrue_Anomaly_values, True_Anomaly_values_sin
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        def vr_fit_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis):
            Mean_Anomaly_values, Eccentric_Anomaly_values, True_Anomaly_values, True_Anomaly_values_sin=Anomalies(obs_time_list, T0, Period, Eccentricity, Argument_Of_Periapsis)

            vr_fit=[]
            
            for i in range(len(True_Anomaly_values)):
                vr_fit_val=v_gama+(K*((Eccentricity*math.cos(Argument_Of_Periapsis))+(math.cos(True_Anomaly_values[i]+Argument_Of_Periapsis))))
                vr_fit.append(vr_fit_val)
            
            return vr_fit
        
        def vr_fit_sin_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis):
            Mean_Anomaly_values, Eccentric_Anomaly_values, True_Anomaly_values, True_Anomaly_values_sin=Anomalies(obs_time_list, T0, Period, Eccentricity, Argument_Of_Periapsis)
            
            vr_fit_sin_list=[]
            
            for i in range(len(True_Anomaly_values_sin)):
                vr_fit_sin=v_gama+(K*((Eccentricity*math.cos(Argument_Of_Periapsis))+(math.cos(True_Anomaly_values_sin[i]+Argument_Of_Periapsis))))
                vr_fit_sin_list.append(vr_fit_sin)
            
            return vr_fit_sin_list
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        vr_fit=vr_fit_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis)
        vr_fit_sin_list=vr_fit_sin_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis)
        
        from lmfit import Model, minimize, Parameters, report_fit, fit_report
        
        #Parametreler oluşturulur.
        RV_Parameters_Fit=Parameters()
        
        #Modeli oluşturuyoruz.
        RV_Model=Model(vr_fit_func)
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        #Oluşturduğumuz Modeldeki Değişkenleri ve Parametreleri print ediyoruz.
        print(80*"*") #Sonucun güzel gözükmesi açısından eklendi.
        print(f"Parameters: {RV_Model.param_names}")
        print(f"Variables: {RV_Model.independent_vars}")
        
        print(80*"*") #Sonucun güzel gözükmesi açısından eklendi.
        
        #Şimdi de bu Parametreleri sınırlayabilmek için ve ilk değerlerini koyabilmek için tanımlıyoruz.
        Parameters=RV_Model.make_params()
        
        #Parametrelerin sınırları yerine koyulur. (vary=False yapılırsa o parametre serbest bırakılmayacaktır!!!!!!)
        Parameters.add("v_gama", value=v_gama, min=-500, max=500, vary=rvv_gama_vary.get())
        Parameters.add("K", value=K, min=0.000001, max=np.inf, vary=rvK_vary.get())
        Parameters.add("T0", value=T0, min=0.000001, max=np.inf, vary=rvT0_vary.get())
        Parameters.add("Period", value=Period, min=0.000001, max=np.inf, vary=rvPeriod_vary.get())
        Parameters.add("Eccentricity", value=Eccentricity, min=0, max=0.99999, vary=rvEccentricity_vary.get())
        Parameters.add("Argument_Of_Periapsis", value=Argument_Of_Periapsis, min=0, max=2*math.pi, vary=rvArgument_Of_Periapsis_vary.get())
        
        #Fit yapılır.
        fit_result=RV_Model.fit(v_radial_list, Parameters, obs_time_list=obs_time_list, weights=weights)
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        #Sonuç yazılır.
        fit_values=[]
        fit_errors=[]
        fit_corrs=[]
        
        fit_values.append(fit_result.params["v_gama"].value)
        fit_values.append(fit_result.params["K"].value)
        fit_values.append(fit_result.params["T0"].value)
        fit_values.append(fit_result.params["Period"].value)
        fit_values.append(fit_result.params["Eccentricity"].value)
        fit_values.append(fit_result.params["Argument_Of_Periapsis"].value)
        
        fit_errors.append(fit_result.params["v_gama"].stderr)
        fit_errors.append(fit_result.params["K"].stderr)
        fit_errors.append(fit_result.params["T0"].stderr)
        fit_errors.append(fit_result.params["Period"].stderr)
        fit_errors.append(fit_result.params["Eccentricity"].stderr)
        fit_errors.append(fit_result.params["Argument_Of_Periapsis"].stderr)
        
        fit_corrs.append(fit_result.params["v_gama"].correl)
        fit_corrs.append(fit_result.params["K"].correl)
        fit_corrs.append(fit_result.params["T0"].correl)
        fit_corrs.append(fit_result.params["Period"].correl)
        fit_corrs.append(fit_result.params["Eccentricity"].correl)
        fit_corrs.append(fit_result.params["Argument_Of_Periapsis"].correl)
        
        rvmodel_parameters_update.rvmodel_parameter_fit_value_list=fit_values.copy()
        rvmodel_parameters_update.rvmodel_parameter_error_value_list=fit_errors.copy()
    
        #Düzeltilen Fit değerleri bir parametre olarak tanımlanır.
        Fit_finalized=fit_result.best_fit
        
        #Gerekli olan Phase değerleri elimizde mevcuttur.
        rvEpoch_List, rvT_T0_values, sinEpoch_List, sinT_T0_values=Epoch_func(obs_time_list, T0, Period)
        
        #Fit değerleri bir listeye çekilir.
        vr_fit=vr_fit_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis)
        vr_fit_sin_list=vr_fit_sin_func(obs_time_list, v_gama, K, T0, Period, Eccentricity, Argument_Of_Periapsis)

        Epoch_List_backup=rvEpoch_List.copy()
        Fit_finalized_backup=Fit_finalized.copy()
        v_radial_list_backup=v_radial_list.copy()
        v_error_list_backup=v_error_list.copy()
        Epoch_List_sin_backup=sinEpoch_List.copy()
        vr_fit_sin_list_backup=vr_fit_sin_list.copy()
        
        Fit_finalized_sorted, Epoch_List_sorted, v_radial_list_sorted, v_error_list_sorted, Epoch_List_sin_sorted, vr_fit_sin_list_sorted = [], [], [], [], [], []
        
        for i in range(len(v_radial_list)):
            min_val=min(Epoch_List_backup) #Minimum zaman
            min_index=Epoch_List_backup.index(min_val) #Minimum zamanın index değeri
            
            Epoch_List_sorted.append(min_val)
            Epoch_List_backup.pop(min_index)
            
            Fit_finalized_sorted.append(Fit_finalized_backup[min_index])
            Fit_finalized_backup.pop(min_index)
            
            v_radial_list_sorted.append(v_radial_list_backup[min_index])
            v_radial_list_backup.pop(min_index)
            
            v_error_list_sorted.append(v_error_list_backup[min_index])
            v_error_list_backup.pop(min_index)
        
        for i in range(len(sinEpoch_List)):
            min_val=min(Epoch_List_sin_backup) #Minimum zaman
            min_index=Epoch_List_sin_backup.index(min_val)
            
            Epoch_List_sin_sorted.append(Epoch_List_sin_backup[min_index])
            Epoch_List_sin_backup.pop(min_index)
            
            vr_fit_sin_list_sorted.append(vr_fit_sin_list_backup[min_index])
            vr_fit_sin_list_backup.pop(min_index)

        rv_result_parameter_values_list.clear()
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        M1=rv_mass_star[0]
        
        mass_func=((1.0385e-7)*(pow((1-pow(fit_values[4],2)),3/2))*pow(fit_values[1],3)*Period)/(pow(math.sin(inclination),3)) #km

        M2_old=0
        M2_new=(1/1000)*13
        
        while abs(M2_old-M2_new)>0.0001:
            M2_old=M2_new
            
            f=(pow(M2_old,3))/(pow((M1+M2_old),2))-mass_func
            f_der=((pow(M2_old,2))*(M2_old+(3*M1)))/(pow((M1+M2_old),3))
        
            M2_new=M2_old-(f/f_der)
        
        M2=M2_new
        
        M_sun=rv_constant_values[0]
        M_jup=rv_constant_values[1]
        
        R_sun=rv_constant_values[2]
        R_jup=rv_constant_values[3]
        
        G=rv_constant_values[4]
        
        a1=(13751*fit_values[1]*Period*math.sqrt(1-pow(fit_values[4],2)))/math.sin(inclination) #km
        a2=(a1*M1)/M2 #km
        a=a1+a2 #km
        
        r_star=rvstar_parameter_values_list[2]
        r_planet=rvplanet_parameter_values_list[5]
        
        R_star=r_star*a
        R_planet=r_planet*a
        
        M1_g=(M1*M_sun)*1000 #g
        M2_g=(M2*M_sun)*1000 #g

        R_star_cm=R_star*(1000*100) #cm
        R_planet_cm=R_planet*(1000*100) #cm

        g_star=G*(M1_g/pow(R_star_cm,2))
        g_planet=G*(M2_g/pow(R_planet_cm,2))
        
        rv_result_parameter_values_list.append(round(M1,3)) #M Sun
        rv_result_parameter_values_list.append(round(((M2*M_sun)/M_jup),3)) #M Jup
        rv_result_parameter_values_list.append(round((a2/149597870.7),3)) #AU
        rv_result_parameter_values_list.append(round((R_planet/R_jup),3)) #R Jup
        rv_result_parameter_values_list.append(round(g_planet,3)) #cm/sn^2
        rv_result_parameter_values_list.append(round(math.log10(g_planet),3)) #unitless
        rv_result_parameter_values_list.append(round((R_star/R_sun),3)) #R Sun
        rv_result_parameter_values_list.append(round(g_star,3)) #cm/sn^2
        rv_result_parameter_values_list.append(round(math.log10(g_star),3)) #unitless
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        for i in range(len(rvmodel_parameter_names_list)):
            parameter_value=tk.Entry(master=frame_middle_rightrv)
            parameter_value.insert(tk.END, rvmodel_parameter_values_list[i])
            parameter_value.grid(row=2+i, column=1, sticky="nsew", padx=10)
            parameter_value.configure(state="readonly")
        
        for i in range(len(fit_values)):
            fit_value=tk.Entry(master=frame_middle_rightrv)
            fit_value.insert(tk.END, round(fit_values[i], rvmodel_parameter_round_list[i]))
            fit_value.grid(row=2+i, column=2, sticky="nsew", padx=10)
            fit_value.configure(state="readonly")
        
        for i in range(len(fit_errors)):
            fit_error=tk.Entry(master=frame_middle_rightrv)
            if fit_errors[i] is not None:
                fit_error.insert(tk.END, f"\u00B1 {round(fit_errors[i], rvmodel_parameter_round_list[i])}")
            
            else:
                fit_error.insert(tk.END, f"\u00B1 {fit_errors[i]}")
            
            fit_error.grid(row=2+i, column=3, sticky="nsew", padx=10)
            fit_error.configure(state="readonly")
        
        for widgets in rv_fitting.winfo_children():
            widgets.destroy()
        
        progress()
        rv_window.update()
        rvp_window.update()
        
        rv_fig=Figure(figsize=(6.0,3.2), dpi=90)
        rv_fig.patch.set_facecolor(main_bg)
        
        rv_fit=rv_fig.add_subplot(111)
        rv_fit.plot(Epoch_List_sorted, v_radial_list_sorted, data_style, color=data_bg)
        rv_fit.plot(Epoch_List_sin_sorted, vr_fit_sin_list_sorted, line_style, color=fit_color)
        rv_fit.plot([0,1], [fit_values[0], fit_values[0]], '--', color=sys_vel_color)
        rv_fit.patch.set_facecolor(tetriary_bg)
        rv_fit.errorbar(Epoch_List_sorted, v_radial_list_sorted, yerr=v_error_list_sorted, fmt=data_style, color=data_bg, ecolor=err_bar_color)
        rv_fig.subplots_adjust(left=0.19, bottom=0.177, top=0.895, right=0.958)
        
        rv_fit.xaxis.label.set_color(text_color_primary_label)
        rv_fit.yaxis.label.set_color(text_color_primary_label)
        
        rv_fit.tick_params(axis='x', colors=text_color_primary_axis)
        rv_fit.tick_params(axis='y', colors=text_color_primary_axis)
        
        rv_fit.spines['left'].set_color(text_color_primary_axis)
        rv_fit.spines['bottom'].set_color(text_color_primary_axis)
        rv_fit.spines['right'].set_color(text_color_primary_axis)
        rv_fit.spines['top'].set_color(text_color_primary_axis)
        
        rv_fit.set_title(graphtitle, fontsize=10, color=text_color_primary_label, fontweight="bold")
        rv_fit.set_ylabel("Radial Velocity (km/sec)", fontsize=10)
        rv_fit.set_xlabel("Phase", fontsize=10)
        
        plot_box=FigureCanvasTkAgg(rv_fig, rv_fitting)
        plot_box.draw()
        plot_box.get_tk_widget().pack()
        
        toolbar=NavigationToolbar2Tk(plot_box, rv_fitting)
        toolbar.update()
        plot_box.get_tk_widget().pack()
        
        for i in range(len(rv_result_parameter_values_list)):
            parametervalue_first=tk.Label(master=frame_bottom_leftrv, text=rv_result_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
        
        progress()
        rv_window.update()
        rvp_window.update()
    
    rvphase_button=tk.Button(master=frame_top_rightrv, text="Press to show Radial Velocity Data", command=rv_phase_show, bg=button_bg)
    rvfit_button=tk.Button(master=frame_top_rightrv, text="Press to calculate Model Radial Velocity Fit", command=rv_fit_first_show, bg=button_bg)
    rvfinalized_fit_button=tk.Button(master=frame_top_rightrv, text="Press to fit for precise Model Parameters", command=rv_fit_improved_show, bg=button_bg)
    
    rvphase_button.grid(row=2, column=1, sticky="nsew", padx=10, pady=(10,5))
    rvfit_button.grid(row=3, column=1, sticky="nsew", padx=10, pady=5)
    rvfinalized_fit_button.grid(row=4, column=1, sticky="nsew", padx=10, pady=5)
    
    #Middle Right Frame -> Results and Errors
    rvfit_results_label=tk.Label(master=frame_middle_rightrv, text="Fit Results", font=("Times New Roman", 10, "bold"), bg=title_bg, height=2, width=80)
    rvparameter_result_label=tk.Label(master=frame_middle_rightrv, text="Parameter Name", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    rvparameter_name_fit_label=tk.Label(master=frame_middle_rightrv, text="Original Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    rvparameter_value_fit_label=tk.Label(master=frame_middle_rightrv, text="Fitted Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    rverror_fit_label=tk.Label(master=frame_middle_rightrv, text="Error", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)

    rvfit_results_label.grid(row=0, column=0, columnspan=4, sticky="nsew", padx=10, pady=(10,5))
    rvparameter_result_label.grid(row=1, column=0, sticky="nsew", padx=20, pady=5)
    rvparameter_name_fit_label.grid(row=1, column=1, sticky="nsew", padx=20, pady=5)
    rvparameter_value_fit_label.grid(row=1, column=2, sticky="nsew", padx=20, pady=5)
    rverror_fit_label.grid(row=1, column=3, sticky="nsew", padx=20, pady=5)
    
    for i in range(len(rvmodel_parameter_names_list)):
        rvparametername=tk.Label(master=frame_middle_rightrv, text=rvmodel_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        rvparametername.grid(row=2+i, column=0, sticky="nsew", padx=20, pady=0)
    
    def rvupdate_parameters():
        if dl_button['text']=="Light Mode":
            main_bg="#0B0B0B"
            secondary_bg="#252525"
            tetriary_bg="#3A3A3A"
            title_bg="#FF7A37"
            
            button_bg="#f0e130"
            
            text_color_primary="white"
            text_color_secondary="black"
            
            color_frame_bg="#0C0B0B"
            
        if dl_button['text']=="Dark Mode":
            main_bg="#ECDBDB"
            secondary_bg="#FFFFFF"
            tetriary_bg="#FFFFFF"
            title_bg="#7A7A7A"
            
            button_bg="#C9C6C6"
            
            text_color_primary="black"
            text_color_secondary="black"
                
            color_frame_bg="#0C0B0B"
        
        try:
            if len(rvmodel_parameters_update.rvmodel_parameter_fit_value_list)==6:                
                for i in range(len(rvmodel_parameters_update.rvmodel_parameter_fit_value_list)):
                    if i<2:
                        rvstar_parameter_values_list[i]=round(rvmodel_parameters_update.rvmodel_parameter_fit_value_list[i], rvmodel_parameter_round_list[i])
                        
                    else:
                        rvplanet_parameter_values_list[i-2]=round(rvmodel_parameters_update.rvmodel_parameter_fit_value_list[i], rvmodel_parameter_round_list[i])
                    
                for i in range(len(rvstar_parameter_values_list)-2):
                    if len(rvstar_parameter_values_list)!=0:
                        parametervalue_first=tk.Label(master=frame_middle_leftrv, text=rvstar_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
                        parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
                
                    elif len(rvstar_parameter_values_list)==0:
                        parametervalue_first=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
                        parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
                        
                for i in range(len(rvplanet_parameter_values_list)-2):
                    if len(rvplanet_parameter_values_list)!=0:
                        parametervalue_first=tk.Label(master=frame_middle_leftrv, text=rvplanet_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
                        parametervalue_first.grid(row=4+i, column=1, sticky="nsew", padx=10)
                
                    elif len(rvplanet_parameter_values_list)==0:
                        parametervalue_first=tk.Label(master=frame_middle_leftrv, text="---", bg=secondary_bg, fg=text_color_primary)
                        parametervalue_first.grid(row=4+i, column=1, sticky="nsew", padx=10)
                    
            show_success=tk.messagebox.showinfo("Successful", "Parameters have been updated!")
            
        except AttributeError:
            show_error=tk.messagebox.showerror("Error", "There aren't any parameters to update.\n\nPlease do the Radial Velocity Modelling first in order to update the parameters.")
    
    def rvexport_func():
        if len(rvplanet_parameter_values_list)!=6 and len(rvstar_parameter_values_list)!=4:
            show_error=tk.messagebox.showerror("Error", "There aren't any parameters to export.\n\nPlease do the Radial Velocity Modelling first in order to export the parameters.")
    
        elif len(rvplanet_parameter_values_list)==6 and len(rvstar_parameter_values_list)==4:
            rvexport.exportplanet_parameter_values_list=rvplanet_parameter_values_list.copy()
            rvexport.exportplanet_parameter_values_list.pop(0)
            
            rvexport.exportstar_parameter_values_list=rvstar_parameter_values_list.copy()
            
            if len(planet_parameter_values_list)==6:
               for i in range(len(rvexport.exportplanet_parameter_values_list)):
                   planet_parameter_values_list[i+1]=rvexport.exportplanet_parameter_values_list[i]
            
            else:
                planet_parameter_values_list.append(0)
                for i in range(len(rvexport.exportplanet_parameter_values_list)):
                    planet_parameter_values_list.insert(i+1, rvexport.exportplanet_parameter_values_list[i])
            
            if len(star_parameter_values_list)==3:
                star_parameter_values_list[0]=rvexport.exportstar_parameter_values_list[2]
            
            else:
                star_parameter_values_list.clear()
                star_parameter_values_list.insert(0, rvexport.exportstar_parameter_values_list[2])
            
            for i in range(len(planet_parameter_values_list)):
                parametervalue=tk.Label(master=frame_middle_left, text=planet_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
                parametervalue.grid(row=2+i, column=1, sticky="nsew")
            
            starparametervalue=tk.Label(master=frame_middle_left, text=star_parameter_values_list[0], bg=secondary_bg, fg=text_color_primary)
            starparametervalue.grid(row=8, column=1, sticky="nsew")

            show_success=tk.messagebox.showinfo("Successful", "Parameters have been exported to Light Curve analysis!\n\nNote: Due to Light Curve T0 being the minimum of the transit and Radial Velocity T0 being passage of periapsis time, it is not exported.")
    
    rvupdate_parameters=tk.Button(master=frame_bottom_rightrv, text="Press to Update the Parameter Values", font=("Times New Roman", 10, "bold"), command=rvupdate_parameters, bg=button_bg)
    rvupdate_parameters.grid(row=0, column=0, pady=(10,5), padx=5)
    
    rvexport_parameters=tk.Button(master=frame_bottom_rightrv, text="Press to Export the Parameter Values to Light Curve", font=("Times New Roman", 10, "bold"), command=rvexport_func, bg=button_bg)
    rvexport_parameters.grid(row=1, column=0, pady=(5,10), padx=5)

system_name_list=[""]
    
def all_parameters_show():
    if dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
            
        button_bg="#f0e130"
        
        text_color_primary="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
            
    if dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        
        button_bg="#C9C6C6"
            
        text_color_primary="black"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
    ap_window=tk.Toplevel(master=main_window)
    ap_window.title("All Parameters and Values")
    ap_window.resizable(width=False, height=False)
    ap_window.iconbitmap(r'icon_lcf.ico')
    ap_window.configure(bg=main_bg)
    ap_window.wm_transient(main_window)

    ap_first_frame=tk.Frame(master=ap_window, bg=secondary_bg, padx=10, pady=5)
    ap_second_frame=tk.Frame(master=ap_window, bg=secondary_bg, padx=10, pady=5)
    ap_third_frame=tk.Frame(master=ap_window, bg=secondary_bg, padx=10, pady=5)
    
    ap_first_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
    ap_second_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
    ap_third_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
    
    #Planet Parameters
    ppall_label=tk.Label(master=ap_first_frame, text="Planet Parameters", padx=5, pady=10, bg=title_bg, font=("Times New Roman", 10, "bold"), width=85)
    ppall_label.grid(row=0, column=0, columnspan=4, sticky="nsew")
    
    ppall_label=tk.Label(master=ap_first_frame, text="Parameter Name", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    ppall_notation_label=tk.Label(master=ap_first_frame, text="Notation", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    ppall_name_label=tk.Label(master=ap_first_frame, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    ppall_unit_label=tk.Label(master=ap_first_frame, text="Unit", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    
    ppall_label.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
    ppall_notation_label.grid(row=1, column=1, sticky="nsew", padx=10, pady=10)
    ppall_name_label.grid(row=1, column=2, sticky="nsew", padx=10, pady=10)
    ppall_unit_label.grid(row=1, column=3, sticky="nsew", padx=10, pady=10)
    
    for i in range(len(all_planet_parameter_names_list)):
        apparametername=tk.Label(master=ap_first_frame, text=all_planet_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametername.grid(row=2+i, column=0, sticky="nsew", padx=10, pady=5)
        
        apparameternotation=tk.Label(master=ap_first_frame, text=all_planet_parameter_notation_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparameternotation.grid(row=2+i, column=1, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_first_frame, text="---", font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=2+i, column=2, sticky="nsew", padx=10, pady=5)
        
        apparameterunit=tk.Label(master=ap_first_frame, text=all_planet_parameter_unit_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparameterunit.grid(row=2+i, column=3, sticky="nsew", padx=10, pady=5)
        
    if len(planet_parameter_values_list)==6:
        for i in range(len(planet_parameter_names_list)):
            apparametervalue=tk.Label(master=ap_first_frame, text=planet_parameter_values_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
            apparametervalue.grid(row=2+i, column=2, sticky="nsew", padx=10, pady=5)
    
    if len(rv_result_parameter_values_list)==9:
        apparametervalue=tk.Label(master=ap_first_frame, text=rv_result_parameter_values_list[2], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=8, column=2, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_first_frame, text=rv_result_parameter_values_list[3], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=9, column=2, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_first_frame, text=rv_result_parameter_values_list[1], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=10, column=2, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_first_frame, text=rv_result_parameter_values_list[4], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=11, column=2, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_first_frame, text=rv_result_parameter_values_list[5], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=12, column=2, sticky="nsew", padx=10, pady=5)
    
    #Star Parameters
    spall_label=tk.Label(master=ap_second_frame, text="Star Parameters", padx=5, pady=10, bg=title_bg, font=("Times New Roman", 10, "bold"), width=85)
    spall_label.grid(row=0, column=0, columnspan=4, sticky="nsew")
    
    spall_label=tk.Label(master=ap_second_frame, text="Parameter Name", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    spall_notation_label=tk.Label(master=ap_second_frame, text="Notation", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    spall_name_label=tk.Label(master=ap_second_frame, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    spall_unit_label=tk.Label(master=ap_second_frame, text="Unit", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    
    spall_label.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
    spall_notation_label.grid(row=1, column=1, sticky="nsew", padx=10, pady=10)
    spall_name_label.grid(row=1, column=2, sticky="nsew", padx=10, pady=10)
    spall_unit_label.grid(row=1, column=3, sticky="nsew", padx=10, pady=10)
    
    for i in range(len(all_star_parameter_names_list)):
        apparametername=tk.Label(master=ap_second_frame, text=all_star_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametername.grid(row=2+i, column=0, sticky="nsew", padx=10, pady=5)
        
        apparameternotation=tk.Label(master=ap_second_frame, text=all_star_parameter_notation_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparameternotation.grid(row=2+i, column=1, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_second_frame, text="---", font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=2+i, column=2, sticky="nsew", padx=10, pady=5)
        
        apparameterunit=tk.Label(master=ap_second_frame, text=all_star_parameter_unit_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparameterunit.grid(row=2+i, column=3, sticky="nsew", padx=10, pady=5)
    
    if len(star_parameter_values_list)==3:
        for i in range(len(star_parameter_names_list)):
            apparametervalue=tk.Label(master=ap_second_frame, text=star_parameter_values_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
            apparametervalue.grid(row=2+i, column=2, sticky="nsew", padx=10, pady=5)
    
    if len(rv_result_parameter_values_list)==9:
        apparametervalue=tk.Label(master=ap_second_frame, text=rv_result_parameter_values_list[6], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=5, column=2, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_second_frame, text=rv_result_parameter_values_list[0], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=6, column=2, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_second_frame, text=rv_result_parameter_values_list[7], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=7, column=2, sticky="nsew", padx=10, pady=5)
        
        apparametervalue=tk.Label(master=ap_second_frame, text=rv_result_parameter_values_list[8], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        apparametervalue.grid(row=8, column=2, sticky="nsew", padx=10, pady=5)
    
    systemnamelabel=tk.Label(master=ap_third_frame, text="Enter System Name:", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
    systemnamelabel.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=10, padx=5)
    
    systemnameentry=tk.Entry(master=ap_third_frame, bg=tetriary_bg, fg=text_color_primary)
    systemnameentry.grid(row=0, column=2, columnspan=2, sticky="nsew", pady=10, padx=5)
    
    if system_name_list[0]!="":
        systemnameentry.insert(tk.END, system_name_list[0])

    def all_parameters_latex_table():
        geometry_options = {"tmargin": "1cm"}
        tabledoc = pylatex.Document(geometry_options=geometry_options)
        
        system_name=systemnameentry.get()
        
        with tabledoc.create(pylatex.MiniPage(align="c")):
            if system_name=="":
                tabledoc.append(pylatex.LargeText(pylatex.utils.bold("System Parameters")))
            
            else:
                tabledoc.append(pylatex.LargeText(pylatex.utils.bold(f"{system_name} System Parameters")))
    
        with tabledoc.create(pylatex.LongTable("l l l", pos="c", row_height=(2), col_space="40px")) as data_table:
            data_table.add_hline()
            data_table.add_row(" ","Planet Parameters"," ")
            data_table.add_hline()
        
            data_table.add_hline()
            data_table.add_row(["Parameter", "Value", "Unit"])
            data_table.add_hline()
            
            if len(planet_parameter_values_list)==6:
                for i in range(len(planet_parameter_names_list)):
                    latex_planet_parameter_values[i]=planet_parameter_values_list[i]
            
            if len(rv_result_parameter_values_list)==9:
                latex_planet_parameter_values[6]=rv_result_parameter_values_list[2]
                latex_planet_parameter_values[7]=rv_result_parameter_values_list[3]
                latex_planet_parameter_values[8]=rv_result_parameter_values_list[1]
                latex_planet_parameter_values[9]=rv_result_parameter_values_list[4]
                latex_planet_parameter_values[10]=rv_result_parameter_values_list[5]
            
            for i in range(len(latex_planet_parameter_names)):
                row=[latex_planet_parameter_names[i], latex_planet_parameter_values[i], latex_planet_parameter_units[i]]
                data_table.add_row(row)
            
            data_table.add_hline()
    
        with tabledoc.create(pylatex.LongTable("l l l", pos="c", row_height=(2), col_space="40px")) as sdata_table:
            sdata_table.add_hline()
            sdata_table.add_row(" ","Star Parameters"," ")
            sdata_table.add_hline()
            
            sdata_table.add_hline()
            sdata_table.add_row(["Parameter", "Value", "Unit"])
            sdata_table.add_hline()
            
            if len(star_parameter_values_list)==3:
                for i in range(len(star_parameter_names_list)):
                    latex_star_parameter_values[i]=star_parameter_values_list[i]
            
            if len(rv_result_parameter_values_list)==9:
                latex_star_parameter_values[3]=rv_result_parameter_values_list[6]
                latex_star_parameter_values[4]=rv_result_parameter_values_list[0]
                latex_star_parameter_values[5]=rv_result_parameter_values_list[7]
                latex_star_parameter_values[6]=rv_result_parameter_values_list[8]
            
            for i in range(len(latex_star_parameter_names)):
                row=[latex_star_parameter_names[i], latex_star_parameter_values[i], latex_star_parameter_units[i]]
                sdata_table.add_row(row)
            
            sdata_table.add_hline()
            sdata_table.add_hline()
        
        # Creating a pdf
        if system_name=="":
            tabledoc.generate_pdf('Parameter Table', clean_tex=False, compiler="pdflatex")
        
        else:
            tabledoc.generate_pdf(f'{system_name} Parameter Table', clean_tex=False, compiler="pdflatex")
        
        tex = tabledoc.dumps()
        
        if system_name=="":
            with open("Parameter Table LaTeX Script", "w") as file:
                file.write(tex)
        
        else:
            with open(f"{system_name} Parameter Table", "w") as file:
                file.write(tex)
        
        success=tk.messagebox.showinfo("Successful", "Parameter values have been exported to a PDF file as a Latex Table!")
        
        system_name_list.insert(0, system_name)
    
    latexpdfbutton=tk.Button(master=ap_third_frame, text="Press to Export Parameter Values as LaTeX Table (PDF Format)", font=("Times New Roman", 10, "bold"), command=all_parameters_latex_table, bg=button_bg, width=85)
    latexpdfbutton.grid(row=1, column=0, columnspan=4, sticky="nsew", pady=10, padx=5)
    
#Top Left Frame --> The Buttons
entry_label=tk.Label(master=frame_top_left, text="Data Entry Area", padx=5, pady=10, bg=title_bg, font=("Times New Roman", 10, "bold"), width=60)
open_button=tk.Button(master=frame_top_left, text="Light Curve Data Entry", padx=5, pady=5, command=lcdataentry_window, bg=button_bg, width=20)
lc_button=tk.Button(master=frame_top_left, text="Show Raw Data Light Curve", padx=5, pady=5, command=lightcurvewindow, bg=button_bg, width=20)
parametersofstar_button=tk.Button(master=frame_top_left, text="Star Parameters", padx=5, pady=5, command=starparameters, bg=button_bg, width=20)
parametersofplanet_button=tk.Button(master=frame_top_left, text="Planet Parameters", padx=5, pady=5, command=planetparameters, bg=button_bg, width=20)
rv_analysis_button=tk.Button(master=frame_top_left, text="Radial Velocity (RV) Analysis", padx=5, pady=5, command=rvanalysis_window, bg=button_bg, width=20)
all_parameters_button=tk.Button(master=frame_top_left, text="All Parameters and Values", padx=5, pady=5, command=all_parameters_show, bg=button_bg, width=20)

entry_label.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=10)
open_button.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
lc_button.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
parametersofstar_button.grid(row=2, column=0, sticky="nsew", padx=5, pady=5)
parametersofplanet_button.grid(row=2, column=1, sticky="nsew", padx=5, pady=5)
rv_analysis_button.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
all_parameters_button.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)

#Middle Left Frame --> Should it be varied or not? + The values of the parameters
parameter_label=tk.Label(master=frame_middle_left, text="Model Parameters", font=("Times New Roman", 10, "bold"), bg=title_bg, width=58)
parameter_name_label=tk.Label(master=frame_middle_left, text="Parameter", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
parameter_value_label=tk.Label(master=frame_middle_left, text="Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
variability_label=tk.Label(master=frame_middle_left, text="Free/Fixed", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)

parameter_label.grid(row=0, column=0, columnspan=3, sticky="nsew", padx=10, pady=(10,5))
parameter_name_label.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
parameter_value_label.grid(row=1, column=1, sticky="nsew", padx=10, pady=5)
variability_label.grid(row=1, column=2, sticky="nsew", padx=10, pady=5)

for i in range(len(model_parameter_names_list)):
    parametername=tk.Label(master=frame_middle_left, text=model_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
    parametername.grid(row=2+i, column=0, sticky="nsew", padx=10)

    parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
    parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)

T0_vary=tk.IntVar()
Period_vary=tk.IntVar()
Eccentricity_vary=tk.IntVar()
Argument_Of_Periapsis_vary=tk.IntVar()
inclination_vary=tk.IntVar()
r_planet_vary=tk.IntVar()

r_star_vary=tk.IntVar()
ua_vary=tk.IntVar()
ub_vary=tk.IntVar()

T0_vary_check=tk.Checkbutton(master=frame_middle_left, variable=T0_vary, onvalue=1, offvalue=0, bg=secondary_bg)
Period_vary_check=tk.Checkbutton(master=frame_middle_left, variable=Period_vary, onvalue=1, offvalue=0, bg=secondary_bg)
Eccentricity_vary_check=tk.Checkbutton(master=frame_middle_left, variable=Eccentricity_vary, onvalue=1, offvalue=0, bg=secondary_bg)
Argument_Of_Periapsis_vary_check=tk.Checkbutton(master=frame_middle_left, variable=Argument_Of_Periapsis_vary, onvalue=1, offvalue=0, bg=secondary_bg)
inclination_vary_check=tk.Checkbutton(master=frame_middle_left, variable=inclination_vary, onvalue=1, offvalue=0, bg=secondary_bg)
r_planet_vary_check=tk.Checkbutton(master=frame_middle_left, variable=r_planet_vary, onvalue=1, offvalue=0, bg=secondary_bg)

r_star_vary_check=tk.Checkbutton(master=frame_middle_left, variable=r_star_vary, onvalue=1, offvalue=0, bg=secondary_bg)
ua_vary_check=tk.Checkbutton(master=frame_middle_left, variable=ua_vary, onvalue=1, offvalue=0, bg=secondary_bg)
ub_vary_check=tk.Checkbutton(master=frame_middle_left, variable=ub_vary, onvalue=1, offvalue=0, bg=secondary_bg)

T0_vary_check.grid(row=2, column=2, padx=10)
Period_vary_check.grid(row=3, column=2, padx=10)
Eccentricity_vary_check.grid(row=4, column=2, padx=10)
Argument_Of_Periapsis_vary_check.grid(row=5, column=2, padx=10)
inclination_vary_check.grid(row=6, column=2, padx=10)
r_planet_vary_check.grid(row=7, column=2, padx=10)

r_star_vary_check.grid(row=8, column=2, padx=10)
ua_vary_check.grid(row=9, column=2, padx=10)
ub_vary_check.grid(row=10, column=2, padx=10)

#Top Right Frame --> Fitting Area
fit_text=tk.Label(master=frame_top_right, text="Light Curve Model", font=("Times New Roman", 14, "bold"), bg=title_bg)
fit_text.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
lc_fitting=tk.Frame(master=frame_top_right, width=(630), height=(400), highlightbackground="red", highlightthickness=0, bg=main_bg)
lc_fitting.grid(row=1, column=1)

def lc_phase_show():
    if len(lc_time)==0:
        inputfit_error=tk.messagebox.showerror("Error", "Please enter the Light Curve Data to proceed.")   
    
    elif len(star_parameter_values_list)!=3:
        lc_show_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
    
    elif len(planet_parameter_values_list)!=6:
        lc_show_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
    
    graphtitle="Light Curve"
    
    if len(preferences_list)==12:
        graphtitle=preferences_list[0]
        data_bg=preferences_list[1]
        main_bg=preferences_list[3]
        tetriary_bg=preferences_list[4]
        text_color_primary_label=preferences_list[5]
        text_color_primary_axis=preferences_list[6]
        sys_vel_color=preferences_list[7]
        err_bar_color=preferences_list[8]
        data_style=preferences_list[10]
        line_style=preferences_list[11]
        
    elif dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        data_bg="red"
        fit_color="blue"
        data_style="."
        line_style="-"
        
        button_bg="#f0e130"
        
        text_color_primary_label="white"
        text_color_primary_axis="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
    elif dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        data_bg="red"
        fit_color="blue"
        data_style="."
        line_style="-"
    
        button_bg="#C9C6C6"
        
        text_color_primary_label="black"
        text_color_primary_axis="black"
        text_color_secondary="black"
    
        color_frame_bg="#0C0B0B"

    T0=planet_parameter_values_list[0]
    period=planet_parameter_values_list[1]    
    
    Delta_phase_List=[]
    
    for time in lc_time:
        #Epoch değerleri elde edilir.
        if ((float(time)-T0)/period)<0:
            Epoch_value=(float(time)-T0)/period-(int((float(time)-T0)/period)-1)
            Epoch_deg_value=((float(time)-T0)/period-(int((float(time)-T0)/period)-1))*360
    
        else:
            Epoch_value=(float(time)-T0)/period-(int((float(time)-T0)/period))
            Epoch_deg_value=((float(time)-T0)/period-(int((float(time)-T0)/period)))*360
        
        Epoch_List.append(Epoch_value)
        
        if Epoch_value<0.2:
            Delta_phase=Epoch_value+1
        
        else:
            Delta_phase=Epoch_value
        
        Delta_phase_List.append(Delta_phase)
    
    for widgets in lc_fitting.winfo_children():
        widgets.destroy()
    
    lc_fig=Figure(figsize=(7,4), dpi=90)
    lc_fig.patch.set_facecolor(main_bg)
    
    lc_fit=lc_fig.add_subplot(111)
    lc_fit.plot(Delta_phase_List, lc_flux, data_style, color=data_bg)
    lc_fit.patch.set_facecolor(tetriary_bg)
    
    lc_fit.xaxis.label.set_color(text_color_primary_label)
    lc_fit.yaxis.label.set_color(text_color_primary_label)
    
    lc_fit.tick_params(axis='x', colors=text_color_primary_axis)
    lc_fit.tick_params(axis='y', colors=text_color_primary_axis)
    
    lc_fit.spines['left'].set_color(text_color_primary_axis)
    lc_fit.spines['bottom'].set_color(text_color_primary_axis)
    lc_fit.spines['right'].set_color(text_color_primary_axis)
    lc_fit.spines['top'].set_color(text_color_primary_axis)
    
    lc_fit.set_title(graphtitle, fontsize=12, color=text_color_primary_label, fontweight="bold")
    lc_fit.set_ylabel("Intensity", fontsize=12)
    lc_fit.set_xlabel("Phase", fontsize=12)
    
    plot_box=FigureCanvasTkAgg(lc_fig, lc_fitting)
    plot_box.draw()
    plot_box.get_tk_widget().pack()
    
    toolbar=NavigationToolbar2Tk(plot_box, lc_fitting)
    toolbar.update()
    plot_box.get_tk_widget().pack()
        
lcphase_button=tk.Button(master=frame_top_right, text="Press to show Light Curve", command=lc_phase_show, bg=button_bg)
lcphase_button.grid(row=2, column=1, sticky="nsew", padx=10, pady=(10,5))

Eclipse_Functions, Light_Curve_Fit=[], []

def input_parameter_fit():
    graphtitle="Fit with Input Parameters"
    
    if len(preferences_list)==12:
        graphtitle=preferences_list[0]
        data_bg=preferences_list[1]
        fit_color=preferences_list[2]
        main_bg=preferences_list[3]
        tetriary_bg=preferences_list[4]
        text_color_primary_label=preferences_list[5]
        text_color_primary_axis=preferences_list[6]
        sys_vel_color=preferences_list[7]
        err_bar_color=preferences_list[8]
        data_style=preferences_list[10]
        line_style=preferences_list[11]
        
        if dl_button['text']=="Light Mode":
            lcpmain_bg="#0B0B0B"
            lcpsecondary_bg="#252525"
            text_color_primary="white"
            lcpbutton_bg="#f0e130"
            
        elif dl_button['text']=="Dark Mode":
            lcpmain_bg="#ECDBDB"
            lcpsecondary_bg="#FFFFFF"
            text_color_primary="black"
            lcpbutton_bg="#C9C6C6"
    
    elif dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        data_bg="red"
        fit_color="blue"
        data_style="."
        line_style="-"
        
        button_bg="#f0e130"
        lcpbutton_bg="#f0e130"
        
        text_color_primary_label="white"
        text_color_primary_axis="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
        lcpmain_bg="#0B0B0B"
        lcpsecondary_bg="#252525"
        text_color_primary="white"
        
    elif dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        data_bg="red"
        fit_color="blue"
        data_style="."
        line_style="-"
    
        button_bg="#C9C6C6"
        lcpbutton_bg="#C9C6C6"
        
        text_color_primary_label="black"
        text_color_primary_axis="black"
        text_color_secondary="black"
    
        color_frame_bg="#0C0B0B"
        
        lcpmain_bg="#ECDBDB"
        lcpsecondary_bg="#FFFFFF"
        text_color_primary="black"
    
    #Define Progress
    def progress_stage():
        return f"Current Progress: %{lcprogressbar['value']}"
    
    def progress():
        if lcprogressbar['value']<100:
            lcprogressbar['value']+=10
            lcprogresslabel['text']=progress_stage()
        
        if lcprogressbar['value']==100:
            progressfinish=tk.messagebox.showinfo("Successful", "The Light Curve Modelling is finished!")
            lcp_window.destroy()
    
    #Define Pop-up Progress Bar
    lcp_window=tk.Toplevel(master=main_window)
    lcp_window.title("Light Curve Modelling Progress")
    lcp_window.iconbitmap(r'icon_lcf.ico')
    lcp_window.configure(bg=lcpmain_bg)
    lcp_window.wm_transient(main_window)
    lcp_window.resizable(width=False, height=False)
    
    window_width=lcp_window.winfo_reqwidth()
    window_height=lcp_window.winfo_reqheight()

    screen_width=lcp_window.winfo_screenwidth()
    screen_height=lcp_window.winfo_screenheight()

    x_cordinate = int((screen_width/2) - (window_width/2))
    y_cordinate = int((screen_height/2) - (window_height/2))
    
    lcp_window.geometry("+{}+{}".format(x_cordinate, y_cordinate))
    
    lcinprogress=tk.Label(master=lcp_window, text="Light Curve Modelling in Progress...\n(Progress may take up to a few minutes)", bg=lcpsecondary_bg, fg=text_color_primary)
    lcinprogress.grid(row=0, column=0, pady=10, padx=10)
    
    lcprogressbar=ttk.Progressbar(master=lcp_window, orient="horizontal", mode="determinate", length=300)
    lcprogressbar.grid(row=1, column=0, pady=10, padx=10)
    
    lcprogresslabel=tk.Label(master=lcp_window, text=progress_stage(), bg=lcpbutton_bg)
    lcprogresslabel.grid(row=2, column=0, pady=10, padx=10)
    
    #Define progressbar value
    lcprogressbar['value']=0
    
    model_parameter_values_list=planet_parameter_values_list+star_parameter_values_list
    
    if len(lc_time)==0:
        inputfit_error=tk.messagebox.showerror("Error", "Please enter the Light Curve Data to proceed.")
        lcp_window.destroy()
    
    elif (len(star_parameter_values_list)!=3):
        inputfit_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
        lcp_window.destroy()
        
    elif (len(planet_parameter_values_list)!=6):
        inputfit_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
        lcp_window.destroy()
    
    T0=planet_parameter_values_list[0]
    Period=planet_parameter_values_list[1]
    Eccentricity=planet_parameter_values_list[2]
    Argument_Of_Periapsis=planet_parameter_values_list[3]
    inclination=planet_parameter_values_list[4]
    r_planet=planet_parameter_values_list[5]
    
    r_star=star_parameter_values_list[0]
    ua=star_parameter_values_list[1]
    ub=star_parameter_values_list[2]
    
    progress()
    main_window.update()
    lcp_window.update()
    
    def Delta(lc_time, T0, Period, Eccentricity, Argument_Of_Periapsis, inclination, r_star, r_planet):
        Epoch_List, Delta_values_List, Delta_phase_List, c_values_List=[], [], [], []

        for time in lc_time:
            #Epoch değerleri elde edilir.
            if ((float(time)-T0)/Period)<0:
                Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
    
            else:
                Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360
        
            Epoch_List.append(Epoch_value)
    
            #T-T0 değeri hesaplanır.
            T_T0=time-T0
            T_T0_values.append(T_T0)
    
            #Delta Değerleri Yörünge Basıklığı değerine göre hesaplanır.
            if Eccentricity>0:
                Mean_Anomaly=((2*math.pi)/Period)*(T_T0)
                Mean_Anomaly_values.append((Mean_Anomaly)) #M değerleri Radyan Biriminde
                
                Ei_old=Mean_Anomaly #Radyan Biriminde
                Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old)) #Radyan Biriminde
                error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                while error>0.001:
                    Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old))
                    error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                    Ei_old=Ei_new
        
                Eccentric_Anomaly_values.append(Ei_new) #E değerleri Radyan Biriminde
                
                True_Anomaly_0=((math.pi/2)-Argument_Of_Periapsis) #ν değerleri Radyan Biriminde
            
                True_Anomaly=2*math.atan(math.sqrt((1+Eccentricity)/(1-Eccentricity))*math.tan((Ei_new)/2))
                True_Anomaly_values.append(True_Anomaly) #ν değerleri Radyan Biriminde
                
                rho=(1-pow(Eccentricity,2))/(1+Eccentricity*math.cos(True_Anomaly))
                Rho_values.append(rho)
       
                delta=rho*math.sqrt((pow(math.sin(True_Anomaly-True_Anomaly_0),2)*pow(math.sin(inclination),2))+pow(math.cos(inclination),2))
                #δ değerleri Radyan Biriminde
                Delta_values_List.append(delta)
       
            elif Eccentricity==0:
                delta=math.sqrt((pow(math.sin(math.radians(Epoch_deg_value)),2)*pow(math.sin(inclination),2))+pow(math.cos(inclination),2))
                #δ değerleri Radyan Biriminde
                Delta_values_List.append(delta)
                
            #Epoch değerleri de belirlenir.
            if Epoch_value<0.2:
                Delta_phase=Epoch_value+1
        
            else:
                Delta_phase=Epoch_value
        
            Delta_phase_List.append(Delta_phase)
        
            #c değerleri de hesaplanır.
            c=delta/(r_star+r_planet)
            if c>1:
                c=1
                
            c_values_List.append(c)
        
        return Epoch_List, Delta_values_List, Delta_phase_List, c_values_List
    
    progress()
    main_window.update()
    lcp_window.update()
    
    def a_and_b_calculation(r_star, r_planet):
        a=r_star/(r_planet+r_star)
        b=1-a
    
        return a, b
    
    def C_constants_calculation(ua, ub):
        u1=ua+(2*ub)
        u2=-ub
    
        C0=(3-(3*u1))/(3-u1)
        C1=(3*u1)/(3-u1)
        C2=(6*u2)/(6-(2*u1)-(3*u2))
    
        return C0, C1, C2
    
    progress()
    main_window.update()
    lcp_window.update()
    
    def Jacobi_Polynomial(h,alpha,beta,z,x):
        Jacobi_Polynomials=list()
    
        Lambda=alpha+beta+1 #İşlem kolaylığı açısından tanımlanan sabit
    
        for n in range(0,x+1):
            if n==0:
                Jacobi=1
                Jacobi_Polynomials.append(Jacobi)
                continue
        
            if n==1:
                Jacobi=-(beta+1)+(Lambda+1)*z
                Jacobi_Polynomials.append(Jacobi)
                continue
        
            if n==2:
                Jacobi=(beta+1)*(beta+2)/2-(beta+2)*(Lambda+2)*z+(Lambda+2)*(Lambda+3)*pow(z,2)/2
                Jacobi_Polynomials.append(Jacobi)
                continue
    
            if n>2:
                for n in range(2,x):
                    Constant_1=(2*n+Lambda)*(2*n+Lambda+1)/((n+1)*(n+Lambda))
                    Constant_2=((2*n+Lambda)*(pow(alpha,2)-pow(beta,2)+1)-pow((2*n+Lambda),3))/(2*(n+1)*(n+Lambda)*(2*n+Lambda-1))
                    Constant_3=((n+alpha)*(n+beta)*(2*n+Lambda+1))/((n+1)*(n+Lambda)*(2*n+Lambda-1))
                    Jacobi=(Constant_1*z+Constant_2)*Jacobi_Polynomials[n]-Constant_3*Jacobi_Polynomials[n-1]
                    Jacobi_Polynomials.append(Jacobi)
                    continue
        
            return Jacobi_Polynomials
    
    progress()
    main_window.update()
    lcp_window.update()
    
    def Fractional_Alpha_Function(h, a, b, c):
        #Nu ifadesi tanımlanır. (Her alfa fonksiyonu için ayrı değere sahiptir!)
        nu=(h+2)/2
        
        #Toplam Sembollerinden önce bulunan 1. ifadenin hesabı yapılır.
        alpha_before_sum=pow(b,2)*math.gamma(nu)*pow((1-pow(c,2)),(nu+1))
        
        #Toplam Sembolü içinde bulunan ifadelerin hesabı için de makalede bulunan denklem kullanılacaktır.
        #Önce Jacobi Polinomlarımızın her ikisini de her bir yapılacak toplam için elde etmemiz gerekmektedir.
        #Sonrasında bunu Toplam Sembolündeki ilk kısım ile çarparak sonucu elde edebiliriz.
        Jacobi_1=Jacobi_Polynomial(h, 1, nu, a, 80)
        Jacobi_2=Jacobi_Polynomial(h, 1+nu, 0, pow(c,2), 80)
    
        sum_total=0
        for n in range(0,81):
            sum_total+=(((math.factorial(n)*(nu+2*n+2))/((n+1)*math.gamma(nu+n+1)))*pow(Jacobi_1[n],2)*Jacobi_2[n])
    
        #İncelenen gözlem zamanı için Küçük Alfa Fonksiyonlarından biri elde edilir.
        alpha=alpha_before_sum*sum_total
        return alpha
    
    progress()
    main_window.update()
    lcp_window.update()
    
    #Formüllerde iterasyon 0'dan sonsuza yapılmakta.
    #Kod kapsamında iterasyonumuz 0'dan 80'e kadar yapılacaktır.
    def Alpha_Function(lc_time, T0, Period, r_star, r_planet, Eccentricity, Argument_Of_Periapsis, inclination, ua, ub):
        Eclipse_Functions, Light_Curve_Fit=[], []
        
        #Verilen lc_time için Epoch, Delta ve Phase değeri elde edilir.
        Epoch_Fit_List, Delta_values_Fit_List, Delta_phase_Fit_List, c_values_Fit_List = Delta(lc_time, T0, Period, Eccentricity, Argument_Of_Periapsis, inclination, r_star, r_planet)
    
        #a ve b parametreleri hesaplanır.
        #Gözlem zamanı ile değişmediğinden döngüye gerek kalmadan hesap yapılabilir.
        a, b = a_and_b_calculation(r_star, r_planet)
    
        #Alfa Fonksiyonu hesabında gerekli olan C0, C1 ve C2 değerleri hesaplanır.
        #Gözlem zamanı ile değişmediğinden döngüye gerek kalmadan hesap yapılabilir.
        C0, C1, C2 = C_constants_calculation(ua, ub)
    
        #lc_time zamanı için Kesirsel Alfa Fonksiyonları hesaplanır.
        for c_value in c_values_Fit_List:
            frac_alpha_0=Fractional_Alpha_Function(0, a, b, c_value) #alpha_0
            frac_alpha_1=Fractional_Alpha_Function(1, a, b, c_value) #alpha_1
            frac_alpha_2=Fractional_Alpha_Function(2, a, b, c_value) #alpha_2
    
            #lc_time zamanı için Tutulma Fonksiyonunun (Alfa Fonksiyonunun) değeri hesaplanır.
            Eclipse_Function=(C0*frac_alpha_0)+(C1*frac_alpha_1)+(C2*frac_alpha_2)
            Eclipse_Functions.append(Eclipse_Function)
            Light_Curve_Fit.append(1-Eclipse_Function)
        
        return Light_Curve_Fit
    
    progress()
    main_window.update()
    lcp_window.update()
    
    Epoch_List, Delta_values_List, Delta_phase_List, c_values_List=Delta(lc_time, T0, Period, Eccentricity, Argument_Of_Periapsis, inclination, r_star, r_planet)
    Light_Curve_Fit=Alpha_Function(lc_time, T0, Period, r_star, r_planet, Eccentricity, Argument_Of_Periapsis, inclination, ua, ub)
    
    progress()
    main_window.update()
    lcp_window.update()
    
    #Bazen alınan datalardaki zaman değerleri farklı zaman araklıklarında olabilir.
    #Örneğin 360.15 zamanından sonra 630.15, sonrasında da tekrardan 361.15 zamanlarında gözlemler alınmış olup bize datada öyle gelebilir.
    #Bu nedenle Fit değerlerini, Akı değerlerini ve Evre Değerlerini minimumdan maksimuma sıralamamız faydalı olacaktır.
    #Bu sayede fit çizgimizde bir problem oluşması önlenecektir.
    
    Delta_phase_sort_list=Delta_phase_List.copy()
    lc_flux_sort_list=lc_flux.copy()
    Light_Curve_Fit_sort_list=Light_Curve_Fit.copy()
    
    Delta_phase_sorted, lc_flux_sorted, Light_Curve_Fit_sorted = [], [], []
    
    for i in range(len(Delta_phase_List)):
        min_val=min(Delta_phase_sort_list) #Minimum zaman
        min_index=Delta_phase_sort_list.index(min_val) #Minimum zamanın index değeri
    
        Delta_phase_sorted.append(min_val)
        Delta_phase_sort_list.pop(min_index)
    
        lc_flux_sorted.append(lc_flux_sort_list[min_index])
        lc_flux_sort_list.pop(min_index)
    
        Light_Curve_Fit_sorted.append(Light_Curve_Fit_sort_list[min_index])
        Light_Curve_Fit_sort_list.pop(min_index)
    
    progress()
    main_window.update()
    lcp_window.update()
    
    progress()
    main_window.update()
    lcp_window.update()
    
    for i in range(len(model_parameter_names_list)):
        parameter_value=tk.Entry(master=frame_middle_right)
        parameter_value.insert(tk.END, model_parameter_values_list[i])
        parameter_value.grid(row=2+i, column=1, sticky="nsew", padx=10)
        parameter_value.configure(state="readonly")
    
    for i in range(len(model_parameter_names_list)):
        fit_value=tk.Entry(master=frame_middle_right)
        fit_value.insert(tk.END, model_parameter_values_list[i])
        fit_value.grid(row=2+i, column=2, sticky="nsew", padx=10)
        fit_value.configure(state="readonly")
    
    for i in range(len(model_parameter_names_list)):
        fit_error=tk.Entry(master=frame_middle_right)
        fit_error.insert(tk.END, f"\u00B1 None")
        fit_error.grid(row=2+i, column=3, sticky="nsew", padx=10)
        fit_error.configure(state="readonly")
    
    model_parameters_update.model_parameter_fit_value_list=model_parameter_values_list.copy()
    model_parameters_update.model_parameter_error_value_list=["\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None", "\u00B1 None"]
    
    for widgets in lc_fitting.winfo_children():
        widgets.destroy()
    
    lc_fig=Figure(figsize=(7,4), dpi=90)
    lc_fig.patch.set_facecolor(main_bg)
    
    lc_fit=lc_fig.add_subplot(111)
    lc_fit.plot(Delta_phase_sorted, lc_flux_sorted, data_style, color=data_bg)
    lc_fit.plot(Delta_phase_sorted, Light_Curve_Fit_sorted, line_style, color=fit_color, label="Fit")
    lc_fit.patch.set_facecolor(tetriary_bg)
    
    lc_fit.xaxis.label.set_color(text_color_primary_label)
    lc_fit.yaxis.label.set_color(text_color_primary_label)
    
    lc_fit.tick_params(axis='x', colors=text_color_primary_axis)
    lc_fit.tick_params(axis='y', colors=text_color_primary_axis)
    
    lc_fit.spines['left'].set_color(text_color_primary_axis)
    lc_fit.spines['bottom'].set_color(text_color_primary_axis)
    lc_fit.spines['right'].set_color(text_color_primary_axis)
    lc_fit.spines['top'].set_color(text_color_primary_axis)
    
    lc_fit.set_title(graphtitle, fontsize=12, color=text_color_primary_label, fontweight="bold")
    lc_fit.set_ylabel("Intensity", fontsize=12)
    lc_fit.set_xlabel("Phase", fontsize=12)
    
    plot_box=FigureCanvasTkAgg(lc_fig, lc_fitting)
    plot_box.draw()
    plot_box.get_tk_widget().pack()
    
    toolbar=NavigationToolbar2Tk(plot_box, lc_fitting)
    toolbar.update()
    plot_box.get_tk_widget().pack()
        
    progress()
    main_window.update()
    lcp_window.update()

def fixed_parameter_fit():
    graphtitle="Fit with more precise Parameters"
    
    if len(preferences_list)==12:
        graphtitle=preferences_list[0]
        data_bg=preferences_list[1]
        fit_color=preferences_list[2]
        main_bg=preferences_list[3]
        tetriary_bg=preferences_list[4]
        text_color_primary_label=preferences_list[5]
        text_color_primary_axis=preferences_list[6]
        sys_vel_color=preferences_list[7]
        err_bar_color=preferences_list[8]
        data_style=preferences_list[10]
        line_style=preferences_list[11]
        
        if dl_button['text']=="Light Mode":
            lcpmain_bg="#0B0B0B"
            lcpsecondary_bg="#252525"
            text_color_primary="white"
            lcpbutton_bg="#f0e130"
            
        elif dl_button['text']=="Dark Mode":
            lcpmain_bg="#ECDBDB"
            lcpsecondary_bg="#FFFFFF"
            text_color_primary="black"
            lcpbutton_bg="#C9C6C6"
    
    elif dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        data_bg="red"
        fit_color="blue"
        data_style="."
        line_style="-"
        
        button_bg="#f0e130"
        lcpbutton_bg="#f0e130"
        
        text_color_primary_label="white"
        text_color_primary_axis="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
        lcpmain_bg="#0B0B0B"
        lcpsecondary_bg="#252525"
        text_color_primary="white"
        
    elif dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        data_bg="red"
        fit_color="blue"
        data_style="."
        line_style="-"
    
        button_bg="#C9C6C6"
        lcpbutton_bg="#C9C6C6"
        
        text_color_primary_label="black"
        text_color_primary_axis="black"
        text_color_secondary="black"
    
        color_frame_bg="#0C0B0B"
        
        lcpmain_bg="#ECDBDB"
        lcpsecondary_bg="#FFFFFF"
        text_color_primary="black"
    
    #Define Progress
    def progress_stage():
        return f"Current Progress: %{lcprogressbar['value']}"
    
    def progress():
        if lcprogressbar['value']<100:
            lcprogressbar['value']+=10
            lcprogresslabel['text']=progress_stage()
        
        if lcprogressbar['value']==100:
            progressfinish=tk.messagebox.showinfo("Successful", "The Light Curve Modelling is finished!")
            lcp_window.destroy()
    
    #Define Pop-up Progress Bar
    lcp_window=tk.Toplevel(master=main_window)
    lcp_window.title("Light Curve Modelling Progress")
    lcp_window.iconbitmap(r'icon_lcf.ico')
    lcp_window.configure(bg=lcpmain_bg)
    lcp_window.wm_transient(main_window)
    lcp_window.resizable(width=False, height=False)
    
    window_width=lcp_window.winfo_reqwidth()
    window_height=lcp_window.winfo_reqheight()

    screen_width=lcp_window.winfo_screenwidth()
    screen_height=lcp_window.winfo_screenheight()

    x_cordinate = int((screen_width/2) - (window_width/2))
    y_cordinate = int((screen_height/2) - (window_height/2))
    
    lcp_window.geometry("+{}+{}".format(x_cordinate, y_cordinate))
    
    lcinprogress=tk.Label(master=lcp_window, text="Light Curve Modelling in Progress...\n(Progress may take up to a few minutes)", bg=lcpsecondary_bg, fg=text_color_primary)
    lcinprogress.grid(row=0, column=0, pady=10, padx=10)
    
    lcprogressbar=ttk.Progressbar(master=lcp_window, orient="horizontal", mode="determinate", length=300)
    lcprogressbar.grid(row=1, column=0, pady=10, padx=10)
    
    lcprogresslabel=tk.Label(master=lcp_window, text=progress_stage(), bg=lcpbutton_bg)
    lcprogresslabel.grid(row=2, column=0, pady=10, padx=10)
    
    #Define progressbar value
    lcprogressbar['value']=0
    
    model_parameter_values_list=planet_parameter_values_list+star_parameter_values_list
    
    if len(lc_time)==0:
        inputfit_error=tk.messagebox.showerror("Error", "Please enter the Light Curve Data to proceed.")
        lcp_window.destroy()
    
    elif (len(star_parameter_values_list)!=3):
        inputfit_error=tk.messagebox.showerror("Error", "Please enter all the required Star Parameters to proceed.")
        lcp_window.destroy()
    
    elif (len(planet_parameter_values_list)!=6):
        inputfit_error=tk.messagebox.showerror("Error", "Please enter all the required Planet Parameters to proceed.")
        lcp_window.destroy()
    
    T0=planet_parameter_values_list[0]
    Period=planet_parameter_values_list[1]
    Eccentricity=planet_parameter_values_list[2]
    Argument_Of_Periapsis=planet_parameter_values_list[3]
    inclination=planet_parameter_values_list[4]
    r_planet=planet_parameter_values_list[5]
    
    r_star=star_parameter_values_list[0]
    ua=star_parameter_values_list[1]
    ub=star_parameter_values_list[2]
    
    progress()
    main_window.update()
    lcp_window.update()
    
    def Delta(lc_time, T0, Period, Eccentricity, Argument_Of_Periapsis, inclination, r_star, r_planet):
        Epoch_List, Delta_values_List, Delta_phase_List, c_values_List=[], [], [], []
        
        # if len(Epoch_List)!=0:
        #     Epoch_List.clear()
        #     Delta_values_List.clear() 
        #     Delta_phase_List.clear() 
        #     c_values_List.clear()
            
        for time in lc_time:
            #Epoch değerleri elde edilir.
            if ((float(time)-T0)/Period)<0:
                Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period)-1)
                Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)-1))*360
    
            else:
                Epoch_value=(float(time)-T0)/Period-(int((float(time)-T0)/Period))
                Epoch_deg_value=((float(time)-T0)/Period-(int((float(time)-T0)/Period)))*360
        
            Epoch_List.append(Epoch_value)
    
            #T-T0 değeri hesaplanır.
            T_T0=time-T0
            T_T0_values.append(T_T0)
    
            #Delta Değerleri Yörünge Basıklığı değerine göre hesaplanır.
            if Eccentricity>0:
                Mean_Anomaly=((2*math.pi)/Period)*(T_T0)
                Mean_Anomaly_values.append((Mean_Anomaly)) #M değerleri Radyan Biriminde
                
                Ei_old=Mean_Anomaly #Radyan Biriminde
                Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old)) #Radyan Biriminde
                error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                while error>0.001:
                    Ei_new=Mean_Anomaly+(Eccentricity*math.sin(Ei_old))
                    error=abs(math.degrees(Ei_new)-math.degrees(Ei_old))
                    Ei_old=Ei_new
        
                Eccentric_Anomaly_values.append(Ei_new) #E değerleri Radyan Biriminde
                
                True_Anomaly_0=((math.pi/2)-Argument_Of_Periapsis) #ν değerleri Radyan Biriminde
            
                True_Anomaly=2*math.atan(math.sqrt((1+Eccentricity)/(1-Eccentricity))*math.tan((Ei_new)/2))
                True_Anomaly_values.append(True_Anomaly) #ν değerleri Radyan Biriminde
                
                rho=(1-pow(Eccentricity,2))/(1+Eccentricity*math.cos(True_Anomaly))
                Rho_values.append(rho)
       
                delta=rho*math.sqrt((pow(math.sin(True_Anomaly-True_Anomaly_0),2)*pow(math.sin(inclination),2))+pow(math.cos(inclination),2))
                #δ değerleri Radyan Biriminde
                Delta_values_List.append(delta)
       
            elif Eccentricity==0:
                delta=math.sqrt((pow(math.sin(math.radians(Epoch_deg_value)),2)*pow(math.sin(inclination),2))+pow(math.cos(inclination),2))
                #δ değerleri Radyan Biriminde
                Delta_values_List.append(delta)
                
            #Epoch değerleri de belirlenir.
            if Epoch_value<0.2:
                Delta_phase=Epoch_value+1
        
            else:
                Delta_phase=Epoch_value
        
            Delta_phase_List.append(Delta_phase)
        
            #c değerleri de hesaplanır.
            c=delta/(r_star+r_planet)
            if c>1:
                c=1
                
            c_values_List.append(c)
        
        return Epoch_List, Delta_values_List, Delta_phase_List, c_values_List

    def a_and_b_calculation(r_star, r_planet):
        a=r_star/(r_planet+r_star)
        b=1-a
    
        return a, b
    
    def C_constants_calculation(ua, ub):
        u1=ua+(2*ub)
        u2=-ub
    
        C0=(3-(3*u1))/(3-u1)
        C1=(3*u1)/(3-u1)
        C2=(6*u2)/(6-(2*u1)-(3*u2))
    
        return C0, C1, C2
    
    progress()
    main_window.update()
    lcp_window.update()
    
    def Jacobi_Polynomial(h,alpha,beta,z,x):
        Jacobi_Polynomials=list()
    
        Lambda=alpha+beta+1 #İşlem kolaylığı açısından tanımlanan sabit
    
        for n in range(0,x+1):
            if n==0:
                Jacobi=1
                Jacobi_Polynomials.append(Jacobi)
                continue
        
            if n==1:
                Jacobi=-(beta+1)+(Lambda+1)*z
                Jacobi_Polynomials.append(Jacobi)
                continue
        
            if n==2:
                Jacobi=(beta+1)*(beta+2)/2-(beta+2)*(Lambda+2)*z+(Lambda+2)*(Lambda+3)*pow(z,2)/2
                Jacobi_Polynomials.append(Jacobi)
                continue
    
            if n>2:
                for n in range(2,x):
                    Constant_1=(2*n+Lambda)*(2*n+Lambda+1)/((n+1)*(n+Lambda))
                    Constant_2=((2*n+Lambda)*(pow(alpha,2)-pow(beta,2)+1)-pow((2*n+Lambda),3))/(2*(n+1)*(n+Lambda)*(2*n+Lambda-1))
                    Constant_3=((n+alpha)*(n+beta)*(2*n+Lambda+1))/((n+1)*(n+Lambda)*(2*n+Lambda-1))
                    Jacobi=(Constant_1*z+Constant_2)*Jacobi_Polynomials[n]-Constant_3*Jacobi_Polynomials[n-1]
                    Jacobi_Polynomials.append(Jacobi)
                    continue
        
            return Jacobi_Polynomials
    
    progress()
    main_window.update()
    lcp_window.update()
    
    def Fractional_Alpha_Function(h, a, b, c):
        #Nu ifadesi tanımlanır. (Her alfa fonksiyonu için ayrı değere sahiptir!)
        nu=(h+2)/2
        
        #Toplam Sembollerinden önce bulunan 1. ifadenin hesabı yapılır.
        alpha_before_sum=pow(b,2)*math.gamma(nu)*pow((1-pow(c,2)),(nu+1))
        
        #Toplam Sembolü içinde bulunan ifadelerin hesabı için de makalede bulunan denklem kullanılacaktır.
        #Önce Jacobi Polinomlarımızın her ikisini de her bir yapılacak toplam için elde etmemiz gerekmektedir.
        #Sonrasında bunu Toplam Sembolündeki ilk kısım ile çarparak sonucu elde edebiliriz.
        Jacobi_1=Jacobi_Polynomial(h, 1, nu, a, 80)
        Jacobi_2=Jacobi_Polynomial(h, 1+nu, 0, pow(c,2), 80)
    
        sum_total=0
        for n in range(0,81):
            sum_total+=(((math.factorial(n)*(nu+2*n+2))/((n+1)*math.gamma(nu+n+1)))*pow(Jacobi_1[n],2)*Jacobi_2[n])
    
        #İncelenen gözlem zamanı için Küçük Alfa Fonksiyonlarından biri elde edilir.
        alpha=alpha_before_sum*sum_total
        return alpha
    
    progress()
    main_window.update()
    lcp_window.update()
    
    #Formüllerde iterasyon 0'dan sonsuza yapılmakta.
    #Kod kapsamında iterasyonumuz 0'dan 80'e kadar yapılacaktır.
    def Alpha_Function(lc_time, T0, Period, r_star, r_planet, Eccentricity, Argument_Of_Periapsis, inclination, ua, ub):
        Eclipse_Functions, Light_Curve_Fit=[], []
        
        #Verilen lc_time için Epoch, Delta ve Phase değeri elde edilir.
        Epoch_Fit_List, Delta_values_Fit_List, Delta_phase_Fit_List, c_values_Fit_List = Delta(lc_time, T0, Period, Eccentricity, Argument_Of_Periapsis, inclination, r_star, r_planet)
    
        #a ve b parametreleri hesaplanır.
        #Gözlem zamanı ile değişmediğinden döngüye gerek kalmadan hesap yapılabilir.
        a, b = a_and_b_calculation(r_star, r_planet)
    
        #Alfa Fonksiyonu hesabında gerekli olan C0, C1 ve C2 değerleri hesaplanır.
        #Gözlem zamanı ile değişmediğinden döngüye gerek kalmadan hesap yapılabilir.
        C0, C1, C2 = C_constants_calculation(ua, ub)
    
        #lc_time zamanı için Kesirsel Alfa Fonksiyonları hesaplanır.
        for c_value in c_values_Fit_List:
            frac_alpha_0=Fractional_Alpha_Function(0, a, b, c_value) #alpha_0
            frac_alpha_1=Fractional_Alpha_Function(1, a, b, c_value) #alpha_1
            frac_alpha_2=Fractional_Alpha_Function(2, a, b, c_value) #alpha_2
    
            #lc_time zamanı için Tutulma Fonksiyonunun (Alfa Fonksiyonunun) değeri hesaplanır.
            Eclipse_Function=(C0*frac_alpha_0)+(C1*frac_alpha_1)+(C2*frac_alpha_2)
            Eclipse_Functions.append(Eclipse_Function)
            Light_Curve_Fit.append(1-Eclipse_Function)
        
        return Light_Curve_Fit
    
    from lmfit import Model, minimize, Parameters, report_fit, fit_report
    
    Parameters_Fit=Parameters()
    
    progress()
    main_window.update()
    lcp_window.update()
    
    #Modeli oluşturuyoruz.
    Alpha_Model=Model(Alpha_Function)

    #Oluşturduğumuz Modeldeki Değişkenleri ve Parametreleri print ediyoruz.
    print(f"Parameters: {Alpha_Model.param_names}")
    print(f"Variables: {Alpha_Model.independent_vars}")

    #Şimdi de bu Parametreleri sınırlayabilmek için ve ilk değerlerini koyabilmek için tanımlıyoruz.
    Parameters=Alpha_Model.make_params()
    
    Parameters.add("T0", value=T0, min=0.000001, max=np.inf, vary=T0_vary.get())
    Parameters.add("Period", value=Period, min=0.000001, max=np.inf, vary=Period_vary.get())
    Parameters.add("Eccentricity", value=Eccentricity, min=0, max=0.99999, vary=Eccentricity_vary.get())
    Parameters.add("Argument_Of_Periapsis", value=Argument_Of_Periapsis, min=0, max=2*math.pi, vary=Argument_Of_Periapsis_vary.get())
    Parameters.add("inclination", value=inclination, min=0.87, max=math.pi/2, vary=inclination_vary.get()) #En küçük eğim 50deg alındı. Değiştirilebilir.
    Parameters.add("r_planet", value=r_planet, min=0, max=0.5, vary=r_planet_vary.get())
    Parameters.add("r_star", value=r_star, min=0, max=0.8, vary=r_star_vary.get())
    Parameters.add("ua", value=ua, min=0, max=1, vary=ua_vary.get())
    Parameters.add("ub", value=ub, min=0, max=1, vary=ub_vary.get())
    
    progress()
    main_window.update()
    lcp_window.update()
    
    #Fit yapılır.
    fit_result=Alpha_Model.fit(lc_flux, Parameters, lc_time=lc_time)

    #Sonuç yazılır.
    fit_values=[]
    fit_errors=[]
    fit_corrs=[]
    
    fit_values.append(fit_result.params["T0"].value)
    fit_values.append(fit_result.params["Period"].value)
    fit_values.append(fit_result.params["Eccentricity"].value)
    fit_values.append(fit_result.params["Argument_Of_Periapsis"].value)
    fit_values.append(fit_result.params["inclination"].value)
    fit_values.append(fit_result.params["r_planet"].value)
    fit_values.append(fit_result.params["r_star"].value)
    fit_values.append(fit_result.params["ua"].value)
    fit_values.append(fit_result.params["ub"].value)
    
    fit_errors.append(fit_result.params["T0"].stderr)
    fit_errors.append(fit_result.params["Period"].stderr)
    fit_errors.append(fit_result.params["Eccentricity"].stderr)
    fit_errors.append(fit_result.params["Argument_Of_Periapsis"].stderr)
    fit_errors.append(fit_result.params["inclination"].stderr)
    fit_errors.append(fit_result.params["r_planet"].stderr)
    fit_errors.append(fit_result.params["r_star"].stderr)
    fit_errors.append(fit_result.params["ua"].stderr)
    fit_errors.append(fit_result.params["ub"].stderr)
    
    fit_corrs.append(fit_result.params["T0"].correl)
    fit_corrs.append(fit_result.params["Period"].correl)
    fit_corrs.append(fit_result.params["Eccentricity"].correl)
    fit_corrs.append(fit_result.params["Argument_Of_Periapsis"].correl)
    fit_corrs.append(fit_result.params["inclination"].correl)
    fit_corrs.append(fit_result.params["r_planet"].correl)
    fit_corrs.append(fit_result.params["r_star"].correl)
    fit_corrs.append(fit_result.params["ua"].correl)
    fit_corrs.append(fit_result.params["ub"].correl)

    #Düzeltilen Fit değerleri bir parametre olarak tanımlanır.
    Fit_finalized=fit_result.best_fit
    
    progress()
    main_window.update()
    lcp_window.update()
    
    #Gerekli olan Phase değerleri listeye çekilir.
    Epoch_List, Delta_values_List, Delta_phase_List, c_values_List=Delta(lc_time, T0, Period, Eccentricity, Argument_Of_Periapsis, inclination, r_star, r_planet)

    #Fit değerleri bir listeye çekilir.
    Light_Curve_Fit=Alpha_Function(lc_time, T0, Period, r_star, r_planet, Eccentricity, Argument_Of_Periapsis, inclination, ua, ub)

    #Bazen alınan datalardaki zaman değerleri farklı zaman araklıklarında olabilir.
    #Örneğin 360.15 zamanından sonra 630.15, sonrasında da tekrardan 361.15 zamanlarında gözlemler alınmış olup bize datada öyle gelebilir.
    #Bu nedenle Fit değerlerini, Akı değerlerini ve Evre Değerlerini minimumdan maksimuma sıralamamız faydalı olacaktır.
    #Bu sayede fit çizgimizde bir problem oluşması önlenecektir.
    
    Delta_phase_sort_list=Delta_phase_List.copy()
    lc_flux_sort_list=lc_flux.copy()
    Light_Curve_Fit_sort_list=Light_Curve_Fit.copy()
    Fit_finalized_sort_list=Fit_finalized.copy()
    
    Delta_phase_sorted, lc_flux_sorted, Light_Curve_Fit_sorted, Fit_finalized_sorted = [], [], [], []
    
    for i in range(len(Delta_phase_List)):
        min_val=min(Delta_phase_sort_list) #Minimum zaman
        min_index=Delta_phase_sort_list.index(min_val) #Minimum zamanın index değeri
    
        Delta_phase_sorted.append(min_val)
        Delta_phase_sort_list.pop(min_index)
    
        lc_flux_sorted.append(lc_flux_sort_list[min_index])
        lc_flux_sort_list.pop(min_index)
    
        Light_Curve_Fit_sorted.append(Light_Curve_Fit_sort_list[min_index])
        Light_Curve_Fit_sort_list.pop(min_index)
    
        Fit_finalized_sorted.append(Fit_finalized_sort_list[min_index])
        Fit_finalized_sort_list.pop(min_index)
    
    progress()
    main_window.update()
    lcp_window.update()
    
    for i in range(len(model_parameter_names_list)):
        parameter_value=tk.Entry(master=frame_middle_right)
        parameter_value.insert(tk.END, model_parameter_values_list[i])
        parameter_value.grid(row=2+i, column=1, sticky="nsew", padx=10)
        parameter_value.configure(state="readonly")
    
    for i in range(len(model_parameter_names_list)):
        fit_value=tk.Entry(master=frame_middle_right)
        fit_value.insert(tk.END, round(fit_values[i], model_parameter_round_list[i]))
        fit_value.grid(row=2+i, column=2, sticky="nsew", padx=10)
        fit_value.configure(state="readonly")
    
    for i in range(len(model_parameter_names_list)):
        fit_error=tk.Entry(master=frame_middle_right)
        if fit_errors[i] is not None:
                fit_error.insert(tk.END, f"\u00B1 {round(fit_errors[i], model_parameter_round_list[i])}")
            
        else:
            fit_error.insert(tk.END, f"\u00B1 {fit_errors[i]}")
            
        fit_error.grid(row=2+i, column=3, sticky="nsew", padx=10)
        fit_error.configure(state="readonly")
    
    progress()
    main_window.update()
    lcp_window.update()
    
    model_parameters_update.model_parameter_fit_value_list=fit_values.copy()
    model_parameters_update.model_parameter_error_value_list=fit_errors.copy()
    
    for widgets in lc_fitting.winfo_children():
        widgets.destroy()
    
    lc_fig=Figure(figsize=(7,4), dpi=90)
    lc_fig.patch.set_facecolor(main_bg)
    
    lc_fit=lc_fig.add_subplot(111)
    lc_fit.plot(Delta_phase_sorted, lc_flux_sorted, data_style, color=data_bg)
    lc_fit.plot(Delta_phase_sorted, Fit_finalized_sorted, line_style, color=fit_color, label="Fit")
    lc_fit.patch.set_facecolor(tetriary_bg)
    
    lc_fit.xaxis.label.set_color(text_color_primary_label)
    lc_fit.yaxis.label.set_color(text_color_primary_label)
    
    lc_fit.tick_params(axis='x', colors=text_color_primary_axis)
    lc_fit.tick_params(axis='y', colors=text_color_primary_axis)
    
    lc_fit.spines['left'].set_color(text_color_primary_axis)
    lc_fit.spines['bottom'].set_color(text_color_primary_axis)
    lc_fit.spines['right'].set_color(text_color_primary_axis)
    lc_fit.spines['top'].set_color(text_color_primary_axis)
    
    lc_fit.set_title(graphtitle, fontsize=12, color=text_color_primary_label, fontweight="bold")
    lc_fit.set_ylabel("Intensity", fontsize=12)
    lc_fit.set_xlabel("Phase", fontsize=12)
    
    plot_box=FigureCanvasTkAgg(lc_fig, lc_fitting)
    plot_box.draw()
    plot_box.get_tk_widget().pack()
    
    toolbar=NavigationToolbar2Tk(plot_box, lc_fitting)
    toolbar.update()
    plot_box.get_tk_widget().pack()
    
    progress()
    main_window.update()
    lcp_window.update()
    
fit_button=tk.Button(master=frame_top_right, text="Press to calculate Model Light Curve", command=input_parameter_fit, bg=button_bg)
finalized_fit_button=tk.Button(master=frame_top_right, text="Press to fit for precise Model Parameters", command=fixed_parameter_fit, bg=button_bg)

fit_button.grid(row=3, column=1, sticky="nsew", padx=10, pady=5)
finalized_fit_button.grid(row=4, column=1, sticky="nsew", padx=10, pady=5)

#Middle Right Frame -> Results and Errors
fit_results_label=tk.Label(master=frame_middle_right, text="Fit Results", font=("Times New Roman", 10, "bold"), bg=title_bg)
parameter_result_label=tk.Label(master=frame_middle_right, text="Parameter Name", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
parameter_name_fit_label=tk.Label(master=frame_middle_right, text="Original Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
parameter_value_fit_label=tk.Label(master=frame_middle_right, text="Fitted Value", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)
error_fit_label=tk.Label(master=frame_middle_right, text="Error", font=("Times New Roman", 10, "bold"), bg=main_bg, fg=text_color_primary)

fit_results_label.grid(row=0, column=0, columnspan=4, sticky="nsew", padx=10, pady=(10,5))
parameter_result_label.grid(row=1, column=0, sticky="nsew", padx=100, pady=5)
parameter_name_fit_label.grid(row=1, column=1, sticky="nsew", padx=100, pady=5)
parameter_value_fit_label.grid(row=1, column=2, sticky="nsew", padx=100, pady=5)
error_fit_label.grid(row=1, column=3, sticky="nsew", padx=100, pady=5)

for i in range(len(model_parameter_names_list)):
    parametername=tk.Label(master=frame_middle_right, text=model_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
    parametername.grid(row=2+i, column=0, sticky="nsew", padx=10)

#Bottom Right Frame -> Update Parameters Button
def update_parameters():
    if dl_button['text']=="Light Mode":
        main_bg="#0B0B0B"
        secondary_bg="#252525"
        tetriary_bg="#3A3A3A"
        title_bg="#FF7A37"
        
        button_bg="#f0e130"
        
        text_color_primary="white"
        text_color_secondary="black"
        
        color_frame_bg="#0C0B0B"
        
    if dl_button['text']=="Dark Mode":
        main_bg="#ECDBDB"
        secondary_bg="#FFFFFF"
        tetriary_bg="#FFFFFF"
        title_bg="#7A7A7A"
        
        button_bg="#C9C6C6"
        
        text_color_primary="black"
        text_color_secondary="black"
            
        color_frame_bg="#0C0B0B"

    try:
        if len(model_parameters_update.model_parameter_fit_value_list)==9:
            star_parameter_values_list.clear()
            planet_parameter_values_list.clear()
            for i in range(len(model_parameters_update.model_parameter_fit_value_list)):
                if i<6:
                    planet_parameter_values_list.append(round(model_parameters_update.model_parameter_fit_value_list[i], model_parameter_round_list[i]))
                
                else:
                    star_parameter_values_list.append(round(model_parameters_update.model_parameter_fit_value_list[i], model_parameter_round_list[i]))
            
            for i in range(len(planet_parameter_values_list)):
                if len(planet_parameter_values_list)!=0:
                    parametervalue_first=tk.Label(master=frame_middle_left, text=planet_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
                    parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
            
                elif len(planet_parameter_values_list)==0:
                    parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
                    parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
            
            for i in range(len(star_parameter_values_list)):
                if len(star_parameter_values_list)!=0:
                    parametervalue_first=tk.Label(master=frame_middle_left, text=star_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
                    parametervalue_first.grid(row=8+i, column=1, sticky="nsew", padx=10)
            
                elif len(star_parameter_values_list)==0:
                    parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
                    parametervalue_first.grid(row=8+i, column=1, sticky="nsew", padx=10)
            
            show_success=tk.messagebox.showinfo("Successful", "Parameters have been updated!")
    
    except AttributeError:
        show_error=tk.messagebox.showerror("Error", "There aren't any parameters to update.\n\nPlease do the Light Curve Modelling first in order to update the parameters.")

def lcexport_func():
    if len(planet_parameter_values_list)!=6 and len(star_parameter_values_list)!=3:
        show_error=tk.messagebox.showerror("Error", "There aren't any parameters to export.\n\nPlease do the Light Curve Modelling first in order to export the parameters.")

    elif len(planet_parameter_values_list)==6 and len(star_parameter_values_list)==3:
        lcexport.lcexportplanet_parameter_values_list=planet_parameter_values_list.copy()
        lcexport.lcexportplanet_parameter_values_list.pop(0)
        
        lcexport.lcexportstar_parameter_values_list=star_parameter_values_list.copy()
        
        if len(rvplanet_parameter_values_list)==6:
           for i in range(len(lcexport.lcexportplanet_parameter_values_list)):
               rvplanet_parameter_values_list[i+1]=lcexport.lcexportplanet_parameter_values_list[i]
        
        else:
            rvplanet_parameter_values_list.append(0)
            for i in range(len(lcexport.lcexportplanet_parameter_values_list)):
                rvplanet_parameter_values_list.insert(i+1, lcexport.lcexportplanet_parameter_values_list[i])
        
        if len(rvstar_parameter_values_list)==4:
            star_parameter_values_list[2]=lcexport.lcexportstar_parameter_values_list[0]
        
        else:
            rvstar_parameter_values_list.clear()
            rvstar_parameter_values_list.append(0)
            rvstar_parameter_values_list.append(0)
            rvstar_parameter_values_list.append(lcexport.lcexportstar_parameter_values_list[0])

        show_success=tk.messagebox.showinfo("Successful", "Parameters have been exported to Radial Velocity analysis!\n\nNote: Due to Light Curve T0 being the minimum of the transit and Radial Velocity T0 being passage of periapsis time, it is not exported.")

update_parameters=tk.Button(master=frame_bottom_right, text="Press to Update the Parameter Values", font=("Times New Roman", 10, "bold"), command=update_parameters, bg=button_bg)
update_parameters.grid(row=0, column=0, pady=5)

lcexport_parameters=tk.Button(master=frame_bottom_right, text="Press to Export the Parameter Values to Radial Velocity", font=("Times New Roman", 10, "bold"), command=lcexport_func, bg=button_bg)
lcexport_parameters.grid(row=1, column=0, pady=5, padx=5)

#Bottom Left Frame --> "How to use?" button + Dark Mode Light Mode Button + Preferences Button
def how_to_use():
    webbrowser.open("Light Curve Modelling (Manual).html")
   
htu_button=tk.Button(master=frame_bottom_left, text="How to use?", command=how_to_use).grid(row=1, column=0, columnspan=2, pady=5)

#Bottom Left Frame --> Dark Light Mode Switch + Button
def dark_mode():
    main_bg="#0B0B0B"
    secondary_bg="#252525"
    tetriary_bg="#3A3A3A"
    title_bg="#FF7A37"
    
    button_bg="#f0e130"
    
    text_color="white"
    color_frame_bg="#0C0B0B"
    
    main_window.configure(bg=main_bg)
    
    frame_top_left.configure(bg=secondary_bg)
    frame_middle_left.configure(bg=secondary_bg)
    frame_bottom_left.configure(bg=secondary_bg)
    frame_top_right.configure(bg=secondary_bg)
    frame_middle_right.configure(bg=secondary_bg)
    frame_bottom_right.configure(bg=secondary_bg)
    
    entry_label.configure(bg=title_bg)
    open_button.configure(bg=button_bg)
    lc_button.configure(bg=button_bg)
    parametersofplanet_button.configure(bg=button_bg)
    parametersofstar_button.configure(bg=button_bg)
    rv_analysis_button.configure(bg=button_bg)
    all_parameters_button.configure(bg=button_bg)
    
    fit_results_label.configure(bg=title_bg)
    parameter_label.configure(bg=title_bg)
    parameter_name_label.configure(bg=main_bg, fg=text_color_primary)
    parameter_value_label.configure(bg=main_bg, fg=text_color_primary)
    variability_label.configure(bg=main_bg, fg=text_color_primary)
    
    lc_fitting.configure(bg=main_bg)
    
    for i in range(len(planet_parameter_names_list)):
        parametername=tk.Label(master=frame_middle_left, text=planet_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=2+i, column=0, sticky="nsew", padx=10)
        
        if len(planet_parameter_values_list)!=0:
           parametervalue_first=tk.Label(master=frame_middle_left, text=planet_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
           parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
        
        elif len(planet_parameter_values_list)==0:
            parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
        
        parametername=tk.Label(master=frame_middle_right, text=planet_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=2+i, column=0, sticky="nsew", padx=10)
    
    for i in range(len(star_parameter_names_list)):
        parametername=tk.Label(master=frame_middle_left, text=star_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=8+i, column=0, sticky="nsew", padx=10)
        
        if len(star_parameter_values_list)!=0:
           parametervalue_first=tk.Label(master=frame_middle_left, text=star_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
           parametervalue_first.grid(row=8+i, column=1, sticky="nsew", padx=10)
        
        elif len(star_parameter_values_list)==0:
            parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=8+i, column=1, sticky="nsew", padx=10)
        
        parametername=tk.Label(master=frame_middle_right, text=star_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=8+i, column=0, sticky="nsew", padx=10)
    
    T0_vary_check.configure(bg=secondary_bg)
    Period_vary_check.configure(bg=secondary_bg)
    Eccentricity_vary_check.configure(bg=secondary_bg)
    Argument_Of_Periapsis_vary_check.configure(bg=secondary_bg)
    inclination_vary_check.configure(bg=secondary_bg)
    r_planet_vary_check.configure(bg=secondary_bg)
    
    r_star_vary_check.configure(bg=secondary_bg)
    ua_vary_check.configure(bg=secondary_bg)
    ub_vary_check.configure(bg=secondary_bg)
    
    fit_text.configure(bg=title_bg)
    lcphase_button.configure(bg=button_bg)
    fit_button.configure(bg=button_bg)
    finalized_fit_button.configure(bg=button_bg)
    
    parameter_result_label.configure(bg=main_bg, fg=text_color_primary)
    parameter_name_fit_label.configure(bg=main_bg, fg=text_color_primary)
    parameter_value_fit_label.configure(bg=main_bg, fg=text_color_primary)
    error_fit_label.configure(bg=main_bg, fg=text_color_primary)
    
    update_parameters.configure(bg=button_bg)
    lcexport_parameters.configure(bg=button_bg)
    
    preferences_button.configure(bg=button_bg, fg="black")
    
    dl_button.configure(bg=button_bg, fg="black", text="Light Mode")
    dl_button.configure(command=light_mode)
    
def light_mode():
    main_bg="#ECDBDB"
    secondary_bg="#FFFFFF"
    tetriary_bg="#3A3A3A"
    title_bg="#7A7A7A"
    
    text_color_primary="black"
    button_bg="#C9C6C6"
    
    color_frame_bg="#0C0B0B"
    
    main_window.configure(bg=main_bg)
    
    frame_top_left.configure(bg=secondary_bg)
    frame_middle_left.configure(bg=secondary_bg)
    frame_bottom_left.configure(bg=secondary_bg)
    frame_top_right.configure(bg=secondary_bg)
    frame_middle_right.configure(bg=secondary_bg)
    frame_bottom_right.configure(bg=secondary_bg)
    
    entry_label.configure(bg=title_bg)
    open_button.configure(bg=button_bg)
    lc_button.configure(bg=button_bg)
    parametersofplanet_button.configure(bg=button_bg)
    parametersofstar_button.configure(bg=button_bg)
    rv_analysis_button.configure(bg=button_bg)
    all_parameters_button.configure(bg=button_bg)
    
    parameter_label.configure(bg=title_bg)
    parameter_name_label.configure(bg=main_bg, fg=text_color_primary)
    parameter_value_label.configure(bg=main_bg, fg=text_color_primary)
    variability_label.configure(bg=main_bg, fg=text_color_primary)
    
    lc_fitting.configure(bg=main_bg)
    
    for i in range(len(planet_parameter_names_list)):
        parametername=tk.Label(master=frame_middle_left, text=planet_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=2+i, column=0, sticky="nsew", padx=10)
        
        if len(planet_parameter_values_list)!=0:
           parametervalue_first=tk.Label(master=frame_middle_left, text=planet_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
           parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
        
        elif len(planet_parameter_values_list)==0:
            parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=2+i, column=1, sticky="nsew", padx=10)
        
        parametername=tk.Label(master=frame_middle_right, text=planet_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=2+i, column=0, sticky="nsew", padx=10)
    
    for i in range(len(star_parameter_names_list)):
        parametername=tk.Label(master=frame_middle_left, text=star_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=8+i, column=0, sticky="nsew", padx=10)
        
        if len(star_parameter_values_list)!=0:
           parametervalue_first=tk.Label(master=frame_middle_left, text=star_parameter_values_list[i], bg=secondary_bg, fg=text_color_primary)
           parametervalue_first.grid(row=8+i, column=1, sticky="nsew", padx=10)
        
        elif len(star_parameter_values_list)==0:
            parametervalue_first=tk.Label(master=frame_middle_left, text="---", bg=secondary_bg, fg=text_color_primary)
            parametervalue_first.grid(row=8+i, column=1, sticky="nsew", padx=10)
        
        parametername=tk.Label(master=frame_middle_right, text=star_parameter_names_list[i], font=("Calibri", 10, "bold"), bg=secondary_bg, fg=text_color_primary)
        parametername.grid(row=8+i, column=0, sticky="nsew", padx=10)
    
    T0_vary_check.configure(bg=secondary_bg)
    Period_vary_check.configure(bg=secondary_bg)
    Eccentricity_vary_check.configure(bg=secondary_bg)
    Argument_Of_Periapsis_vary_check.configure(bg=secondary_bg)
    inclination_vary_check.configure(bg=secondary_bg)
    r_planet_vary_check.configure(bg=secondary_bg)
    
    r_star_vary_check.configure(bg=secondary_bg)
    ua_vary_check.configure(bg=secondary_bg)
    ub_vary_check.configure(bg=secondary_bg)
    
    fit_results_label.configure(bg=title_bg)
    fit_text.configure(bg=title_bg)
    lcphase_button.configure(bg=button_bg)
    fit_button.configure(bg=button_bg)
    finalized_fit_button.configure(bg=button_bg)
    
    parameter_result_label.configure(bg=main_bg, fg=text_color_primary)
    parameter_name_fit_label.configure(bg=main_bg, fg=text_color_primary)
    parameter_value_fit_label.configure(bg=main_bg, fg=text_color_primary)
    error_fit_label.configure(bg=main_bg, fg=text_color_primary)
    
    update_parameters.configure(bg=button_bg)
    lcexport_parameters.configure(bg=button_bg)
    
    preferences_button.configure(bg=button_bg, fg="black")
    
    dl_button.configure(bg=button_bg, fg="black", text="Dark Mode")
    dl_button.configure(command=dark_mode)

dl_button=tk.Button(master=frame_bottom_left, text="Light Mode", command=light_mode, bg=button_bg, font=("Times New Roman", 10, "bold"))
dl_button.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

# splash_window.after(2000, main_window_init)

tk.mainloop()